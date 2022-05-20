import csv
from curses.ascii import isdigit
import glob
import re
import pandas as pd
from unidecode import unidecode
import openpyxl
from openpyxl import Workbook
import openpyxl_dictreader
# coding: utf-8

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

nomProfil : str = "JSCD"
extensionDebut : str = ".xlsx"
extensionFin : str = ".xlsx"
j : int = 1

# On load le le tarif au format xlsx dans un workbook
for filename in glob.glob('*' + extensionDebut):
    
    reader = pd.read_excel(filename, header = 3)
    reader.to_excel(nomProfil + extensionFin, index = False, header=False)

# new_file = openpyxl_dictreader.DictReader('jscd.xlsx', "Sheet1")
new_file = openpyxl_dictreader.DictReader(nomProfil + extensionFin, "Sheet1")

# On prépare un nouveau workbook
wb2 = Workbook()
#dest_filename_bis = 'sortie_'+ 'bis_' + nomProfil.upper()+extensionFin
dest_filename = 'sortie_'+ nomProfil.upper()+extensionFin
ws = wb2.active
ws.title = "JSCD"

for row in new_file :
    region = row['Couleur']
    iAppellation = row['Apellation']
    millesime = row['Millesime']
    iChateau = row['Chateaux/Domaine']
    iFormatB = row['Format']
    iCond = row['Condi.']
    quantite = row['Quantité']
    prix : str = str(row['PrixHTEUR'])
    prix : str = ft.formaterPrix(prix);
    commentaires = row['Remarque']
    iCom_App = 0
    iCom_Cond = 0

    if iChateau is None :
        continue
    
    if "  " in iChateau :
        chateau = str(iChateau).replace('  ','')

    if region=='Bordeaux Blanc' or region=='Bordeaux Rouge':
        chateau=unidecode(iChateau)
    else:
        chateau=unidecode(iAppellation)+' '+unidecode(iChateau)
    
    if iChateau.find("Assortiment") != -1 :

        i = iChateau.find("Assortiment")
        chateau = iAppellation+' '+iChateau[:i+11]       
        conditionnement = 'COLLEC'
        iCom_App = 1

    elif re.findall('\d+[ ]', iChateau) and ',' in iChateau :

        chateau = iAppellation +' collection'
        conditionnement = 'COLLEC'
        iCom_App = 1

    elif ': ' in iChateau :

        i = iChateau.find(": ")
        chateau = iAppellation+' '+iChateau[:i]
        conditionnement = 'COLLEC'
        iCom_App = 1
    
    chateau = chateau.strip()

    if ( ( commentaires is None ) or  ( "pic" in commentaires ) ) :
        commentaires = ''

    [chateau, commentairesBis] = ft.testCommentaireParenthese(chateau)

    if( commentairesBis != "" ):
        commentaires = commentaires + " " + commentairesBis

    formatB = ft.formaterFormatBouteille(iFormatB)

    iCom_defaut : int = 0

    if iCond is None :

        conditionnement = ft.conditionnementParDefaut(formatB, quantite)
        iCom_defaut = 1

    else:
        conditionnement = str(iCond).replace('CB','CBO')

        if '+' in str(iCond) :
            conditionnement = conditionnement.split('+')[0]
            iCom_Cond = 1

    verif_cdt_str : str = 'verif cdt -'

    if iCom_defaut == 1 :
        commentaires = verif_cdt_str 

    if iCom_App == 1 :
        commentaires = commentaires + ' ' + iChateau

    if iCom_Cond == 1 :
        commentaires = commentaires + ' ' + conditionnement
    
    # ancien
    # if iCom_App == 1 :
    #     commentaires = iChateau+' '+commentaires

    #     if iCom_defaut == 1 :
    #         commentaires = 'verif cdt - '+commentaires 

    # elif iCom_Cond == 1 :
    #     commentaires = conditionnement+' '+commentaires

    #     if iCom_defaut == 1 :
    #         commentaires = 'verif cdt - '+commentaires 


    #On affecte les valeurs dans l'ordre des colonnes voulues.
    ft.affectationLignes(j, ws, chateau, millesime, formatB, prix, quantite, conditionnement, commentaires)
    j += 1

#Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)