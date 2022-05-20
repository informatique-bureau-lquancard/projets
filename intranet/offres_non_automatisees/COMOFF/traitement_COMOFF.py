# coding: utf-8
import csv
from email import header
import glob
from typing import OrderedDict
from numpy import False_
#import codecs
from unidecode import unidecode
from openpyxl import Workbook
import pandas as pd

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

nomProfil : str = "COMOFF"
extensionDebut : str = ".xlsx"
extensionFin : str = ".xlsx"

for filename in glob.glob('*' + extensionDebut):
    reader = pd.read_excel(filename)
    reader.dropna(how = "all", inplace = True)
    reader.to_csv('comoff.csv', index = False)
    reader = pd.read_csv('comoff.csv', header = 1)
    reader.to_csv('comoff.csv', index = False, header = False)

wb = Workbook()
dest_filename = 'sortie_' + nomProfil + extensionFin
ws = wb.active
ws.title = nomProfil

with open('comoff.csv', newline='') as csvfile :
    csvreader = csv.DictReader(csvfile)

    j :int = 1

    for row in csvreader :

        # for key, value in row.items() :
               
        vin = unidecode(row['Château']).upper()
        prix = ft.formaterPrix(str(row['€/unit']))

        # if 'quantite' in key :
        #     quantite = 0
        #     commentaire = row['quantite']
        # else:
        #     quantite = 0
        #     commentaire = ''

        designation_prix = ["€/unit"]

        if( ft.bLigneIncorrecte(prix, designation_prix, vin)):
            continue;

        millesime = row['Vintage']
    
        formatB = ft.formaterFormatBouteille(str(row['size ml']))

        quantite = ft.formaterQuantite(str(row['Quantity in bts ']))

        conditionnement = row['PACK'] 

        tab_cond_CB6 = ['6 CC', 'OW6', 'C06B', 'OCW6']

        if conditionnement in tab_cond_CB6:
            conditionnement = 'CBO6'

        if conditionnement == 'CT04':
            conditionnement = 'CC4'


        conditionnement = unidecode(str(conditionnement)).upper().replace('OWC','CBO').replace('PK CARDBOARD','CC').replace('LOOSE','UNITE').replace('6CC','CC6').replace('12CC','CC12')

        commentaire = ''

        tarif_officieux = 1

        prix = prix.replace("'", '')
        prix = prix.replace(",", '.')

        ft.affectationLignes2(j, ws, vin, millesime, formatB, float(prix), quantite, conditionnement, commentaire, tarif_officieux)

        j = j + 1

index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))       

wb.save(dest_filename)