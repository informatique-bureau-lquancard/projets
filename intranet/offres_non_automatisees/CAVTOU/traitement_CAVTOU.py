import csv
import glob
import re
import time
import datetime
from unidecode import unidecode
import openpyxl
from openpyxl import Workbook
import pandas as pd
# coding: utf-8

import codecs

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename = 'sortie_CAVTOU.xlsx'
    ws = wb2.active
    ws.title = "CAVTOU °‿‿° "
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    iMillesime_iRegion = sheet['B']
    iFormatB = sheet['C']
    iQte = sheet['D']
    iDomaine = sheet['E']
    iAppellation = sheet['F']
    iPrix = sheet['G']

    j = 1
    
 
    row_count = sheet.max_row
    column_count = sheet.max_column

    for index_ligne  in range (1, row_count) :
        region = ''
        if iPrix[index_ligne].value is None and iDomaine[index_ligne].value is None :
            continue
        elif str(iFormatB[index_ligne].value) == 'Format':
            continue
        elif iMillesime_iRegion[index_ligne].value is not None and iPrix[index_ligne].value is None :
            region = str(iMillesime_iRegion[index_ligne]).value
            #continue

        if 'BORDEAUX' in region:
            vin : str = unidecode(str(iAppellation[index_ligne].value).upper())
        else:
            vin : str = unidecode(str(iDomaine[index_ligne].value).upper()+' '+str(iAppellation[index_ligne].value).upper())
        
        millesime = ft.formaterAnnee(str(iMillesime_iRegion[index_ligne].value))

        formatB : str = ft.formaterFormatBouteille(iFormatB[index_ligne].value)

        prix = str(iPrix[index_ligne].value).replace('.',',')

        quantite = int(iQte[index_ligne].value)

        conditionnement = ft.conditionnementParDefaut(formatB, quantite)

        commentaire = 'verif cdt'


        #On affecte les valeurs dans l'ordre des colonnes voulues.
        ft.affectationLignes(j, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaire)
        j += 1
        

#Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)

# read_file = pd.read_excel(dest_filename)
# read_file.to_csv('sortie_CAVTOU.csv', index = False, encoding="utf-8", sep = ';')