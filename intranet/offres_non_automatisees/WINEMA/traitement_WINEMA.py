# coding: utf-8
import csv
import glob
import codecs
from unidecode import unidecode
from time import sleep
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook


# Conversion du fichier xls en xlsx
for xls_file in glob.glob('*.xls'):
    reader = pd.read_excel(xls_file)
    reader.to_excel('winemania.xlsx', index = False)

sleep(1)

# On load le le tarif au format xlsx dans un workbook
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_WINEMA.xlsx'
ws = wb2.active
ws.title = "WINEMA"

# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['D']
iAppellation = sheet['E']
iMillesime = sheet['A']
iFormatb = sheet['B']
iPrix = sheet['F']
iQte = sheet['C']
#iCond = sheet['M']
#iDate_offre = sheet['I']  #validité
#iRegie = sheet['J']
#iType_vendeur = sheet['H']
#iCom = sheet['N']

row_count = sheet.max_row
column_count = sheet.max_column

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if iMillesime[i].value is None :
        continue
    elif '+33 (0)' in iMillesime[i].value :
        continue
    elif  iMillesime[i].value is not None and iAppellation[i].value is None :
        continue
    else:

        rec_vin = unidecode(str(iVin[i].value).upper())
        appellation = str(iAppellation[i].value)
        if rec_vin == 'NONE' :
            vin = unidecode(appellation).upper()
            if '-' in str(iMillesime[i].value) :
                millesime = 'NV'
            else:
                millesime = iMillesime[i].value
        else:
            vin = rec_vin
            millesime = iMillesime[i].value
        
        prix = iPrix[i].value

        iSize = str(iFormatb[i].value).replace(' ','')
        if iSize == '37,5cl' :
            formatb = 'DE'
        elif iSize == '50cl':
            formatb = 'CL'
        elif iSize == '75cl' or iSize == '70cl':
            formatb = 'BO'
        elif iSize == '100cl':
            formatb = 'L'
        elif iSize == '150cl':
            formatb = 'MG'
        elif iSize == '300cl':
            formatb = 'DM'
        elif iSize == '500cl':
            formatb = 'JE'
        elif iSize == '600cl':
            formatb = 'IM'
        elif iSize == '900cl':
            formatb = 'SA'
        elif iSize == '1200cl':
            formatb = 'BA'
        elif iSize == '1500cl':
            formatb = 'NA'
        elif iSize == '1800cl':
            formatb = 'ME'
        elif iSize == '2700cl':
            formatb = 'BABY'
        else:
            formatb = iSize

        quantite = iQte[i].value

        rec_cond = re.findall('[C][BOC][L0-9ON]?[L ]?[EX]?[C ]?[T\d]+?[I]?[O]?[N]?', vin)
        if rec_cond :
            conditionnement = rec_cond[0].replace(' ','').replace('X','').replace('0','').replace('COLLECTION','COLLEC')
            vin = vin.replace(rec_cond[0],'')
            uCond = re.findall('\d+',conditionnement)
            if uCond :
                if conditionnement == 'CB'+uCond[0]:
                    conditionnement = conditionnement.replace('CB','CBO')
                quantite = int(quantite)*int(uCond[0])
        else:
            conditionnement = 'UNITE'
        
        rec_com = re.findall('[ENC][ATI][PAIV][SQ]?[:]',vin)
        if rec_com :
            iCom = vin.find(rec_com[0])
            commentaire = vin[int(iCom):]
            vin = vin.replace(commentaire,'')
        else:
            commentaire = ''

        # On affecte les valeurs dans l'ordre des colonnes voulues.
        c1 = ws.cell(row = i, column = 1)
        c1.value = vin

        c2 = ws.cell(row = i, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = i, column = 3)
        c3.value = formatb

        c4 = ws.cell(row = i, column = 4)
        c4.value = prix

        c5 = ws.cell(row = i, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = i, column = 6)
        c6.value = conditionnement

        c7 = ws.cell(row = i, column = 7)
        c7.value = commentaire

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)



   

