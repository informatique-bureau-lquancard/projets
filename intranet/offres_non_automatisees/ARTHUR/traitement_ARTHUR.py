# coding: utf-8
from cmath import e
import glob
import codecs
from operator import index
from unidecode import unidecode
from time import sleep
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook
import random

############################################################
#   Script basé sur /TWMOFF by Mr Clement Escande          #
############################################################


Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC','CBN1','CBN2','CBN3','CBN4','CBN6','CBN12','CC','CBO','CBN']

def is_float(n):
    try:
        float(n)   # Type-casting the string to `float`.
                   # If string is not a valid `float`, 
                   # it'll raise `ValueError` exception
    except ValueError:
        return False
    return True

def is_int(n):
    try:
        int(n)   # Type-casting the string to `float`.
                   # If string is not a valid `float`, 
                   # it'll raise `ValueError` exception
    except ValueError:
        return False
    return True


def a():
    wb = Workbook()

    for filename in glob.glob('*.xlsx'):
        wb = openpyxl.load_workbook(filename)

    #row_count = 0
    index = 0
    #on prépare un nouveau workbook
    wb2 = Workbook()

    dest_filename = 'sortie_ARTHUR.xlsx'

    ws = wb2.active
    ws.title = "ARTHUR ¯|(ツ)_|¯  "
    
    first_sheet = wb['Wines']
    exchange_rate = first_sheet['E10'].value
    print(exchange_rate)

    for sheet in wb :
        
        if sheet.title == 'Wines':
            iVin = sheet['E']
            iDomaine = sheet['D']
            iAppellation = sheet['F']
            iMillesime = sheet['C']
            iFormatb = sheet['I']
            iPrix = sheet['L']
            iQte = sheet['K']
            iCond = sheet['J']
            iRegion = sheet['G']
            iCom = sheet['A']
        elif sheet.title == 'Spirits':
            iVin = sheet['E']
            iDomaine = sheet['D']
            iMillesime = sheet['C']
            iFormatb = sheet['I']
            iPrix = sheet['L']
            iQte = sheet['K']
            iCond = sheet['J']
            iRegion = sheet['G']
            iCom = sheet['A']
        elif sheet.title == 'Champagne':
            iVin = sheet['E']
            iDomaine = sheet['D']
            iMillesime = sheet['C']
            iFormatb = sheet['G']
            iPrix = sheet['J']
            iQte = sheet['I']
            iCond = sheet['H']
            iCom = sheet['A']
            #iRegion = 'Champagne'

        for index_cellule  in range (sheet.max_row):
            # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
            if iMillesime[index_cellule].value is None :
                index = index + 1
                continue
            elif 'Vintage' in str(iMillesime[index_cellule].value) :
                index = index + 1
                continue
            else:
                rec_vin = unidecode(str(iVin[index_cellule].value).upper())
                region = unidecode(str(iRegion[index_cellule].value).upper())
                domaine = unidecode(str(iDomaine[index_cellule].value).upper())
                if region == 'BORDEAUX':
                    if 'VERTICALE' in rec_vin :
                        vin = domaine+' '+rec_vin
                    else:
                        vin = rec_vin
                else:
                    vin = domaine+' '+rec_vin
                
                if iMillesime[index_cellule].value is None or iMillesime[index_cellule].value == '':
                    millesime = 'NV'
                else:
                    millesime = iMillesime[index_cellule].value

                iSize = str(iFormatb[index_cellule].value).replace(' ','').replace('.',',')
                if iSize == '37,5cl':
                    formatb = 'DE'
                elif iSize == '50cl':
                    formatb = 'CL'
                elif iSize == '75cl' or iSize == '70cl':
                    formatb = 'BO'
                elif iSize == '100cl' or iSize == '1L':
                    formatb = 'L'
                elif iSize == '150cl' or iSize == '1,5L':
                    formatb = 'MG'
                elif iSize == '300cl' or iSize == '3L':
                    formatb = 'DM'
                elif iSize == '500cl' or iSize == '5L':
                    formatb = 'JE'
                elif iSize == '600cl' or iSize == '6L':
                    formatb = 'IM'
                elif iSize == '900cl' or iSize == '9L':
                    formatb = 'SA'
                elif iSize == '1200cl' or iSize == '12L':
                    formatb = 'BA'
                elif iSize == '1500cl' or iSize == '15L':
                    formatb = 'NA'
                elif iSize == '1800cl' or iSize == '18L':
                    formatb = 'ME'
                elif iSize == '2700cl' or iSize == '27L':
                    formatb = 'BABY'
                else:
                    formatb = iSize

                prix = iPrix[index_cellule].value
                if(not is_float(prix)) :
                    prix = 0
                else:
                    prix = prix * exchange_rate

                quantite = iQte[index_cellule].value

                conditionnement = str(iCond[index_cellule].value).upper().replace('CTN','CC').replace('BTLE','UNITE').replace('OC','CC').replace('OWC','CBO')
                Cond_Com = 0
                if 'COLLECTION' in vin :
                    conditionnement = 'COLLEC'
                    #quantite = re.findall('[0-9][0-9]?[0-9]?[0-9]?',quantite)[0]
                elif conditionnement not in Dict_cond :
                    conditionnement = 'UNITE'
                    Cond_Com = 1
                    
                
                if Cond_Com == 1 :
                    if iCom[index_cellule].value is None :
                        commentaire = 'VERIF CDT - '+unidecode(str(iCond[index_cellule].value))
                    else:
                        commentaire = 'VERIF CDT - '+unidecode(str(iCond[index_cellule].value))+' - '+unidecode(str(iCom[index_cellule].value))
                else:
                    if iCom[index_cellule].value is None :
                        commentaire = ''
                    else:
                        commentaire = unidecode(str(iCom[index_cellule].value))
    

                # On affecte les valeurs dans l'ordre des colonnes voulues.
                c1 = ws.cell(row = index, column = 1)
                c1.value = vin

                c2 = ws.cell(row = index, column = 2)
                c2.value = millesime

                c3 = ws.cell(row = index, column = 3)
                c3.value = formatb

                c4 = ws.cell(row = index, column = 4)
                c4.value = prix

                c5 = ws.cell(row = index, column = 5)
                c5.value = quantite

                c6 = ws.cell(row = index, column = 6)
                c6.value = conditionnement

                c7 = ws.cell(row = index, column = 7)
                c7.value = commentaire

                index = index + 1

        index_row = []

        for n in range(ws.max_row) :
            if ws.cell(n+1, 1).value is None:
                index_row.append(n+1)

        for row_del in range(len(index_row)):
            ws.delete_rows(idx=index_row[row_del], amount=1)
            index_row = list(map(lambda k: k -1, index_row))
        
        wb2.save(dest_filename)

a()
