# coding: utf-8
import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook


# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_LDVOFF.xlsx'
ws = wb2.active
ws.title = "LDVOFF"


# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['B']
iMillesime = sheet['C']
iFormatb = sheet['D']
iPrix = sheet['G']
iQte = sheet['K']
iCond = sheet['M']
#iDate_offre = sheet['I']  #validité
#iRegie = sheet['J']
#iType_vendeur = sheet['H']
#iCom = sheet['N']

row_count = sheet.max_row
column_count = sheet.max_column

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if iVin[i].value is None :
        continue
    elif '+33 (0)' in iVin[i].value :
        continue
    elif  '@' in iVin[i].value :
        continue
    elif iPrix[i].value is None:
        continue
    elif str(iVin[i].value) == 'WINE' :
        continue
    elif str(iVin[i].value) == '.' :
        continue
    elif str(iVin[i].value) == '(AOC)' :
        continue
    elif str(iVin[i].value) == '' :
        continue
    else:  
        vin = unidecode(str(iVin[i].value).upper())
        
        millesime = iMillesime[i].value

        prix = iPrix[i].value

        iSize = str(iFormatb[i].value)
        if iSize == '37,5cl' :
            formatb = 'DE'
        elif iSize == '50cl':
            formatb = 'CL'
        elif iSize == '75cl' or '70cl':
            formatb = 'BO'
        elif iSize == '100cl':
            formatb = 'L'
        elif iSize == '150cl':
            formatb = 'MG'
        elif iSize == '300cl':
            formatb = 'DM'
        elif iSize == '500cl':
            formatb = 'JE'
        elif iSize == '600cl':
            formatb = 'IM'
        elif iSize == '900cl':
            formatb = 'SA'
        elif iSize == '1200cl':
            formatb = 'BA'
        elif iSize == '1500cl':
            formatb = 'NA'
        elif iSize == '1800cl':
            formatb = 'ME'
        elif iSize == '2700cl':
            formatb = 'BABY'
        else:
            formatb = iSize

        quantite = iQte[i].value
        #quantite = int(quantite)

        rec_cond = str(iCond[i].value).upper().replace('OWC','CBO').replace('0','').replace('NWC','CBN').replace('BTL','UNITE').replace('GB','GIBO')
        # Dictionnaire des conditionnements reconnus
        Dict_Cond = ['CBO','CBN','CBO1','CBO2','CBO3','CBO4','CBO6','CBO12','CBO12DE','CBO24DE','CBN1','CBN2','CBN3','CBN4','CBN6','CBN12','CBN12DE','CBN24DE','CC','CC1','CC2','CC3','CC4','CC6','CC12','CC12DE','CC24DE','CCN','UNITE','GIBO1']

        if rec_cond in Dict_Cond :
            conditionnement = rec_cond
            commentaire = ''
        else:
            conditionnement = 'UNITE'
            commentaire = rec_cond


        # On affecte les valeurs dans l'ordre des colonnes voulues.
        c1 = ws.cell(row = i, column = 1)
        c1.value = vin

        c2 = ws.cell(row = i, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = i, column = 3)
        c3.value = formatb

        c4 = ws.cell(row = i, column = 4)
        c4.value = prix

        c5 = ws.cell(row = i, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = i, column = 6)
        c6.value = conditionnement

        c7 = ws.cell(row = i, column = 7)
        c7.value = commentaire

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)



   

