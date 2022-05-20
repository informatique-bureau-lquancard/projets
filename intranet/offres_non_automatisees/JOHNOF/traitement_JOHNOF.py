# coding: utf-8
import csv
from email import header
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook
import pandas as pd

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

nomProfil : str = "JOHNOF"
extensionDebut : str = ".xls"
extensionFin : str = ".xlsx"

# for filename in glob.glob('*' + extensionDebut):
#     xlsconverter = pd.read_excel(filename)
#     xlsconverter.to_excel('johnof.xlsx', index=False, header=False)

# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_JOHNOF.xlsx'
ws = wb2.active
# ws.title = "JOHNOF ლ(•́•́ლ) "
ws.title = nomProfil

# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['B']
iMillesime = sheet['C']
iFormatb = sheet['A']
iPrix = sheet['H']
iQte_iCond = sheet['D']
#iCond = sheet['I']
#iDate_offre = sheet['I']  #validité
#iRegie = sheet['J']
#iType_vendeur = sheet['H']
iCom = sheet['D']
iAppellation = sheet['G']

row_count = sheet.max_row
column_count = sheet.max_column

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if iVin[i].value is None :
        continue
    elif iQte_iCond[i].value == 'Quantity' or iQte_iCond[i].value == 'Colonne5' :
        continue
    else:  
        vin = unidecode(str(iVin[i].value).upper()).replace('CH. ','').strip()
        
        millesime : str = ft.formaterAnnee( str(iMillesime[i].value) )

        if( not millesime.isnumeric() ):
            millesime = 'NV'

        prix = iPrix[i].value

        iSize = str(iFormatb[i].value).upper()
        if iSize == 'HALVES' :
            formatb = 'DE'
        elif iSize == 'BLLE':
            formatb = 'BO'
        elif iSize == 'MAG':
            formatb = 'MG'
        elif iSize == 'DMAG':
            formatb = 'DM'
        elif iSize == 'JERO':
            formatb = 'JE'
        elif iSize == 'MARJEA':
            formatb = 'MJ'
        elif iSize == 'BALT':
            formatb = 'BA'
        elif iSize == 'SALM':
            formatb = 'SA'
        elif iSize == 'IMP(6L)':
            formatb = 'IM'
        else:
            formatb = iSize

        if iQte_iCond[i].value is None  or iQte_iCond[i].value == '':
            continue

        Qte_Cond = str(iQte_iCond[i].value)

        if 'cs & ' in Qte_Cond or 'cs + ' in Qte_Cond :
            Qte_Cond = Qte_Cond.replace('cs ','cs/12')
                
        rec_qte_cs_cc = re.findall('\d+[c][cs]?',Qte_Cond)
        rec_uqte_cs_cc = re.findall('[/]\d+',Qte_Cond)
        rec_qte_unite = re.findall('\d+[ ][u]',Qte_Cond)

        
        
        if len(rec_qte_cs_cc) > 1 :
            q = 0
            
            for q in range(len(rec_qte_cs_cc)) :
                qte_cs = int(re.findall('\d+',rec_qte_cs_cc[q])[0])
                qte_u = int(re.findall('\d+',rec_uqte_cs_cc[q])[0])

                quantite1 = qte_cs * qte_u
                conditionnement = 'CBO'+str(qte_u)
                q = q+1
                qte_cs = int(re.findall('\d+',rec_qte_cs_cc[q])[0])
                qte_u = int(re.findall('\d+',rec_uqte_cs_cc[q])[0])
                
                quantite2 = qte_cs * qte_u
                quantite = quantite1 +quantite2
                
                break

        elif rec_qte_cs_cc and not rec_uqte_cs_cc :    
            qte_cs = int(re.findall('\d+',str(rec_qte_cs_cc[0]))[0])
            quantite = qte_cs * 12
            conditionnement = 'CBO12'

        elif rec_qte_cs_cc and rec_uqte_cs_cc :
            qte_cs = int(re.findall('\d+',str(rec_qte_cs_cc[0]))[0])
            qte_u = int(re.findall('\d+',str(rec_uqte_cs_cc[0]))[0])
            quantite = qte_cs * qte_u
            conditionnement = 'CBO'+str(qte_u)

        else:
            quantite = re.findall('\d+',Qte_Cond)[0]
            conditionnement = 'UNITE'


        # On affecte les valeurs dans l'ordre des colonnes voulues.
        c1 = ws.cell(row = i, column = 1)
        c1.value = vin

        c2 = ws.cell(row = i, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = i, column = 3)
        c3.value = formatb

        c4 = ws.cell(row = i, column = 4)
        c4.value = prix

        c5 = ws.cell(row = i, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = i, column = 6)
        c6.value = conditionnement

        # c7 = ws.cell(row = i, column = 7)
        # c7.value = commentaire

#Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)



   

