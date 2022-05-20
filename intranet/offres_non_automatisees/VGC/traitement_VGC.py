# coding: utf-8
import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook
import random
import pandas as pd

Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC']

# Le traitement marche en local mais plus sur le site !!!


nomProfil : str = "VGC"
extensionDebut = ".xls"
extensionFin = ".csv"


# On load le le tarif au format xlsx
for filename in glob.glob("*"+extensionDebut):

    with open(filename, mode="r", encoding = 'utf-8') as file :

        reader = pd.read_excel(filename, header = None, engine = 'xlrd')
    #reader = reader(header = 15)
    #index_with_nan = reader.index[reader.iloc[:,1].isnull()]
    #reader.drop(index_with_nan,0, inplace=True)
        reader.sort_values(by = [0], axis = 0, ascending=True, inplace=True)
    #print(reader.columns[1])
        reader.to_excel('vgc.xlsx', index = False, header = False, engine = 'openpyxl')
        wb = openpyxl.load_workbook('vgc.xlsx')

    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename_bis = 'sortie_'+ 'bis_' + nomProfil.upper()+extensionFin
    dest_filename = 'sortie_'+ nomProfil.upper()+extensionFin
    ws = wb2.active
    ws.title = "VGC"

    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    iVin = sheet['A']
    iMillesime = sheet['B']
    iFormatb = sheet['C']
    iPrix = sheet['D']
    iQte = sheet['E']
    iCond = sheet['F']
    iCom = sheet['G']

    row_count = sheet.max_row
    column_count = sheet.max_column

    for i  in range (1, row_count):
        # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
        if iVin[i].value is None or iVin[i].value == 'Nom Vin' :
            continue
        else:
            
            vin = unidecode(str(iVin[i].value).upper())
            spliter = re.findall('[ ][0-9]?[0-9]?[0-9]?[0-9]?[ ][0-9][0-9][,.]?[0-9]?[ ][C][L]', vin)
            if spliter :
                vin = vin.split(spliter[0])[0]      
            
            if iMillesime[i].value is None :
                millesime = 'NV'
            else:
                millesime = iMillesime[i].value

            prix = iPrix[i].value

            formatb = iFormatb[i].value
        
            quantite = iQte[i].value

            conditionnement = iCond[i].value

            commentaire = unidecode(str(iCom[i].value))



            c1 = ws.cell(row = i, column = 1)
            c1.value = vin

            c2 = ws.cell(row = i, column = 2)
            c2.value = millesime

            c3 = ws.cell(row = i, column = 3)
            c3.value = formatb

            c4 = ws.cell(row = i, column = 4)
            c4.value = prix

            c5 = ws.cell(row = i, column = 5)
            c5.value = quantite

            c6 = ws.cell(row = i, column = 6)
            c6.value = conditionnement

            c7 = ws.cell(row = i, column = 7)
            c7.value = commentaire      
        
            if i == 1:
                i = i
            else:
                if  ws.cell(row = i, column = 1).value == ws.cell(row = i-1, column = 1).value and ws.cell(row = i, column = 2).value == ws.cell(row = i-1, column = 2).value and ws.cell(row = i, column = 3).value == ws.cell(row = i-1, column = 3).value :
                    while ws.cell(row = i, column = 6).value == ws.cell(row = i-1, column = 6).value :
                        c6.value = random.choice(Dict_cond)
                        c7.value = 'ATTN : VERIF CDT - DOUBLON LIGNE PRECEDENTE - '+c7.value

    #Fonction de suppression des lignes vides d'un workbook
    index_row = []
    for n in range(1, ws.max_row):
        if ws.cell(n, 1).value is None:
            index_row.append(n)

    for row_del in range(len(index_row)):
        ws.delete_rows(idx=index_row[row_del], amount=1)
        index_row = list(map(lambda k: k -1, index_row))

    wb2.save(dest_filename_bis)

    read_file = pd.read_excel(dest_filename_bis)
    read_file.to_csv(dest_filename, index = False, encoding="utf-8", sep = ';')