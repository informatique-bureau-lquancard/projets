# coding: utf-8
import csv
import string
import os
import glob
import codecs
from unidecode import unidecode
from time import sleep
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

# fonction pour tester si tous les caractères d'une chaine sont en majuscule
# def is_all_uppercase(a_str):
#     for c in a_str:
#         if c not in string.ascii_uppercase:
#             return False
#     return True

# Récupération des fichiers
# paths = []

# cheminAcces = "php/CDP"
# cheminAccesAncient = "/Users/Julien/Documents/TRAITEMENT TARIFS/scripts"

# for root, dirs, files in os.walk(r""+cheminAcces):
#     for file in files:
#         if file.endswith(".xls"):
#              s = os.path.join(root, file)
#              print(s)
#              paths.append(s)

# # Collecte des données
# all_data = pd.DataFrame()
# for f in paths:
#     df = pd.read_excel(f, skiprows = 0)
#     all_data = all_data.append(df,ignore_index=False)

# # Nettoyage des données
# old_colnames = (all_data.columns.values)
# new_colnames = (all_data.iloc[0])
# all_data.rename(columns={i:j for i,j in zip(old_colnames,new_colnames)}, inplace=True)
# all_data.drop(0, inplace = True)

# # Export en Excel
# writer = pd.ExcelWriter('full_cpd.xlsx')
# all_data.to_excel(writer,'sheet1', index = False)
# writer.save()


extension_entree = ".xls"

# On load le tarif au format xlsx
for filename in glob.glob('*'+extension_entree):
    reader = pd.read_excel(filename)
    reader.to_excel('cdp.xlsx', engine = 'openpyxl', index = False, header = False)

    wb = openpyxl.load_workbook('cdp.xlsx')

    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename = 'sortie_CDP.xlsx'
    ws = wb2.active
    ws.title = "CDP"

    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    # iDomaine_appellation_millesime = sheet['A']
    iMillesime_Region = sheet['A']
    iDomaine = sheet['D']
    iAppellation = sheet['E']
    iFormatB = sheet['B']
    iPrix = sheet['F']
    iQte = sheet['C']

    row_count = sheet.max_row
    column_count = sheet.max_column
    
    j = 1

    for i  in range (1, row_count):

        if iMillesime_Region[i].value is None :
            continue
        elif str(iMillesime_Region[i].value) == 'Mill.':
            continue
        elif iMillesime_Region[i].value is not None and iDomaine[i].value is None :
            region = str(iMillesime_Region[i].value)
            continue

        prix = str(iPrix[i].value)

        prix = prix.upper()
        prix = float(prix)

        vin = unidecode(str(iDomaine[i].value).upper()+' '+str(iAppellation[i].value).upper())

        millesime = iMillesime_Region[i].value

        formatb = ft.formaterFormatBouteille(iFormatB[i].value)

        quantite = int(iQte[i].value)

        conditionnement = ft.conditionnementParDefaut(formatb, quantite)

        commentaire = ''

        #On affecte les valeurs dans l'ordre des colonnes voulues.
        ft.affectationLignes(j, ws, vin, millesime, formatb, prix, quantite, conditionnement, commentaire)

        j = j + 1

    # Fonction de suppression des lignes vides d'un workbook
    index_row = []
    for n in range(1, ws.max_row):
        if ws.cell(n, 1).value is None:
            index_row.append(n)

    for row_del in range(len(index_row)):
        ws.delete_rows(idx=index_row[row_del], amount=1)
        index_row = list(map(lambda k: k -1, index_row))

    wb2.save(dest_filename)
    csv_converter = pd.read_excel(dest_filename)
    csv_converter.to_csv("sortie_CDP.csv", index = False, encoding = "utf-8", sep = ";")




   

