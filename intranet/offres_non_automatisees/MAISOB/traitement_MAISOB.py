import gspread
import pandas as pd
from unidecode import unidecode
import openpyxl
from openpyxl import Workbook
import glob
import re
import random

import csv

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

# MAISOB

extensionDebut : str = '.xlsx'

#Initialisation et création du fichier de  xlsx
def initialisation_google_sheet(tab_onglets_ok, flux, chemin_fichier_sortie):

    # print ("  initialisation_google_sheet  ")
    # print(chemin_fichier_sortie)

    # chemin_fichier_sortie_xlsx = chemin_fichier_sortie+'.xlsx'

    # authentification à google API,  compte j.martin.blq@gmail.com
    # fichier credential.json dans le répertoire du script
    gc = gspread.service_account(filename='/var/www/html/projets/bureau-laurent-quancard/gestion-des-offres/gestion_tarifs_automatises/Tarifs/tests_tarifs/credential.json')

    # ouverture du spreadsheet google de Maion B
    sh = gc.open_by_key(flux)

    # Récupération de la liste des onglets
    worksheet_list = sh.worksheets()

    workbook = pd.ExcelWriter(chemin_fichier_sortie+"1" + extensionDebut, engine = 'openpyxl')
    #pd.to_excel(workbook,sheet_name="Sheet1", startrow=1, startcol=1,index=False, header=True)

    sheets = []

    nombre_onglets_entree = len(worksheet_list)

    # print("nombre_onglets_entree : "+str(nombre_onglets_entree))

    #A revoir si le tableau tab_onglets n'est pas renseigné
    #if(len(tab_onglets) < 1):
    index_onglet_sortie = 0

    # Parcours des onglet et création de dataframe Pandas par onglets 
    for index_onglet in range(nombre_onglets_entree) :

        active_worksheet = sh.get_worksheet(index_onglet)
        sheet_title = unidecode(str(active_worksheet).split(' ')[1].replace("'",""))
        
        # Test les onglets à récupérer
        if len(tab_onglets_ok) != 0 :
            if sheet_title not in tab_onglets_ok:
                continue

        sheets.append(sheet_title)   

        # print('titre : '+sheet_title)
        # print('index_onglet_sortie : '+str(index_onglet_sortie))

        sheets[index_onglet_sortie] = pd.DataFrame(active_worksheet.get_all_records())

        sheets[index_onglet_sortie].to_excel(workbook,sheet_title, index = False)

        index_onglet_sortie = index_onglet_sortie + 1

    workbook.save()

def mise_en_forme_csv(tab_designation, chemin_fichier_sortie):
    
    with open(chemin_fichier_sortie+'2.csv', newline='', encoding="utf-16") as csvfile :
        reader = csv.DictReader(csvfile)
        with open(chemin_fichier_sortie+"3.csv",'w',newline='', encoding="utf-16") as monFichierSortie:
            writer = csv.writer(monFichierSortie, delimiter=';', quoting=csv.QUOTE_NONE, escapechar=';', quotechar='')
            
            # print(tab_designation)

            for row in reader :

                newRow = []

                for index_designation in range(len(tab_designation)):

                    #Test sur la valeur du vin
                    if(index_designation == 0):
                        colonne_vin = unidecode(row[tab_designation[index_designation]]).upper().strip().replace(';','').replace(', ','-')

                        newRow.append(colonne_vin)
                        continue

                    #Test sur la valeur du prix
                    if(index_designation == 3):
                        taille_colonne_prix = len(str(row[tab_designation[index_designation]]))

                        if( taille_colonne_prix < 1):
                            newRow.append(0)
                            continue

                    newRow.append(row[tab_designation[index_designation]])

                writer.writerow(newRow)

def recuperation_colonnes(active_sheet):
    tab_colonnes = []

    max_column : int = active_sheet.max_column

    # print("max_column : "+str(max_column))

    #Récupération des colonnes dans un tableau
    for index_colonnes in range(max_column):

        #Utilisation du tableau ASCII : 65 = A

        indice_lettre : int = 65 + index_colonnes

        #print("indice_lettre : "+str(indice_lettre))

        lettre_colonne : str = chr(indice_lettre)

        #print("lettre_colonne : "+lettre_colonne)

        # print(active_sheet[lettre_colonne])

        tab_colonnes.append(active_sheet[lettre_colonne])

    return tab_colonnes

nomProfil : str = "MAISOB"
flux : str = "1aqe6HRgndilOFTEmuSncSGNly7VesKExNjIouGXW0UM"
chemin_dossier_tarifs : str = "/var/www/html/php/"
chemin_fichier_sortie : str = chemin_dossier_tarifs + nomProfil + "/sortie_" + nomProfil

nom_tarif_profil : str = nomProfil
id_negociant : int = 4

extensionFin_bis : str = '.xlsx'
extensionFin : str = '.csv'

def fct_flux_google_sheets2(flux, chemin_fichier_sortie, nom_tarif_profil, id_negociant):

    #Initialisation des onglets faisant parti du fichier final, quand vide prendre tout
    tab_onglets_ok = []
    
    #Initialisation et création du fichier de  xlsx
    initialisation_google_sheet(tab_onglets_ok, flux, chemin_fichier_sortie)

    #transforme en csv afin d'utiliser DictReader
    reader = pd.read_excel(chemin_fichier_sortie+"1.xlsx")

    # reader = reader.sort_values(by = [0], axis = 0, ascending=True, inplace=True)

    reader.to_csv(chemin_fichier_sortie+'2.csv', index = False, encoding="utf-16")

    vin_desi = 'Château / Domaine'
    millesime_desi = 'Millésime'
    formatB_desi = 'Format'
    prix_desi = '€/btl.'
    quantite_desi = 'Qté'
    conditionnement_desi = 'Condit.'
    # commentaires_desi = 'Stock Stock Physique'

    appellation = 'Appellation'
    couleur = 'Couleur'
    regie = 'Régie'

    tab_designation = [vin_desi, millesime_desi, formatB_desi, prix_desi, quantite_desi, conditionnement_desi, appellation, couleur, regie]

    mise_en_forme_csv(tab_designation, chemin_fichier_sortie)

    #transforme en xlsx afin de pouvoir utiliser openpyxl
    new_pd_reader = pd.read_csv(chemin_fichier_sortie+"3.csv", sep = ';', encoding="utf-16")
    new_pd_reader.to_excel(chemin_fichier_sortie+"4.xlsx", index = False, encoding="utf-16")

    # On load le le tarif au format xlsx dans un workbook
    wb = openpyxl.load_workbook(chemin_fichier_sortie+"4.xlsx")

    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename_bis = chemin_fichier_sortie + extensionFin_bis
    dest_filename = chemin_fichier_sortie + extensionFin

    # dest_filename_bis = 'sortie_bis_' + nomProfil + extensionFin_bis
    # dest_filename = 'sortie_' + nomProfil + extensionFin
    
    ws = wb2.active
    ws.title = nom_tarif_profil

    # print("Nombre d'onglet : "+str(len(wb.sheetnames)))

    # Lecture du workbook d'origine
    for j in range(len(wb.sheetnames)) :

        # print(j)
        sheet_list = wb.sheetnames
        active_sheet = wb[sheet_list[j]]

        tab_colonnes = recuperation_colonnes(active_sheet)

        iVin = tab_colonnes[0]
        iProducteur = ''
        iCouleur = ''
        iMillesime = tab_colonnes[1]
        iFormatB = tab_colonnes[2]
        iPrix = tab_colonnes[3]
        iQte = tab_colonnes[4]
        iCond = tab_colonnes[5]
        # iCom = tab_colonnes[6]
        iAppellation = tab_colonnes[6]
        iCouleur = tab_colonnes[7]
        iRegie = tab_colonnes[8]

        row_count = active_sheet.max_row
        column_count = active_sheet.max_column

        index_sortie : int = 1

        for index_ligne in range(row_count):

            exclude_appellation = []

            appellation : str = ft.formaterAppellation( str(iAppellation[index_ligne].value), exclude_appellation)
            couleur : str = str(iCouleur[index_ligne].value)

            vin : str = unidecode(str(iVin[index_ligne].value))
            vin = ft.formaterVin(appellation, vin, couleur)

            # print("   ")
            # print("index : " + str(index_sortie))

            iMillesime_bis : str = re.findall('[0-9][0-9][0-9][0-9]', str(iMillesime[index_ligne].value))

            # print("iMillesime : " + str(iMillesime_bis))

            if( not (len(iMillesime_bis) > 0) ):
                millesime = "NV"
            else:
                millesime : str = ft.formaterAnnee( str(iMillesime_bis[0]) )

            # print("millesime : " + str(millesime))
            
            formatB = str(iFormatB[index_ligne].value)
            formatB = ft.formaterFormatBouteille(formatB)

            prix = float(unidecode(str(iPrix[index_ligne].value)).replace(' ','').replace('EUR',''))

            # print("prix : "+str(prix))

            quantite = iQte[index_ligne].value

            Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC'] # pour le random / doublons
            rec_cond = str(iCond[index_ligne].value).replace('UNIT','UNITE').replace('CT','CC').replace('CO','CC')
            
            if 'COLLECTION' in vin :
                conditionnement = 'COLLEC'
            else:
                conditionnement = rec_cond

            # commentaires = str(iRegie[index_ligne].value)+' '+str(iCom[index_ligne].value)

            commentaires = ""

            ft.affectationLignes(index_sortie, ws, vin, str(millesime), formatB, str(prix), str(quantite), conditionnement, commentaires)

            index_sortie += 1    

    # Fonction de suppression des lignes vides d'un workbook
    index_row = []
    for n in range(1, ws.max_row):
        if ws.cell(n, 1).value is None:
            index_row.append(n)

    for row_del in range(len(index_row)):
        ws.delete_rows(idx=index_row[row_del], amount=1)
        index_row = list(map(lambda k: k -1, index_row))

    wb2.save(dest_filename_bis)

    read_file = pd.read_excel(dest_filename_bis)
    read_file.to_csv(dest_filename, index = False, encoding="utf-8", sep = ';')

fct_flux_google_sheets2(flux, chemin_fichier_sortie, nom_tarif_profil, id_negociant)


# Ancien
# # authentification à google API,  compte j.martin.blq@gmail.com
# # fichier credential.json dans le répertoire du script
# gc = gspread.service_account(filename='credential.json')

# # ouverture du spreadsheet google de Maion B
# sh = gc.open_by_key("1aqe6HRgndilOFTEmuSncSGNly7VesKExNjIouGXW0UM")

# # Récupération de la liste des onglets
# worksheet_list = sh.worksheets()
# w =0

# writer = pd.ExcelWriter('sortie_MAISOB.xlsx', engine = 'openpyxl')
# sheets = []

# # Parcours des onglet et création de dataframe Pandas par onglets 
# while w in range(len(worksheet_list)) :
#     active_worksheet = sh.get_worksheet(w)
#     sheet_title = unidecode(str(active_worksheet).split(' ')[1].replace("'",""))
#     sheets.append(sheet_title)
#     sheets[w] = pd.DataFrame(active_worksheet.get_all_records())
#     sheets[w].to_excel(writer,sheet_title, index = False)
#     w = w+1

# writer.save()

# nomProfil : str = 'MAISOB'
# extensionDebut : str = '.xlsx'
# extensionFin_bis : str = '.xlsx'
# extensionFin : str = '.csv'

# # On load le le tarif au format xlsx dans un workbook
# for filename in glob.glob('*' + extensionDebut):
#     wb = openpyxl.load_workbook(filename)

# # On prépare un nouveau workbook
# wb2 = Workbook()
# dest_filename_bis = 'sortie_bis_' + nomProfil + extensionFin_bis
# dest_filename = 'sortie_' + nomProfil + extensionFin
# ws = wb2.active
# ws.title = nomProfil

# # Lecture du workbook d'origine
# s = 0
# while s in range(len(sheets)) :
#     sheet_list = wb.sheetnames
#     active_sheet = wb[sheet_list[s]]
#     if sheet_list[s] == 'Bordeaux':
#         iVin = active_sheet['A']
#         iProducteur = ''
#         iCouleur = ''
#         iMillesime = active_sheet['B']
#         iFormatb = active_sheet['C']
#         iPrix = active_sheet['E']
#         iQte = active_sheet['D']
#         iCond = active_sheet['F']
#         iRegie = active_sheet['G']
#         iCom = active_sheet['H']

#         row_count = active_sheet.max_row
#         column_count = active_sheet.max_column
#     else:
#         iProducteur = active_sheet['A']
#         iVin = active_sheet['B']
#         iCouleur = active_sheet['D']
#         iMillesime = active_sheet['E']
#         iFormatb = active_sheet['F']
#         iPrix = active_sheet['H']
#         iQte = active_sheet['G']
#         iCond = active_sheet['I']
#         iRegie = active_sheet['K']
#         iCom = active_sheet['L']

#         row_count = active_sheet.max_row
#         column_count = active_sheet.max_column
    
#     s = s+1

#     for i  in range (1, row_count):
#         # if iMillesime[i].value == 'Millésime':
#         #     continue 
        
#         if iProducteur:
#             if iProducteur[i].value is None :
#                 continue
#             else:
#                 exclude_regions = ['Languedoc','Loire','USA']
#                 if str(iProducteur[i].value) in exclude_regions :
#                     vin = unidecode(str(iVin[i].value)).upper()
#                 else:
#                     vin = unidecode(str(iProducteur[i].value)+' '+str(iVin[i].value)).upper()
#         else:
#             vin = unidecode(str(iVin[i].value)).upper()
        
#         millesime = re.findall('[0-9][0-9][0-9][0-9]',str(iMillesime[i].value))
#         if millesime :
#             millesime = int(millesime[0])
#         else:
#             millesime = 'NV'
        
#         iSize = str(iFormatb[i].value).replace(' ','')
#         if iSize == '37,5cl' :
#             formatb = 'DE'
#         elif iSize == '50cl':
#             formatb = 'CL'
#         elif iSize == '75cl' or iSize == '70cl':
#             formatb = 'BO'
#         elif iSize == '100cl':
#             formatb = 'L'
#         elif iSize == '150cl':
#             formatb = 'MG'
#         elif iSize == '300cl':
#             formatb = 'DM'
#         elif iSize == '500cl':
#             formatb = 'JE'
#         elif iSize == '600cl':
#             formatb = 'IM'
#         elif iSize == '900cl':
#             formatb = 'SA'
#         elif iSize == '1200cl':
#             formatb = 'BA'
#         elif iSize == '1500cl':
#             formatb = 'NA'
#         elif iSize == '1800cl':
#             formatb = 'ME'
#         elif iSize == '2700cl':
#             formatb = 'BABY'
#         else:
#             formatb = iSize

#         prix = float(unidecode(str(iPrix[i].value)).replace(' ','').replace('EUR',''))

#         quantite = iQte[i].value

#         Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC'] # pour le random / doublons
#         rec_cond = str(iCond[i].value).replace('UNIT','UNITE').replace('CT','CC').replace('CO','CC')
#         if 'COLLECTION' in vin :
#             conditionnement = 'COLLEC'
#         else:
#             conditionnement = rec_cond
        
#         commentaire = str(iRegie[i].value)+' '+str(iCom[i].value)

#         c1 = ws.cell(row = i, column = 1)
#         c1.value = vin

#         c2 = ws.cell(row = i, column = 2)
#         c2.value = millesime

#         c3 = ws.cell(row = i, column = 3)
#         c3.value = formatb

#         c4 = ws.cell(row = i, column = 4)
#         c4.value = prix

#         c5 = ws.cell(row = i, column = 5)
#         c5.value = quantite

#         c6 = ws.cell(row = i, column = 6)
#         c6.value = conditionnement

#         c7 = ws.cell(row = i, column = 7)
#         c7.value = commentaire      

        
#         # if  ws.cell(row = i, column = 1).value == ws.cell(row = i-1, column = 1).value and ws.cell(row = i, column = 2).value == ws.cell(row = i-1, column = 2).value and ws.cell(row = i, column = 3).value == ws.cell(row = i-1, column = 3).value :
#         #     while ws.cell(row = i, column = 6).value == ws.cell(row = i-1, column = 6).value :
#         #         c6.value = random.choice(Dict_cond)
#         #         c7.value = 'ATTN : VERIF CDT - DOUBLON LIGNE PRECEDENTE - '+c7.value


# # Fonction de suppression des lignes vides d'un workbook
# index_row = []
# for n in range(1, ws.max_row):
#     if ws.cell(n, 1).value is None:
#         index_row.append(n)

# for row_del in range(len(index_row)):
#     ws.delete_rows(idx=index_row[row_del], amount=1)
#     index_row = list(map(lambda k: k -1, index_row))

# wb2.save(dest_filename_bis)

# read_file = pd.read_excel(dest_filename_bis)
# read_file.to_csv(dest_filename, index = False, encoding="utf-8", sep = ';')
