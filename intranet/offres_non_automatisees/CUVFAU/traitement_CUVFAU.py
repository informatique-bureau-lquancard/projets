import glob
import random
import re
import gspread
import openpyxl
import pandas as pd
from openpyxl import Workbook
from unidecode import unidecode

Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC']

# authentification à google API,  compte j.martin.blq@gmail.com
# fichier credential.json dans le répertoire du script
gc = gspread.service_account(filename='credential.json')

# ouverture du spreadsheet google de Maion B
sh = gc.open_by_key("1dWHRCDE99uXQ1IbTZNm1XYLzRwMwLyg3TcJzmzINV-M")

# Récupération de la liste des onglets
worksheet_list = sh.worksheets()
w =0

writer = pd.ExcelWriter('sortie_CUVFAU.xlsx', engine = 'openpyxl')
sheets = []

# Parcours des onglet et création de dataframe Pandas par onglets 
while w in range(len(worksheet_list)) :
    active_worksheet = sh.get_worksheet(w)
    sheet_title = unidecode(str(active_worksheet).split(' ')[1].replace("'",""))
    sheets.append(sheet_title)
    sheets[w] = pd.DataFrame(active_worksheet.get_all_records())
    # sheets[w].reindex(1,len(sheets[w].index+1))
    sheets[w].to_excel(writer,sheet_title, index = False)
    w = w+1

writer.save()

# On load le le tarif au format xlsx dans un workbook
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_CUVFAU.xlsx'
ws = wb2.active
ws.title = "CUVFAU"


# Lecture du workbook d'origine
s = 0
while s in range(len(sheets)) :
    sheet_list = wb.sheetnames
    active_sheet = wb[sheet_list[s]]
    if sheet_list[s] == 'RESULTAT':
        iVin = active_sheet['A']
        iMillesime = active_sheet['B']
        iAppellation = active_sheet['C']
        iRegie = active_sheet['D']
        iFormatb = active_sheet['E']
        iCond = active_sheet['F']
        iCouleur = active_sheet['G']
        iQte = active_sheet['K']
        iCom = active_sheet['L']
        iPrix = active_sheet['M']

        row_count = active_sheet.max_row
        column_count = active_sheet.max_column

        
    s = s+1

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    # if iVin[i].value is None or str(iVin[i].value) == 'Product Info Name' :
    #     continue
    #else:
    if iQte[i].value is None and iCom[i].value is None :
        continue
    exclude_appellation = ['SAUTERNES','BARSAC']
    couleur = str(iCouleur[i].value).upper()
    vin = unidecode(str(iVin[i].value)).upper().strip()
    appellation = unidecode(str(iAppellation[i].value)).upper()
    if appellation not in exclude_appellation :
        if couleur == 'BLANC' and ' BL' not in vin :
            vin = vin+' BLANC'      
        
    millesime = iMillesime[i].value

    if iPrix[i].value is None :
        prix = 0
    else:
        prix = iPrix[i].value

    #Reconnaissance du format de bouteille, on cherche uniquement les caractères numérique
    iSize = str(iFormatb[i].value)
    if iSize is not None :
        if iSize == '375' :
            formatb = 'DE'
        elif iSize == '500':
            formatb = 'CL'
        elif iSize == '750' or iSize == '700':
            formatb = 'BO'
        elif iSize == '1000':
            formatb = 'L'
        elif iSize == '1500' or iSize == '1,5':
            formatb = 'MG'
        elif iSize == '3000' or iSize == '3':
            formatb = 'DM'
        elif iSize == '5000' or iSize == '5' :
            formatb = 'JE'
        elif iSize == '6000' or iSize == '6' or iSize == '540' or iSize == 'IMPERIALE':
            formatb = 'IM'
        elif iSize == '9000' or iSize == '9':
            formatb = 'SA'
        elif iSize == '12000' or iSize == '12':
            formatb = 'BA'
        elif iSize == '15000' or iSize == '15' or 'NABU' in iSize :
            formatb = 'NA'
        elif iSize == '18000' or iSize == '18':
            formatb = 'ME'
        elif iSize == '27000' or iSize == '27':
            formatb = 'BABY'
        else:
            formatb = iSize
    else:
        formatb == iSize
        

    if iQte[i].value is None :
        quantite = 0
    else:
        quantite = iQte[i].value

        rec_cond =str(iCond[i].value).split(' ')[0].replace('0','').replace('CB','CBO').replace('CT','CC')
        rec_cond = re.findall('[C][BC][O]?[0-9]?[0-9]?',rec_cond)
        if rec_cond :
            conditionnement = rec_cond[0]
        else:
            conditionnement = 'UNITE'


        if iCom[i].value is None :
            commentaire = ''
        else:
            commentaire = 'Stock Phy : '+str(iCom[i].value)
                

        c1 = ws.cell(row = i, column = 1)
        c1.value = vin

        c2 = ws.cell(row = i, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = i, column = 3)
        c3.value = formatb

        c4 = ws.cell(row = i, column = 4)
        c4.value = prix

        c5 = ws.cell(row = i, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = i, column = 6)
        c6.value = conditionnement

        c7 = ws.cell(row = i, column = 7)
        c7.value = commentaire      

        if i  == 1 :
            i = i
            #print(ws.cell(row = i, column = 1).value)
        else:
            if ws.cell(row = i, column = 1).value == ws.cell(row = i-1, column = 1).value and ws.cell(row = i, column = 2).value == ws.cell(row = i-1, column = 2).value and ws.cell(row = i, column = 3).value == ws.cell(row = i-1, column = 3).value :
                while ws.cell(row = i, column = 6).value == ws.cell(row = i-1, column = 6).value :
                    c6.value = random.choice(Dict_cond)
                    c7.value = 'ATTN : VERIF CDT - DOUBLON LIGNE PRECEDENTE - '+c7.value

#Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)

       