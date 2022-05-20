import csv
import re
import glob
from unidecode import unidecode
from collections import Counter
import openpyxl
from openpyxl import Workbook


for filename in glob.glob('*.csv'):
    with open(filename, newline='', encoding='utf-8') as monFichierEntre:
        reader = csv.reader(monFichierEntre, delimiter=';', doublequote=False)
        nom_sortie='new_file.csv'
        with open(nom_sortie,'w',newline='', encoding='utf-8') as monFichierSortie:
            writer = csv.writer(monFichierSortie, delimiter=';', quoting=csv.QUOTE_NONE, escapechar=';', quotechar='')

            enTete = ['Numero Bordereaux','Statut','Nom Acheteur','Vendeur','Date','Vin','Millesime','Prix','Quantite','Total courtage']
            writer.writerow(enTete)

            for row in reader :
              if 'Bordereau' in row[0] :
                bEcrire = 0
              else:
                bEcrire = 1
                bordereau = row[0]
                statut = unidecode(row[10])
                acheteur = row[1]
                vendeur = row[3]
                vin = row[5]
                millesime = row[4]
                quantite = row[6]
                prix = unidecode(row[7]).replace('EUR','')
                date = row[9]
                
                i = 0
         
                if ',' in millesime :
                  vin = vin.split(',')
                  millesime = millesime.split(',')
                  quantite = quantite.split(',')
                  prix = prix.split(',')
                  Ln = len(millesime)
                  while i < Ln :
                    p_qte = re.findall('[0-9][0-9]?[0-9]?[0-9]?[0-9]?',quantite[i-1])[0]
                    total_courtage = (int(p_qte)*float(prix[i-1]))*0.02 
                    row = [bordereau, statut, acheteur, vendeur, date, vin[i-1],millesime[i-1],prix[i-1],quantite[i-1],total_courtage]
                    writer.writerow(row)
                    i = i+1
                else:
                  p_qte = re.findall('[0-9][0-9]?[0-9]?[0-9]?[0-9]?',quantite)[0]
                  total_courtage = (int(p_qte)*float(prix))*0.02 
                  row = [bordereau, statut, acheteur,vendeur, date, vin, millesime,prix,quantite,total_courtage]
                  writer.writerow(row)

wb = Workbook()
ws = wb.active
with open('new_file.csv', 'r') as f:
  reader = csv.reader(f, delimiter = ';', quotechar='|')
  for row in reader:
    ws.append(row)
wb.save('csv_converted.xlsx')

# #########################################################################
# # On load le le tarif au format xlsx
for filename in glob.glob('csv_converted.xlsx'):
    wb2 = openpyxl.load_workbook(filename)

# # On prépare un nouveau workbook
final_wb = Workbook()
dest_filename = 'sortie_EXTRACT_BORDEREAUX.xlsx'
ws2 = final_wb.active
ws2.title = "Jolie Extraction Bordereaux"


# # Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iBordereau = sheet['A']
iStatut = sheet['B']
iAcheteur = sheet['C']
iVendeur = sheet['D']
iDate = sheet['E']
iVin = sheet['F']
iMillesime = sheet['G']
iPrix = sheet['H']
iQte_formatb = sheet['I']
iCourtage = sheet['J']

row_count = sheet.max_row
column_count = sheet.max_column

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if iVin[i].value is None :
        continue
    else:
        num_bordereau = iBordereau[i].value

        statut = iStatut[i].value
        
        acheteur = iAcheteur[i].value

        vendeur = iVendeur[i].value

        vin = unidecode(str(iVin[i].value).upper())

        date = iDate[i].value
        
        millesime = iMillesime[i].value

        prix = iPrix[i].value

        iSize = str(iQte_formatb[i].value)
        formatb = re.findall('[A-Z][A-Z]',iSize)[0]

        quantite = iSize.replace(formatb,'').replace('()','').strip()

        courtage = (iCourtage[i].value)


        # On affecte les valeurs dans l'ordre des colonnes voulues.
        c1 = ws2.cell(row = 1, column = 1)
        c1.value = 'num_bordereau'

        c2 = ws2.cell(row = 1, column = 2)
        c2.value = 'Statut'

        c3 = ws2.cell(row = 1, column = 3)
        c3.value = 'acheteur'

        c4 = ws2.cell(row = 1, column = 4)
        c4.value = 'Vendeur'

        c5 = ws2.cell(row = 1, column = 5)
        c5.value = 'vin'

        c6 = ws2.cell(row = 1, column = 6)
        c6.value = 'Date'

        c7 = ws2.cell(row = 1, column = 7)
        c7.value = 'millesime'

        c8 = ws2.cell(row = 1, column = 8)
        c8.value = 'prix'

        c9 = ws2.cell(row = 1, column = 9)
        c9.value = 'formatb'

        c10 = ws2.cell(row = 1, column = 10)
        c10.value = 'quantite'

        c11 = ws2.cell(row = 1, column = 11)
        c11.value = 'Total courtage'

        c1 = ws2.cell(row = i, column = 1)
        c1.value = num_bordereau

        c2 = ws2.cell(row = i, column = 2)
        c2.value = statut

        c3 = ws2.cell(row = i, column = 3)
        c3.value = acheteur

        c4 = ws2.cell(row = i, column = 4)
        c4.value = vendeur

        c5 = ws2.cell(row = i, column = 5)
        c5.value = vin

        c6 = ws2.cell(row = i, column = 6)
        c6.value = date

        c7 = ws2.cell(row = i, column = 7)
        c7.value = millesime

        c8 = ws2.cell(row = i, column = 8)
        c8.value = prix

        c9 = ws2.cell(row = i, column = 9)
        c9.value = formatb

        c10 = ws2.cell(row = i, column = 10)
        c10.value = quantite

        c11 = ws2.cell(row = i, column = 11)
        c11.value = courtage

# # Fonction de suppression des lignes vides d'un workbook
# index_row = []
# for n in range(1, ws2.max_row):
#     if ws2.cell(n, 1).value is None:
#         index_row.append(n)

# for row_del in range(len(index_row)):
#     ws2.delete_rows(idx=index_row[row_del], amount=1)
#     index_row = list(map(lambda k: k -1, index_row))

final_wb.save(dest_filename)