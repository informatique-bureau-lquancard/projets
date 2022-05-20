## coding: utf-8
from cmath import e
import csv
import glob
import re
import io
import time
from unidecode import unidecode

import openpyxl
from openpyxl import Workbook

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

#problème d'onglet seulement 1 prie en compte

def is_float(n):
    try:
        float(n)   # Type-casting the string to `float`.
                   # If string is not a valid `float`, 
                   # it'll raise `ValueError` exception
    except ValueError:
        return False
    return True

def is_int(n):
    try:
        int(n)   # Type-casting the string to `float`.
                   # If string is not a valid `float`, 
                   # it'll raise `ValueError` exception
    except ValueError:
        return False
    return True

def a():
    wb = Workbook()

    # On load le le tarif au format xlsx
    for filename in glob.glob('*.xlsx'):
        wb = openpyxl.load_workbook(filename)

    # On prépare un nouveau workbook
    wb2 = Workbook()
    
    dest_filename = 'sortie_TWMOFF.xlsx'

    ws = wb2.active
    ws.title = "TWMOFF"

    # Lecture du workbook d'origine
    tab_sheet_name_ok = [' 0.75 - Wine', 'Beyond Bordeaux', 'Special Sizes', 'Unit Bottles']

    erreurs = []

    row_count = 0

    l_avec_erreurs_1 = 0
    l_avec_erreurs_2 = 0

    l_sans_erreurs_1 = 0
    l_sans_erreurs_2 = 0

    index = 0

    # try:
    for sheet in wb:

        if(sheet.title not in tab_sheet_name_ok):
            #print(f"sheet not OK : {sheet.title}")
            continue
            
        #print(f"sheet OK : {sheet.title}")

        iVin = sheet['A']
        iMillesime = sheet['D']
        
        if(sheet.title == tab_sheet_name_ok[0]):
            iFormatb = sheet['G']
            iPrix = sheet['H']

        elif(sheet.title == tab_sheet_name_ok[1]):
            iFormatb = sheet['H']
            iPrix = sheet['I']
        
        elif(sheet.title == tab_sheet_name_ok[2]):
            iFormatb = sheet['G']
            iPrix = sheet['I']
        
        elif(sheet.title == tab_sheet_name_ok[3]):
            iFormatb = sheet['G']
            iPrix = sheet['H']
        else:
            erreurs.append(f"Erreurs : {sheet.title}")
        

        iQte = sheet['F']
        iCond = ''

        iAppellation = sheet['B']

        for index_cellule in range(sheet.max_row):
            print("index 1 : "+str(index))

            #prix
            fPrix=iPrix[index_cellule].value
            fPrix=repr(iPrix[index_cellule].value)

            fPrix=fPrix.replace("'","")
            fPrix=fPrix.replace("\\","")
            fPrix=fPrix.replace("u202f","")

            prix=unidecode(fPrix)
            prix=prix.replace(' EUR','')

            #quantite
            quantite = str(iQte[index_cellule].value).replace('<','').replace('>','').replace('on request','0')
            quantite = quantite.replace(',','.')
            quantite = quantite.replace('+','')
            quantite = quantite.replace('-','')
            quantite = quantite.replace('>','')
            quantite = quantite.replace('*','0')
            quantite = quantite.replace(' ','')

            # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
            if ((iVin[index_cellule].value is None) or (iFormatb[index_cellule].value is None) or (iPrix[index_cellule].value is None) or (len(str(iFormatb[index_cellule].value)) == 0) or (len(str(iPrix[index_cellule].value)))) == 0:
                l_avec_erreurs_1 = l_avec_erreurs_1 + 1

                index = index + 1
                print("index 2 : "+str(index))
                continue

            l_sans_erreurs_1 = l_sans_erreurs_1 + 1

            if(not is_float(prix)):
                l_avec_erreurs_2 = l_avec_erreurs_2 + 1

                index = index + 1
                print("index 2 : "+str(index))
                continue

            l_sans_erreurs_2 = l_sans_erreurs_2 + 1

            quantite = int(quantite)

            #prix
            prix = float(prix)

            #vin
            vin = unidecode(
                str(iVin[index_cellule].value
                ).upper())

            if(sheet.title == tab_sheet_name_ok[1]):
                vin = vin + unidecode(
                    str(iAppellation[index_cellule].value
                    ).upper())

            #millesime
            millesime = iMillesime[index_cellule].value

            #prix
            prix = iPrix[index_cellule].value

            #taille bouteille
            formatB = str(iFormatb[index_cellule].value)

            if formatB == 'Btl' or formatB == 'Btl50' :
                formatB='BO'
            elif formatB=='DemiBt':
                formatB='DE'
            elif formatB=='Mgn':
                formatB='MG'
            elif formatB=='DMgn':
                formatB='DM'
            elif formatB=='Imp':
                formatB='IM'
            elif formatB=='Jero':
                formatB='JE'
            else:
                formatB = ft.formaterFormatBouteille(formatB)  

            #conditionnement
            if(sheet.title == tab_sheet_name_ok[0] or sheet.title == tab_sheet_name_ok[0]):
                conditionnement = 'CBO12'

            elif(sheet.title == tab_sheet_name_ok[2]):
                if(formatB == 'MG'):
                    conditionnement = 'CBO6'
                else:
                    conditionnement = 'CBO1'

            elif(sheet.title == tab_sheet_name_ok[3]):
                conditionnement = 'UNITE'
            
            else:
                conditionnement = ft.formaterConditionnement(formatB, quantite)
            
            #Commentaire
            commentaire='VERIF CDT - Qte = '+ str(quantite)

            # On affecte les valeurs dans l'ordre des colonnes voulues.
            c1 = ws.cell(row = index, column = 1)
            c1.value = vin
            print("index : "+str(index)+" vin : "+vin)
            c2 = ws.cell(row = index, column = 2)
            c2.value = millesime
            # print("index_cellule : "+str(index_cellule)+" vin : "+vin)
            c3 = ws.cell(row = index, column = 3)
            c3.value = formatB
            # print("index_cellule : "+str(index_cellule)+" vin : "+vin)
            c4 = ws.cell(row = index, column = 4)
            c4.value = str(prix).replace('.', ',')
            # print("index_cellule : "+str(index_cellule)+" vin : "+vin)
            c5 = ws.cell(row = index, column = 5)
            c5.value = quantite
            # print("index_cellule : "+str(index_cellule)+" vin : "+vin)
            c6 = ws.cell(row = index, column = 6)
            c6.value = conditionnement
            # print("index_cellule : "+str(index_cellule)+" vin : "+vin)
            c7 = ws.cell(row = index, column = 7)
            c7.value = commentaire
            # print("index_cellule : "+str(index_cellule)+" vin : "+vin)

            index = index + 1
            print("index 2 : "+str(index))

    index_row = []

    l_sans_erreurs = 0
    l_avec_erreurs = 0

    print(f"row_count : {row_count}")

    for n in range(ws.max_row):
        l_sans_erreurs = l_sans_erreurs + 1

        if ws.cell(n+1, 1).value is None:
            index_row.append(n+1)
            l_avec_erreurs = l_avec_erreurs + 1
                

    print(f"Ligne sans erreurs : {l_sans_erreurs} Ligne avec erreurs : {l_avec_erreurs}")
    print(f"Ligne avec erreurs 1 : {l_avec_erreurs_1} Ligne avec erreurs 2 : {l_avec_erreurs_2}")
    print(f"Ligne sans erreurs 1 : {l_sans_erreurs_1} Ligne sans erreurs 2 : {l_sans_erreurs_2}")

    #Supprime les lignes vides
    for row_del in range(len(index_row)):
        ws.delete_rows(idx=index_row[row_del], amount=1)
        index_row = list(map(lambda k: k -1, index_row))


    erreurs.__str__

    print("traitement fini")

    wb2.save(dest_filename)

a()