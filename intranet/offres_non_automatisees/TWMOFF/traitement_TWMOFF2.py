## coding: utf-8
import csv
import glob
import re
import io
import time
from unidecode import unidecode

import openpyxl
from openpyxl import Workbook

# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_TWMOFF.xlsx'
ws = wb2.active
ws.title = "TWMOFF"


# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.

for filename in glob.glob('*.csv'):
    with open(filename, newline='', encoding='utf-8') as monFichierEntre:
        reader = csv.reader(monFichierEntre, delimiter=';', doublequote=False)
        nom_sortie='sortie_TWMOFF.csv'
        with open(nom_sortie,'w',newline='', encoding='utf-8') as monFichierSortie:
            writer = csv.writer(monFichierSortie, delimiter=';', quoting=csv.QUOTE_NONE, escapechar='|', quotechar='')

            index_chateau = 0
            vCond = 0
            trou='' # besoin de le faire une seule fois
            
            for row in reader:
                bEcrire=0
                #il ne faut rien faire sur les lignes d'en-tête
                if row[0]=="":
                    bEcrire=0
                elif 'Prix H.T' in row[0]:
                    bEcrire=0
                elif 'PRICE LIST' in row[0]:
                    bEcrire=0
                else:
                    # A partir de là, on traite les lignes et ne surtout pas oublier les vins blancs !
                    bEcrire=1
                    chateau=row[0]
                    

                    # annee
                    if row[2]=='NV':
                        annee=time.strftime('%Y')
                    else:
                        annee=row[2]

                    # Format de bouteille
                    if row[4]=='Btl':
                        formatb='BO'
                    elif row[4]=='DemiBt':
                        formatb='DE'
                    elif row[4]=='Mgn':
                        formatb='MG'
                    elif row[4]=='DMgn':
                        formatb='DM'
                    elif row[4]=='Imp':
                        formatb='IM'
                    elif row[4]=='Jero':
                        formatb='JE'
                    else:
                        formatb=row[4]

                    # Prix , on nettoie la chaine prix
                    fPrix=repr(row[5])
                    fPrix=fPrix.replace("'","")
                    fPrix=fPrix.replace("\\","")
                    fPrix=fPrix.replace("u202f","")
                    prix=unidecode(fPrix)
                    prix=prix.replace(' EUR','')

                    #quantite
                    value=row[3]
                    value=value.replace(',','.')
                    value=value.replace(' ','')
                    value=value.replace('+','')
                    value=value.replace('-','')
                    value=value.replace('>','')
                    quantite='0'


                    if '-' in row[3]:
                        qMax=row[3]
                        qMax=qMax.replace(' ','')
                        qInd=qMax.find('-')
                    else:
                        qInd=0

                    for index in range(qInd, len(value)):
                        if value[index] in '0123456789.':
                            quantite=value[qInd:]

                    #on applique le conditionnement par defaut selon le format bouteille et la quantité
                    conditionnement=''
                    if formatb=='DE' and int(quantite)<24:
                        conditionnement='UNITE'
                    elif formatb=='DE' and int(quantite)>=24:
                        conditionnement='CBO24DE'
                    elif formatb=='BO' and int(quantite)<12:
                        conditionnement='UNITE'
                    elif formatb=='BO' and int(quantite)>=12:
                        conditionnement='CBO12'
                    elif formatb=='MG' and int(quantite)<6:
                        conditionnement='UNITE'
                    elif formatb=='MG' and int(quantite)>=6:
                        conditionnement='CBO6'
                    elif formatb=='DM' and int(quantite)<3:
                        conditionnement='UNITE'
                    elif formatb=='DM' and int(quantite)>=3:
                        conditionnement='CBO3'
                    elif formatb=='JE' and int(quantite)>=1:
                        conditionnement='CBO1'
                    elif formatb=='IM' and int(quantite)>=1:
                        conditionnement='CBO1'
                    elif formatb=='CL' and int(quantite)<12:
                        conditionnement='UNITE'
                    elif formatb=='CL' and int(quantite)>=12:
                        conditionnement='CBO12'
                    else:
                        conditionnement='UNITE'

                    #Commentaire
                    commentaire='VERIF CDT - Qte = '+row[3]

                # écriture de la ligne
                # On fabrique la nouvelle ligne dans l'ordre voulu
                if bEcrire==1 :
                    newRow=[chateau,annee,formatb,prix,quantite,conditionnement,commentaire,trou,trou,trou]
                    writer.writerow(newRow)
                    
                    
            monFichierEntre.close()