# coding: utf-8
import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook
import random

Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC']

# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_PIONSA.xlsx'
ws = wb2.active
ws.title = "PIONSA (っ•́｡•́)♪♬ "


# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iProducteur = sheet['B']
iVin = sheet['C']
iMillesime = sheet['D']
iFormatb = sheet['E']
iPrix = sheet['G']
iQte = sheet['F']
iCom = sheet['J']

row_count = sheet.max_row
column_count = sheet.max_column

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if iVin[i].value is None or iVin[i].value == 'Wine' :
        continue
    else:  
        vin = unidecode(str(iProducteur[i].value).upper())+' '+unidecode(str(iVin[i].value).upper())
        
        millesime = iMillesime[i].value

        prix = iPrix[i].value

        # Reconnaissance du format de bouteille, on cherche uniquement les caractères numérique
        rec_formatb = str(iFormatb[i].value).replace('cl','')
        if 'x' in rec_formatb :
            iSize = rec_formatb.split('x')[1]
        else:
            iSize = rec_formatb

        if iSize :
            #iSize = str(iSize[0])
            if iSize == '37,5' :
                formatb = 'DE'
            elif iSize == '50':
                formatb = 'CL'
            elif iSize == '75' or iSize == '70':
                formatb = 'BO'
            elif iSize == '100':
                formatb = 'L'
            elif iSize == '150' or iSize == '1,5':
                formatb = 'MG'
            elif iSize == '300' or iSize == '3':
                formatb = 'DM'
            elif iSize == '500' or iSize == '5' :
                formatb = 'JE'
            elif iSize == '600' or iSize == '6' or iSize == '540':
                formatb = 'IM'
            elif iSize == '900' or iSize == '9':
                formatb = 'SA'
            elif iSize == '1200' or iSize == '12':
                formatb = 'BA'
            elif iSize == '1500' or iSize == '15':
                formatb = 'NA'
            elif iSize == '1800' or iSize == '18':
                formatb = 'ME'
            elif iSize == '2700' or iSize == '27':
                formatb = 'BABY'
            else:
                formatb = iSize
        else:
            formatb == rec_formatb
        

        quantite = str(iQte[i].value)
        #quantite = int(quantite)

        commentaire = unidecode(str(iCom[i].value))
         
        if 'ASSORT' in vin :
            conditionnement = 'COLLEC'
            vin_assort = vin[int(vin.find('ASSORT')):]
            vin_assort = vin_assort[int(vin_assort.find(' ')):]
            vin = vin.replace(vin_assort, '')
            if iCom[i].value is None :
                commentaire = vin_assort
            else:
                commentaire = vin_assort+' - '+unidecode(str(iCom[i].value))      
        else:
            if 'OWC' in commentaire or 'OC' in commentaire :
                rec_cond = commentaire.replace('*','').replace('OWC','CBO').replace('OC','CC')
                conditionnement = re.findall('[C][CB][O]?[0-9]?[0-9]?',rec_cond)[0]
                if iCom[i].value is not None :
                    commentaire = unidecode(str(iCom[i].value))
                else:
                    commentaire = ''
            else:
                conditionnement = 'UNITE'
                if iCom[i].value is not None :
                    commentaire = unidecode(str(iCom[i].value))
                else:
                    commentaire = ''


        c1 = ws.cell(row = i, column = 1)
        c1.value = vin

        c2 = ws.cell(row = i, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = i, column = 3)
        c3.value = formatb

        c4 = ws.cell(row = i, column = 4)
        c4.value = prix

        c5 = ws.cell(row = i, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = i, column = 6)
        c6.value = conditionnement

        c7 = ws.cell(row = i, column = 7)
        c7.value = commentaire
        
        # Procédure de test si la ligne précédente est identique et modification de la valeur conditionnement en mode random
        if  ws.cell(row = i, column = 1).value == ws.cell(row = i-1, column = 1).value and ws.cell(row = i, column = 2).value == ws.cell(row = i-1, column = 2).value and ws.cell(row = i, column = 3).value == ws.cell(row = i-1, column = 3).value :
            while ws.cell(row = i, column = 6).value == ws.cell(row = i-1, column = 6).value :
                c6.value = random.choice(Dict_cond)
                c7.value = 'ATTN : VERIF CDT - DOUBLON LIGNE PRECEDENTE - '+c7.value

#Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)