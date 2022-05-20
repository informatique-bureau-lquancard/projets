import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook

# coding: utf-8

# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_CRUCOL.xlsx'
ws = wb2.active
ws.title = "CRUCOL"


# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['A']
iAppellation = sheet['B']
iMillesime = sheet['C']
iFormatb = sheet['F']
iPrix = sheet['I']
iQte = sheet['D']
iCond = sheet['G']
iRegie = sheet['J']
iCom = sheet['H']

row_count = sheet.max_row
column_count = sheet.max_column

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if iVin[i].value is None :
        continue
    elif iVin[i].value is not None and iAppellation[i].value is None :
        region = unidecode(str(iVin[i].value)).upper()
        continue
    elif str(iMillesime[i].value) == 'MILLESIME':
        continue
    else:
        if region == 'BORDEAUX' : 
            vin = unidecode(str(iVin[i].value).upper())
        else:
            vin = unidecode(str(iVin[i].value).upper())+' '+unidecode(str(iAppellation[i].value).upper())
        
        millesime = str(iMillesime[i].value).replace('NM','NV')

        prix = iPrix[i].value

        formatb = str(iFormatb[i].value).upper()
        if formatb == '75 CL':
            formatb = 'BO'
        elif formatb == '150 CL':
            formatb = 'MG'
        elif formatb == '300 CL':
            formatb = 'DM'
        elif formatb == '500 CL':
            formatb = 'JE'
        elif formatb == '600 CL':
            formatb = 'IM'
        elif formatb == '900 CL':
            formatb = 'SA'
        elif formatb == '1200 CL':
            formatb = 'BA'
        elif formatb == '1500 CL':
            formatb = 'NA'
        elif formatb == '1800 CL':
            formatb = 'ME'
        elif formatb == '225 CL':
            formatb = 'MJ'
        elif formatb == '450 CL':
            formatb = 'RE'
        elif formatb == '2700 CL':
            formatb = 'BABY'

        quantite = iQte[i].value
        
        conditionnement = str(iCond[i].value).replace('UNIT','UNITE').replace('CTO','CC').replace('CO','CC')
        
        if iCom[i].value is None :
            commentaire = 'Regie : '+str(iRegie[i].value)
        else:
            commentaire = 'Comment : ' + unidecode(str(iCom[i].value))+' - Regie : '+str(iRegie[i].value)

# Fonction Attribution des valeurs dans les bonnes colonnes pour le nouveau workbook
        c1 = ws.cell(row = i, column = 1)
        c1.value = vin

        c2 = ws.cell(row = i, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = i, column = 3)
        c3.value = formatb

        c4 = ws.cell(row = i, column = 4)
        c4.value = prix

        c5 = ws.cell(row = i, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = i, column = 6)
        c6.value = conditionnement

        c7 = ws.cell(row = i, column = 7)
        c7.value = commentaire

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

# Ecriture du nouveau workbook
wb2.save(dest_filename)



   

