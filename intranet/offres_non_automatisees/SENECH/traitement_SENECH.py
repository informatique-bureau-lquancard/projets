# coding: utf-8
import csv
import glob
import codecs
from time import sleep
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook
from unidecode import unidecode

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

nomProfil : str = "SENECH"
extensionDebut : str = ".xlsx"
extensionFin : str = ".xlsx"

# Conversion du fichier xls en xlsx
for xls_file in glob.glob('*.xls'):
    reader = pd.read_excel(xls_file)
    reader.to_excel(nomProfil + extensionDebut, index = False)

# sleep(1)

# On load le le tarif au format xlsx dans un workbook
for filename in glob.glob('*' + extensionDebut):
    wb = openpyxl.load_workbook(filename)

    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename = 'sortie_' + nomProfil + extensionFin
    ws = wb2.active
    ws.title = nomProfil

    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    iVin = sheet['A']
    iMillesime = sheet['C']
    iFormatB = sheet['D']
    iQte = sheet['E']
    iPrix = sheet['F']
    iCond = sheet['G']

    row_count = sheet.max_row
    column_count = sheet.max_column

    j : int = 1

    for index_ligne in range (1, row_count) :

        vin : str = ft.formaterVin( "", str( iVin[index_ligne].value ), "" )
        prix : str = ft.formaterPrix( str( iPrix[index_ligne].value ) )

        if ( iVin[index_ligne].value is None) or ( len( iVin[ index_ligne ].value ) < 1 or ( prix is None ) or ( len( prix ) < 1 ) ) :
            continue

        #### Millesime ####
        millesime : str = ft.formaterAnnee( str( iMillesime[index_ligne].value ) )

        commentaire : str = ""

        if( len(millesime) > 4 ):
            commentaire = millesime
            millesime = "NV"

        #### Format bouteille ####

        formatB : str = str( iFormatB[index_ligne].value )
        formatB = ft.formaterFormatBouteille( formatB )

        #### Quantite ####
        quantite : str = ft.formaterQuantite( str( iQte[index_ligne].value ) )

        #### Conditionnement ####
        conditionnement : str = str( iCond[index_ligne].value )
        conditionnement = ft.formaterConditionnement( formatB, int(quantite), conditionnement, vin )

        [vin, commentairesBis] = ft.testCommentaireParenthese(vin)

        if( commentairesBis != "" ):
            commentaire = commentaire + " " + commentairesBis

        #On affecte les valeurs dans l'ordre des colonnes voulues.
        ft.affectationLignes(j, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaire)
        j += 1

wb2.save(dest_filename)
