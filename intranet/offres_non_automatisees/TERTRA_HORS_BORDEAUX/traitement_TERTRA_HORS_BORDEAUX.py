import csv
import glob
from unidecode import unidecode
import openpyxl
from openpyxl import Workbook

import glob

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

import pandas as pd

nomProfil : str = 'TERTRA_HORS_BORDEAUX'
extensionDebut : str = '.xlsx'
extensionFin_bis : str = '.xlsx'
extensionFin : str = '.csv'

# coding: utf-8

# On load le le tarif au format xlsx
for filename in glob.glob('*' + extensionDebut):

    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename_bis = 'sortie_bis_' + nomProfil + extensionFin_bis
dest_filename = 'sortie_' + nomProfil + extensionFin
ws = wb2.active
ws.title = nomProfil

# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['M']
iMillesime = sheet['K']
iFormatB = sheet['L']
iPrix = sheet['N']
iQte_UK = sheet['H']
iQte_FR = sheet['I']

row_count = sheet.max_row
column_count = sheet.max_column

index_sortie : int = 1

for i  in range (1, row_count):
            
    vin : str = unidecode(str(iVin[i].value)).upper()
    prix : str = ft.formaterPrix(str(iPrix[i].value))

    print("prix : " + prix)

    designation_prix = ["Price"]

    if( ft.bLigneIncorrecte(prix, designation_prix, vin)):
        print("ko : ")
        continue;

    print("vin : "+ vin + " prix : " + prix)

    millesime : str = ft.formaterAnnee(str(iMillesime[i].value))

    formatB : str = ft.formaterFormatBouteille(str(iFormatB[i].value))

    quantite : str = ft.formaterQuantite(str(iQte_UK[i].value + iQte_FR[i].value))

    print("formatB  : " + formatB + " quantite : " + quantite)

    conditionnement : str = ft.conditionnementParDefaut(formatB, int(quantite))

    commentaire : str = "Stock FR + UK"

    # Fonction Attribution des valeurs dans les bonnes colonnes pour le nouveau workbook
    ft.affectationLignes(index_sortie, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaire)

    index_sortie += 1

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

# Ecriture du nouveau workbook
wb2.save(dest_filename_bis)

read_file = pd.read_excel(dest_filename_bis)
read_file.to_csv(dest_filename, index = False, encoding="utf-8", sep = ';')







   

