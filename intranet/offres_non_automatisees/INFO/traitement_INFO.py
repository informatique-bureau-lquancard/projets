import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook
import pandas as pd

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

# coding: utf-8

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_INFO.xlsx'
ws = wb2.active
ws.title = "info"

# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)


    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    iVin : tuple = sheet['A']
    iMillesime : tuple = sheet['B']
    formatB : tuple = 'BO'
    iPrix : tuple = sheet['G']
    iQte : tuple = sheet['E']
    iCond : tuple = sheet['F']
    iDate_offre : tuple = sheet['I']  #validité
    iRegie : tuple = sheet['J']
    iType_vendeur : tuple = sheet['H']
    iCom : tuple = sheet['N']

    j = 1

    row_count = sheet.max_row
    column_count = sheet.max_column

    for i  in range (1, row_count):
        # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
        if iVin[i].value is None :
            continue

        vin : str = unidecode(str(iVin[i].value).upper())
        
        millesime = iMillesime[i].value

        prix = str(iPrix[i].value).replace('.',',')

        if iCond[i].value == 'blles' :
            conditionnement = 'UNITE'
            quantite = iQte[i].value
        else:
            conditionnement = unidecode(str(iCond[i].value).upper())
            conditionnement = conditionnement.replace('CB','CBO')
            conditionnement = conditionnement.replace('CT','CC')
            conditionnement = conditionnement.replace(' A PLAT','PLA')
            uQte = re.findall('\d+',conditionnement)
            if uQte :
                quantite = int(iQte[i].value) * int(uQte[0])
            else:
                quantite = iQte[i].value
        
        if iCom[i].value is None :
            commentaire = 'Date Offre : '+ unidecode(str(iDate_offre[i].value)) + ' - Vendeur : ' + unidecode(str(iType_vendeur[i].value))
        else:
            commentaire = 'Date Offre : '+ unidecode(str(iDate_offre[i].value)) + ' - Vendeur : ' + unidecode(str(iType_vendeur[i].value)) + ' - Comment : ' + unidecode(str(iCom[i].value))

        ft.affectationLignes2(j, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaire, '')

        j = j +1

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)
csv_converter = pd.read_excel(dest_filename)
csv_converter.to_csv('sortie_INFO.csv', index = False, encoding="utf-8", sep = ';')



   

