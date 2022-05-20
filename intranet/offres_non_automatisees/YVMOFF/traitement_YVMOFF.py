# coding: utf-8
import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
from Fonction_tarifs import formatterConditionnementParDefaut
import Fonction_tarifs


# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_YVMOFF.xlsx'
ws = wb2.active
ws.title = "YVMOFF"


# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['B']
iMillesime = sheet['F']
iFormatb = sheet['L']
iPrix = sheet['O']
iQte = sheet['N']
iCond = sheet['M']
iColor = sheet['C']
iRegion = sheet['A']
iAppellation = sheet['D']

row_count = sheet.max_row
column_count = sheet.max_column

for index_cellule  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if iPrix[index_cellule].value is None :
        continue
    elif iVin[index_cellule].value is None :
        continue
    elif str(iRegion[index_cellule].value) == 'Région' :
        continue
    else:
        if 'BORDEAUX' in str(iRegion[index_cellule].value) :
            vin = unidecode(str(iVin[index_cellule].value).upper())
        else:
            vin = unidecode(str(iVin[index_cellule].value).upper())+' '+str(iAppellation[index_cellule].value).upper()
        
        
        if str(iColor[index_cellule].value) == 'Blanc' and 'BLANC' not in vin:
            vin = vin +' BLANC'

        vin = vin.replace('"','').replace("'","")         
       
        millesime = iMillesime[index_cellule].value

        prix = iPrix[index_cellule].value

        iSize = re.findall('[0-9]?[.]?\d+',str(iFormatb[index_cellule].value).replace(',','.'))

        if iSize :
            iSize = iSize[0].replace('.',',')
        
        formatb = Fonction_tarifs.formatterFormatBouteille2(iSize)
        

        quantite = int(re.findall('\d+',str(iQte[index_cellule].value))[0])
        
        i_cond = str(iCond[index_cellule].value)
        unite_cond = re.findall('\d+', i_cond)
        
        if 'CBO' in i_cond or 'OWC' in i_cond :
            type_cond = 'CBO'
        elif 'CRT' in i_cond :
            type_cond = 'CC'
        else:
            type_cond = 'UNITE'

        if unite_cond :
            conditionnement = type_cond+unite_cond[0]
        else:
            conditionnement = type_cond
        


        # On affecte les valeurs dans l'ordre des colonnes voulues.
        c1 = ws.cell(row = index_cellule, column = 1)
        c1.value = vin

        c2 = ws.cell(row = index_cellule, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = index_cellule, column = 3)
        c3.value = formatb

        c4 = ws.cell(row = index_cellule, column = 4)
        c4.value = prix

        c5 = ws.cell(row = index_cellule, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = index_cellule, column = 6)
        c6.value = conditionnement

        # c7 = ws.cell(row = index_cellule, column = 7)
        # c7.value = commentaire

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)



   

