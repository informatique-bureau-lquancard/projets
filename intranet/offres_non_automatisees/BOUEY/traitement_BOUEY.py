# coding: utf-8
import csv
import glob
import codecs
from unidecode import unidecode
from time import sleep
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

nomProfil = "BOUEY"
extensionDebut = ".xlsx"
extensionFin = ".csv"

# Conversion du fichier xls en xlsx
# for xls_file in glob.glob('*.xls'):
#     reader = pd.read_excel(xls_file)
#     reader.to_excel('BOUEY.xlsx', index = False)

# sleep(1)

# On load le le tarif au format xlsx dans un workbook
for filename in glob.glob('*' + extensionDebut):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename_bis = 'sortie_'+ 'bis_' + nomProfil.upper()+extensionFin
dest_filename = 'sortie_'+ nomProfil.upper()+extensionFin
ws = wb2.active
ws.title = nomProfil

# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['C']
iAppellation = sheet['A']
iMillesime = sheet['E']
iFormatB = sheet['G']
iPrix = sheet['J']
iQte = sheet['K']
iCond = sheet['H']
uCond = sheet['I']
iColor = sheet['B'] 

bordeaux_dict = ['PAUILLAC','HAUT MEDOC','ST EMILION GCC','AUSONE', 'MOUTON', 'CHEVAL', 'MISSION', 'YQUEM', 'PETRUS', 'HTBRION', 'LAFITE', 'MARGAUX','PESSAC','ST JULIEN','POMEROL','ST ESTEPHE','MOULIS']

#iRegie = sheet['J']
#iType_vendeur = sheet['H']
iCom = sheet['F']

row_count = sheet.max_row
column_count = sheet.max_column

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    millesime = iMillesime[i].value

    prix : str = ft.formaterPrix(str(iPrix[i].value))

    appellation = unidecode(str(iAppellation[i].value)).upper()

    quantite = ft.formaterQuantite(str(iQte[i].value))

    color = str(iColor[i].value).upper()

    vin = unidecode(str(iVin[i].value)).replace('"','').upper()

    designation_prix = ["PrixTarif"]
    
    if(ft.bLigneIncorrecte(prix, designation_prix, vin)):
        continue

    if 'BLANC' in color and 'BLANC' not in vin:
        vin = vin+' BLANC'

    if millesime is None :
        millesime = 'NV'
    else:
        millesime = iMillesime[i].value
    

    formatB = ft.formaterFormatBouteille(str(iFormatB[i].value))
    
    rec_cond = str(iCond[i].value).upper()
    if rec_cond == "CSE" :
        conditionnement = "CBO"+str(uCond[i].value)
    elif rec_cond == "CRT" :
        conditionnement = "CC"+str(uCond[i].value)
    else:
        conditionnement = rec_cond

    rec_com = str(iCom[i].value)
    if rec_cond == 'NONE':
        if rec_com == 'None' :
            commentaire = ''
        else:
            commentaire = rec_com
    else:
        if rec_com == 'None' :
            commentaire = rec_cond
        else:
            commentaire = rec_com+' '+rec_cond


    #On affecte les valeurs dans l'ordre des colonnes voulues.
    c1 = ws.cell(row = i, column = 1)
    c1.value = vin

    c2 = ws.cell(row = i, column = 2)
    c2.value = millesime

    c3 = ws.cell(row = i, column = 3)
    c3.value = formatB
    c4 = ws.cell(row = i, column = 4)
    c4.value = prix
    c5 = ws.cell(row = i, column = 5)
    c5.value = quantite
    c6 = ws.cell(row = i, column = 6)
    c6.value = conditionnement
    c7 = ws.cell(row = i, column = 7)
    c7.value = commentaire

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename_bis)

read_file = pd.read_excel(dest_filename_bis)
read_file.to_csv(dest_filename, index = False, encoding="utf-8", sep = ';')

   

