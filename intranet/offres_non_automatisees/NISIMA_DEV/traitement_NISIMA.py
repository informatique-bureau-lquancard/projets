# coding: utf-8
import csv
import re
from email import header
import glob
from typing import OrderedDict
from numpy import False_
#import codecs
from unidecode import unidecode
from openpyxl import Workbook
import pandas as pd

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
from Fonction_tarifs import formaterConditionnement
import Fonction_tarifs


tab_conditionnement = ["UNITE","CC1","CC2","CC3","CC4","CC6","CC12","CC24","CC24DE","CBO1","CBO2","CBO3","CBO4","CBO6","CBO12","CBO24","CBO24DE", "CBN","CBN1", "CBN2", "CBN3","CBN4", "CBN6", "CBN12", "COLLEC"]

for filename in glob.glob('*.xlsx'):
    reader = pd.read_excel(filename)
    reader.dropna(how = "all", inplace = True)
    reader.to_csv('nisima.csv', index = False, header = False)
    #reader = pd.read_csv('nisima.csv')
    #reader.to_csv('nisima.csv', index = False, header = False)

wb = Workbook()
dest_filename = 'sortie_NISIMA.xlsx'
ws = wb.active
ws.title = "NISIMA ┌(ㆆ㉨ㆆ)ʃ"

with open('nisima.csv', newline='') as csvfile :
    csvreader = csv.DictReader(csvfile)
    for row in csvreader :
        #for key, value in row.items() :  
            
            ### boucle for inutile à l'heure actuelle, mais
            ### au cas ou les colonnes du fichier viendraient à bouger, il faudra faire un test sur celles-ci
            ## 
            ##  ==> oui mais pourquoi ? Puisqu'on utilise DictReader ...

        
        
        prix = str(row['Prix de vente']).replace(' ','')
        if prix == '':
            continue
        else:
            domaine = unidecode(str(row['Domaine'])).upper()
            appellation = unidecode(str(row['Appellation'])).upper()
            millesime = row['Mill.']
            iSize = row['Unité de volume']
            # prix = row['Prix de vente']
            prix_discount = row['Règlement à la facturation']
            quantite = int(row['Nombres de bouteilles'])                 
            iCond = str(row['Conditionnement']).lower()
            Info = unidecode(str(row['Info']))
            
            # Traitement spécifique fichier "Bordeaux"
            vin = domaine
            i_Collec = 0
            if 'CAISSE' in vin or ':' in vin :
                i_Collec = 1
                liste_vin = vin.split(':')
                vin = liste_vin[0]
                comVin = liste_vin[1]

            # Traitement spécifique fichier "hors bordeaux"

            #Traitement commun
            formatb = Fonction_tarifs.formaterFormatBouteille(iSize)   

            # traitement sur le prix pour la lecture
            prix = float(prix)
            prix = str(prix).replace('.',',')   

            iCB = re.findall('[c][a][i][s][s][e]',iCond)
            iCC = re.findall('[c][a][r][t][o][n]',iCond)
            iORI = re.findall('[o][r][i][g][i][n]',iCond)
            iNEU = re.findall('[n][e][u][t][r][e]',iCond)
            iUNI = re.findall('\d+', iCond)

            if iCB :
                type_cond = 'CB'
            elif iCC :
                type_cond = 'CC'
            else:
                type_cond = 'MISSING INFO'
            
            if iORI :
                style_cond = 'O'
            elif iNEU :
                style_cond = 'N'
            else:
                style_cond = ''
            
            if iUNI :
                unite_cond = iUNI[0]
            else:
                unite_cond = ''

            if type_cond == 'CB':
                new_iCond = type_cond+style_cond+unite_cond
            else:
                new_iCond = type_cond+unite_cond

            if new_iCond in tab_conditionnement :
                conditionnement = new_iCond
            else:
                conditionnement = 'UNITE'


            commentaire = Info+' - '+'Rlt a facturation : '+prix_discount
            if new_iCond not in tab_conditionnement :
                commentaire = 'verif cdt : '+new_iCond+' - '+ Info+' - '+'Rlt a facturation : '+prix_discount
            
            if i_Collec == 1 :
                conditionnement = 'COLLEC'
                commentaire += ' '+comVin

        
            c1 = ws.cell(row = csvreader.line_num,column = 1)
            c1.value = vin

            c2 = ws.cell(row = csvreader.line_num, column = 2)
            c2.value = millesime

            c3 = ws.cell(row = csvreader.line_num, column = 3)
            c3.value = formatb

            c4 = ws.cell(row = csvreader.line_num, column = 4)
            c4.value = prix

            c5 = ws.cell(row = csvreader.line_num, column = 5)
            c5.value = quantite

            c6 = ws.cell(row = csvreader.line_num, column = 6)
            c6.value = conditionnement

            c7 = ws.cell(row = csvreader.line_num, column = 7)
            c7.value = commentaire

index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))       

wb.save(dest_filename)