# coding: utf-8
import glob
from operator import index
from unidecode import unidecode
from time import sleep
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook
import random

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs

Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC']

for xls_file in glob.glob('*.xls'):
        
    # Conversion du fichier xls en xlsx pour traitement classique
    reader = pd.read_excel(xls_file, header = 10)
    reader.to_excel('JOEMEY.xlsx', index = False, header = False)

wb = openpyxl.load_workbook('JOEMEY.xlsx')

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_JOEMEY.xlsx'
ws = wb2.active
ws.title = "JOEMEY"

# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['C']
iDomaine_Region = sheet['A']
iMillesime = sheet['B']
iFormatb = sheet['E']
iPrix = sheet['G']
iQte = sheet['F']
iCond = sheet['D']
iCom = sheet['I']

row_count = sheet.max_row
column_count = sheet.max_column

region = ''
domaine = ''
verif_cdt = 0
for index_cellule  in range (1,sheet.max_row):
    
    if iDomaine_Region[index_cellule].value is not None and iVin[index_cellule].value is None :
        region = unidecode(str(iDomaine_Region[index_cellule].value)).upper()
        continue
    
    
    vin = unidecode(str(iVin[index_cellule].value)).upper()
    spacing = re.findall('[ ][ ]+',vin)
                
    if "BORDEAUX" in region :      
        if spacing :
            cut_vin = vin.find(spacing[0])
            vin = vin[:int(cut_vin)]      
        if "BLANC" in str(iDomaine_Region[index_cellule].value).upper():
            vin = vin +' BLANC'
    else:
        if spacing :
            vin = vin.replace(spacing[0],' ')
        if iDomaine_Region[index_cellule].value is not None and iVin[index_cellule].value is not None :
            domaine = unidecode(str(iDomaine_Region[index_cellule].value)).upper()
        
        vin = domaine+' '+vin
        
    vin = vin.strip()
    
    millesime = iMillesime[index_cellule].value

    formatb = Fonction_tarifs.formaterFormatBouteille(str(iFormatb[index_cellule].value).strip())
    
    prix = iPrix[index_cellule].value

    quantite = str(iQte[index_cellule].value).strip().split(' ')
    if quantite :
        quantite = int(quantite[0])
    else:
        quantite = 'NC'

    
    if iCond[index_cellule].value is not None and iCond[index_cellule].value != '':
        #print(iCond[index_cellule].value)
    #if iCond[index_cellule].value is not None or iCond[index_cellule].value != '':
        verif_cdt = 0
        rec_cond = str(iCond[index_cellule].value).upper().replace(' / ','').replace('/','').replace(' ','')
        i_cond = rec_cond
        i_uqte = re.findall('\d+', rec_cond)
        if 'CARTON' in rec_cond :
            i_cond = 'CC'
        elif 'BOIS' in rec_cond :
            i_cond = 'CBO'
        elif 'COFFRET' in rec_cond :
            i_cond = 'COF'
            
        
        if i_uqte :
            conditionnement = i_cond+i_uqte[0]
        else:
            conditionnement = i_cond
    else:
        conditionnement = Fonction_tarifs.conditionnementParDefaut(formatb, quantite)
        verif_cdt = 1

    print(conditionnement)   
    if iCom[index_cellule].value is not None :
        commentaire = unidecode(str(iCom[index_cellule].value))
    else:
        commentaire = ''

    if verif_cdt == 1 :
        commentaire = 'verif cdt - '+commentaire
    
    # print(vin, millesime, formatb, quantite, conditionnement, commentaire, verif_cdt)
        

    #On affecte les valeurs dans l'ordre des colonnes voulues.

    c1 = ws.cell(row = index_cellule, column = 1)
    c1.value = vin

    c2 = ws.cell(row = index_cellule, column = 2)
    c2.value = millesime

    c3 = ws.cell(row = index_cellule, column = 3)
    c3.value = formatb

    c4 = ws.cell(row = index_cellule, column = 4)
    c4.value = prix

    c5 = ws.cell(row = index_cellule, column = 5)
    c5.value = quantite

    c6 = ws.cell(row = index_cellule, column = 6)
    c6.value = conditionnement

    c7 = ws.cell(row = index_cellule, column = 7)
    c7.value = commentaire

    if index_cellule == 1 :
        index_cellule = index_cellule
    else:
        if  ws.cell(row = index_cellule, column = 1).value == ws.cell(row = index_cellule-1, column = 1).value and ws.cell(row = index_cellule, column = 2).value == ws.cell(row = index_cellule-1, column = 2).value and ws.cell(row = index_cellule, column = 3).value == ws.cell(row = index_cellule-1, column = 3).value :
                while ws.cell(row = index_cellule, column = 6).value == ws.cell(row = index_cellule-1, column = 6).value :
                    c6.value = random.choice(Dict_cond)
                    c7.value = 'ATTN : VERIF CDT - DOUBLON LIGNE PRECEDENTE - '+c7.value


index_row = []
for n in range(ws.max_row) :
    if ws.cell(n+1, 1).value is None:
        index_row.append(n+1)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

    
wb2.save(dest_filename)

            


    