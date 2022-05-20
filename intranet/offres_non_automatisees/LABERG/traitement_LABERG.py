import csv
import glob
import re
from unidecode import unidecode
# coding: utf-8

import openpyxl
from openpyxl import Workbook

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

nomProfil : str = "LABERG"
extensionDebut : str = ".xlsx"
extensionFin : str = ".xlsx"

for filename in glob.glob('*' + extensionDebut):
    wb = openpyxl.load_workbook(filename)

    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename = 'sortie_' + nomProfil + extensionFin
    ws = wb2.active
    ws.title = nomProfil

    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    iVin = sheet['B']
    iMillesime = sheet['E']
    iFormatB = sheet['F']
    iQte = sheet['G']
    iPrix = sheet['H']
    iCond = sheet['I']
    iCom = sheet['J']

    iAppellation = sheet['D']

    iRegion = sheet['A']
    iCouleur = sheet['A']

    row_count = sheet.max_row
    column_count = sheet.max_column

    j : int = 1

    for index_ligne in range (1, row_count) :

        vin : str = ft.formaterVin( "", unidecode ( str( iVin[index_ligne].value ) ), "" )
        # print("-" + vin + "-")

        if( ( len(vin) < 1 ) or ( vin == "REGION" ) ):
            continue

        # année
        millesime = ft.formaterAnnee( str( iMillesime[index_ligne].value ) )

        # quantite
        quantite = ft.formaterQuantite( str( iQte[index_ligne].value ) )

        # Format de la bouteille
        formatB = ft.formaterFormatBouteille( str( iFormatB[index_ligne].value ) )

        # Prix
        prix = ft.formaterPrix( str( iPrix[index_ligne].value ) )

        # couleur
        couleur = iCouleur[index_ligne].value

        blanc_str : str =  "BLANC"
        rose_str : str =  "ROSE"

        if( blanc_str in couleur ):
            vin = vin + blanc_str

        elif( rose_str in couleur ):
            vin = vin + rose_str
        
        # appelation
        appelation : str = unidecode( str( iAppellation[index_ligne].value ) )

        # conditionnement
        conditionnement : str = str( iCond[index_ligne].value )

        conditionnement = ft.formaterConditionnement(formatB, quantite, conditionnement, vin)

        # Commentaire
        commentaire = str( iCom[index_ligne].value )

        #On affecte les valeurs dans l'ordre des colonnes voulues.
        ft.affectationLignes(j, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaire)
        j += 1

        wb2.save(dest_filename)
