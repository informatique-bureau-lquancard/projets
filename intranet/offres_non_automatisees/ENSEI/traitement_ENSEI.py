import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook
import requests as req

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

URL = 'https://lenseignedubordeaux.fr/Tarifs.xlsx'
file = req.get(URL, allow_redirects=True, verify = False)

open('Tarifs.xlsx', 'wb').write(file.content)

# coding: utf-8

# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_ENSEI.xlsx'
ws = wb2.active
ws.title = "ENSEI"


# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['C']
iMillesime = sheet['B']
iFormatb = sheet['A']
iPrix = sheet['E']
iQte = sheet['D']
iCond = sheet['H']
#iRegie = sheet['H']
iCom = sheet['G']

row_count = sheet.max_row
column_count = sheet.max_column

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if iPrix[i].value is None :
        continue
    elif iMillesime[i].value == 'Mill.':
        continue
    elif iFormatb[i].value is None:
        continue
    else:  
        vin = unidecode(str(iVin[i].value).upper())
        
        millesime = iMillesime[i].value

        formatb = str(iFormatb[i].value).upper()
        if formatb == 'BOUTEILLE':
            formatb = 'BO'
        elif formatb == 'DEMI' :
            formatb = 'DE'
        elif formatb == 'DMAG' or 'DOUBLE' in formatb:
            formatb = 'DM'
        elif formatb == 'MAGNUM':
            formatb = 'MG' 

        prix = iPrix[i].value

        quantite = iQte[i].value

        cond_rec = unidecode(str(iCond[i].value)).upper().replace(' ','').replace('CB0','CBO')
        conditionnement = ft.formaterConditionnement(formatb, quantite, cond_rec, vin)

        if cond_rec == '':
            commentaire = 'Verif cdt - '+str(iCom[i].value)
        else:
            commentaire = str(iCom[i].value)
        
        commentaire = commentaire.replace('None','')

        c1 = ws.cell(row = i, column = 1)
        c1.value = vin

        c2 = ws.cell(row = i, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = i, column = 3)
        c3.value = formatb

        c4 = ws.cell(row = i, column = 4)
        c4.value = prix

        c5 = ws.cell(row = i, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = i, column = 6)
        c6.value = conditionnement

        c7 = ws.cell(row = i, column = 7)
        c7.value = commentaire

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)
