import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook

# coding: utf-8

# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_ENSEI3.xlsx'
ws = wb2.active
ws.title = "ENSEI"


# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['C']
iMillesime = sheet['B']
iFormatb = sheet['A']
iPrix = sheet['G']
iQte = sheet['F']
iCond = sheet['J']
#iRegie = sheet['J']
iCom = sheet['I']

row_count = sheet.max_row
column_count = sheet.max_column

def conditionnement_par_defaut(formatb,quantite):
    if formatb == 'DE':
        if quantite < 24 :
            conditionnement = 'UNITE'
        else:
            conditionnement = 'CBO24DE'
    elif formatb == 'BO' or 'CL':
        if quantite < 6  :
            conditionnement = 'UNITE'
        elif quantite > 6 and quantite < 12 :
            conditionnement = 'CBO6'
        else:
            conditionnement = 'CBO12'
    elif formatb == 'MG' :
        if quantite < 6 :
            conditionnement = 'UNITE'
        else:
            conditionnement = 'CBO6'
    elif formatb == 'DM':
        if quantite < 3 :
            conditionnement = 'UNITE'
        else:
            conditionnement = 'CBO3'
    else:
        conditionnement = 'CBO1'
    
    return conditionnement



for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if iVin[i].value is None :
        continue
    else:  
        vin = unidecode(str(iVin[i].value).upper())
        
        millesime = iMillesime[i].value

        formatb = str(iFormatb[i].value).upper()
        if formatb == 'BOUTEILLE':
            formatb = 'BO'
        elif formatb == 'MAGNUM':
            formatb = 'MG'
        elif formatb == 'DOUBLE MAGNUM':
            formatb = 'DM'
        elif formatb == 'JEROBOAM':
            formatb = 'JE'
        elif formatb == 'IMPERIALE':
            formatb = 'IM'
        elif formatb == 'DEMI':
            formatb = 'DE'
        # elif formatb == '900 CL':
        #     formatb = 'SA'
        # elif formatb == '1200 CL':
        #     formatb = 'BA'
        # elif formatb == '1500 CL':
        #     formatb = 'NA'
        # elif formatb == '1800 CL':
        #     formatb = 'ME'
        # elif formatb == '225 CL':
        #     formatb = 'MJ'
        # elif formatb == '450 CL':
        #     formatb = 'RE'
        # elif formatb == '2700 CL':
        #     formatb = 'BABY'

        cond_rec = str(iCond[i].value)
        if 'OWC' in cond_rec:
            uqte = re.findall('\d+',cond_rec)
            prix = float(iPrix[i].value)/int(uqte[0])
        else:
            prix = iPrix[i].value
          
        quantite = iQte[i].value

        cond_Com = 0
        if iCond[i].value is None :
            conditionnement = conditionnement_par_defaut(formatb, quantite)
            cond_Com = 1
        else:
            conditionnement = cond_rec.replace('OWC','CBO')
            cond_Com = 0
        
        if iCom[i].value is not None :
            if cond_Com == 1 :
                commentaire = 'Verif cdt - '+iCom[i].value
            else :
                commentaire = iCom[i].value
        else:
            if cond_Com == 1 :
                commentaire = 'Verif cdt '
            else:
                commentaire = ''
        
        c1 = ws.cell(row = i, column = 1)
        c1.value = vin

        c2 = ws.cell(row = i, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = i, column = 3)
        c3.value = formatb

        c4 = ws.cell(row = i, column = 4)
        c4.value = prix

        c5 = ws.cell(row = i, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = i, column = 6)
        c6.value = conditionnement

        c7 = ws.cell(row = i, column = 7)
        c7.value = commentaire

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)



   

