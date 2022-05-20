# coding: utf-8
import csv
import os
import glob
import re
import io
import time
from typing import ItemsView
from unidecode import unidecode
from lxml import etree
import sys
import xml.etree.ElementTree as ET
import urllib
import urllib.request
from urllib.request import Request, urlopen
import feedparser
import openpyxl
from openpyxl import Workbook
import pandas as pd

nom_sortie = 'sortie_ANGWIN_HEBDO_NON_AUTOMATIQUE.csv'

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

nomProfil : str = 'ANGWIN_HEBDO'
extensionDebut : str = '.xlsx'
extensionFin_bis : str = '.xlsx'
extensionFin : str = '.csv'

# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename_bis = 'sortie_bis_' + nomProfil + extensionFin_bis
dest_filename = 'sortie_' + nomProfil + extensionFin
ws = wb2.active
ws.title = nomProfil

# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['D']
iCouleur = sheet['E']
iMillesime = sheet['F']
iFormatB = sheet['G']
iPrix = sheet['I']
iQte = sheet['H']
iCond_formatb = sheet['I']

row_count = sheet.max_row
column_count = sheet.max_column

j : int = 1

for i  in range (1, row_count):

        vin : str = unidecode(str(iVin[i].value)).upper()

        indice_vin : int = -1

        if( re.search('[0-9][0-9][0-9][0-9]', vin) is not None ):
            indice_vin = (re.search('[0-9][0-9][0-9][0-9]', vin)).start()
        
        if indice_vin != -1 :
            vin = vin[0 : (indice_vin - 1)]

        index_vin2 = vin.find("- PROMO")

        if( index_vin2 != -1 ):
            vin = vin[0 : (index_vin2 - 1)]

        # print("vin : " + vin)

        prix : str = ft.formaterPrix(str(iPrix[i].value))

        designation_prix = ["€/unit"]

        if( ft.bLigneIncorrecte(prix, designation_prix, vin)):
            continue;

        quantite : str = str( iQte[i].value )
        couleur : str = str( iCouleur[i].value )
    
        rec_millesime = re.findall('[0-9][0-9][0-9][0-9]', str( iMillesime[i].value ))
        
        if rec_millesime :
            millesime = rec_millesime[0]
        else:
            millesime = 'NV'

        commentaire : str = ""       

        iCom_Vin = ''
        rec_assort = 0
        if("ASSORTIMENT" in vin):
            rec_assort = 1
            liste_vin = vin.split('ASSORTIMENT')

            if len(liste_vin) == 2 :

                if 'LOT ' in liste_vin[0]:
                    vin = liste_vin[1]
                    rec_details = re.findall('\d+[ ][B][O][U][T][E][I][L][L][E][S][ ]', vin)

                    if rec_details :
                        vin = vin.replace(rec_details[0],'')
                        iCom_Vin = liste_vin[0]+rec_details[0]

                else:
                    vin = liste_vin[0]
                    iCom_Vin = liste_vin[1]

            vin += ' ASSORTIMENT'
        
        vin = vin.replace(' '+millesime, '').replace(';','').replace('"','').strip()

        rec_formatb = re.findall('[0-9]?[,]?[0-9][eE0-9][cC]?[lL]', str( iFormatB[i].value ))
        formatB_dict = ['Magnum','Double Magnum', 'Jeroboam', 'Imper']

        formatB : str = ft.formaterFormatBouteille( str( iFormatB[i].value ) )

        for f in formatB_dict :

            if f in formatB :
                iFormat = formatB.find(f)
                formatB = formatB[ int(iFormat): ]
            else:
                break

        if rec_formatb and formatB == '':
            formatB = rec_formatb[0]

        elif 'Imperiale' in formatB :
            formatB = 'IM'

        elif formatB == '' :
            formatB = 'BO'

        formatB = formatB.replace('Double Magnum','DM').replace('Magnum','MG').replace('Jeroboam','JE').replace('62cl','BO').replace('70cl','BO').replace('300cl','DM').replace('300CL','DM').replace(' 3L','')
        
        if rec_assort == 1 :
            conditionnement = 'COLLEC'
            commentaire += ' '+iCom_Vin
        else:
            conditionnement = ft.conditionnementParDefaut(formatB,int(quantite))
            commentaire = 'Verif CDT ' + commentaire
        
        # Fonction Attribution des valeurs dans les bonnes colonnes pour le nouveau workbook
        ft.affectationLignes(j, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaire)

        j = j + 1
# Ecriture du nouveau workbook
wb2.save(dest_filename_bis)

read_file = pd.read_excel(dest_filename_bis)
read_file.to_csv(dest_filename, index = False, encoding="utf-8", sep = ';')