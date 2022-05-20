# coding: utf-8
from cmath import nan
import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook
import pandas as pd
import random

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

nomProfil : str = "TERTRA"
extensionDebut : str = ".xlsx"
extensionFin : str = ".csv"

chemin_sortie_csv = 'sortie_tertra.csv'
Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC']

# On load le le tarif au format xlsx
for filename in glob.glob('*'+extensionDebut):
    reader = pd.read_excel(filename)
    for x in reader.itertuples() :
        #print(x[11])
        if 'Vintage' in str(x[11]) :
            header_index = x.Index

            reader = pd.read_excel(filename, header = header_index)
        # reader = pd.read_excel(filename)

        #     break
    # reader.to_csv('tertra.csv', index = False, index_label= 'firstcolumn' )
        reader.to_csv(nomProfil + extensionFin, index = False, header=False)

# with open('sortie_tertra.csv', newline='') as csvfile :
with open(nomProfil + extensionFin, newline='') as csvfile :
    reader = csv.DictReader(csvfile)
    with open(chemin_sortie_csv,'w',newline='', encoding='utf-8') as monFichierSortie:
    # with open(chemin_sortie_csv,'w',newline='', encoding='utf-8') as monFichierSortie:
        writer = csv.writer(monFichierSortie, delimiter=';', quoting=csv.QUOTE_NONE, escapechar='', quotechar='')
        
        for row in reader :

            vin = unidecode(row['Wine'])
            vin = vin.strip().replace(';','').replace(', ','-')
            millesime = row['Vintage']
            formatb = row['Format']
            prix = row['€ Price']
            quantite_uk = row['UK']
            quantite_France = row['France']
        
            newRow=[vin,millesime,formatb,prix,quantite_uk + quantite_France]
            writer.writerow(newRow)

new_pd_reader = pd.read_csv('sortie_tertra.csv', sep = ';')
# new_pd_reader.to_excel('tertra.xlsx', index = False,)
# new_pd_reader = pd.read_csv(chemin_sortie_csv, sep = ';')
new_pd_reader.to_excel('tertra.xlsx', index = False,)


wb = openpyxl.load_workbook('tertra.xlsx')

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename_bis = 'sortie_'+ 'bis_' + nomProfil.upper()+extensionFin
dest_filename = 'sortie_'+ nomProfil.upper()+extensionFin
ws = wb2.active
ws.title = nomProfil


# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

       
iVin = sheet['M']
iMillesime = sheet['K']
iFormatb = sheet['L']
iPrix = sheet['N']
iQte = sheet['I']
iQte_bis = sheet['H']

row_count = sheet.max_row
column_count = sheet.max_column

for i  in range (1, row_count):

    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.

    if (iVin[i].value is None) :
        continue

    elif iQte[i].value is None:

        if(iQte[i].value == "QUANTITE" or iQte_bis[i].value is None):
            continue

        iQte[i].value = iQte_bis[i].value

    else:  
        vin = unidecode(str(iVin[i].value).upper())
        
        millesime = iMillesime[i].value
        # if iMillesime[i].value is not int :
        #     millesime = 'NV'
        # else:
        #     millesime = iMillesime[i].value


        prix = iPrix[i].value

        quantite = iQte[i].value
        #print(quantite)

        # Reconnaissance du format de bouteille, on cherche uniquement les caractères numérique
        rec_formatb = str(iFormatb[i].value)
        # formatb = ft.formaterFormatBouteille(str(iFormatb[i].value))

        iSize = re.findall('\d+[.]?\d+?',rec_formatb)

        if iSize :
            formatb = ft.formaterFormatBouteille(str(iSize[0]))
            iSize = str(iSize[0])
        else:
            formatb = rec_formatb



        if ' C' in vin :
            r = vin.find(' C')
            rec_cond = re.findall('[ ][C][ TS_][ ]?[0-9]?[0-9]?',vin[int(r):])
        elif 'OWC' in vin :
            rec_cond = re.findall('[(O][O_SW][W C0-9][C )]?[ 0-9)]?[0-9]?[)]?',vin)
        else:
            rec_cond = []
        
        commentaire : str = "Stock FR + UK "
       
        if rec_cond :
            f = vin.find(rec_cond[0])
            vin = vin[0:int(f)]
            conditionnement = rec_cond[0].replace(' ','').replace('CS','CBO').replace('_','C').replace('CT','CC').replace('OWC','CBO').replace('(','').replace(')','').replace('C1','CC1')
            commentaire = ''
        else:

            if(quantite == "ENQ"):
                quantite = 0

            conditionnement = ft.conditionnementParDefaut(formatb, int(quantite))
            commentaire += 'verif cdt'
        
        #On affecte les valeurs dans l'ordre des colonnes voulues.
        c1 = ws.cell(row = i, column = 1)
        c1.value = vin

        c2 = ws.cell(row = i, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = i, column = 3)
        c3.value = formatb

        c4 = ws.cell(row = i, column = 4)
        c4.value = prix

        c5 = ws.cell(row = i, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = i, column = 6)
        c6.value = conditionnement

        c7 = ws.cell(row = i, column = 7)
        c7.value = commentaire

        if i == 1:
            i = i
        else:
        #rec_same_line = dict()
            if  ws.cell(row = i, column = 1).value == ws.cell(row = i-1, column = 1).value and ws.cell(row = i, column = 2).value == ws.cell(row = i-1, column = 2).value and ws.cell(row = i, column = 3).value == ws.cell(row = i-1, column = 3).value :
                # rec_same_line["vin"]=ws.cell(row = i, column = 1)
                # rec_same_line["millesime"]=ws.cell(row = i, column = 2)
                # rec_same_line["formatb"]=ws.cell(row = i, column = 3)
                # rec_same_line["cond"]=ws.cell(row = i, column = 6)
                while ws.cell(row = i, column = 6).value == ws.cell(row = i-1, column = 6).value :
                    c6.value = random.choice(Dict_cond)
                    c7.value = 'ATTN : VERIF CDT - DOUBLON LIGNE PRECEDENTE - '+c7.value

#Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename_bis)

read_file = pd.read_excel(dest_filename_bis)
read_file.to_csv(dest_filename, index = False, encoding="utf-8", sep = ';')



   

