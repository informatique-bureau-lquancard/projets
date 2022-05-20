# coding: utf-8
import csv
from email import header
import glob
from typing import OrderedDict
from numpy import False_
#import codecs
from unidecode import unidecode
import openpyxl
from openpyxl import Workbook
import pandas as pd

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

nomProfil : str = "WINSOU"
extensionDebut : str = ".xlsx"
extensionFin : str = ".xlsx"

profil_bis : str = "winsou"

wb = Workbook()

for filename in glob.glob('*' + extensionDebut):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_' + nomProfil + extensionFin
ws = wb2.active
ws.title = nomProfil

# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['A']
iMillesime = sheet['B']
# Pas de format
# iFormatB = sheet['F']
formatB : str = "NON RECONNU"
iPrix = sheet['G']
iQte = sheet['C']
iCond = sheet['D']

j : int = 1

for index_cellule  in range (1,sheet.max_row):
    vin : str = unidecode(str(iVin[index_cellule].value)).upper()

    prix : str = str( iPrix[index_cellule].value )
    
    designation_prix = ["EUR", "None"]

    if( ft.bLigneIncorrecte(prix, designation_prix, vin) ):
        continue

    prix : float = ft.formaterPrix3(prix)

    millesime = ft.formaterAnnee(iMillesime[index_cellule].value)
    quantite = ft.formaterQuantite(str(iQte[index_cellule].value))
    conditionnement = str(iCond[index_cellule].value)

    conditionnement = conditionnement.replace('OWC', 'CBO')
    conditionnement = conditionnement.replace(' ', '')
    conditionnement = conditionnement.replace('0', '')
    conditionnement = conditionnement.replace('NEUTRAL', 'N')

    commentaire = ''

    # On affecte les valeurs dans l'ordre des colonnes voulues.

    ft.affectationLignes(j, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaire)

    j = j + 1

index_row = []

for n in range(ws.max_row) :
    if ws.cell(n+1, 1).value is None:
        index_row.append(n+1)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)
