# coding: utf-8

from cmath import e
import glob
import codecs
from operator import index
from unidecode import unidecode
from time import sleep
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook
import random


# import des fonctions
import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
from Fonction_tarifs import formatterConditionnementParDefaut

############################################################
#   Script basé sur /TWMOFF by Mr Clement Escande          #
############################################################


Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC','CBN1','CBN2','CBN3','CBN4','CBN6','CBN12','CC','CBO','CBN']

def is_float(n):
    try:
        float(n)   # Type-casting the string to `float`.
                   # If string is not a valid `float`, 
                   # it'll raise `ValueError` exception
    except ValueError:
        return False
    return True

def is_int(n):
    try:
        int(n)   # Type-casting the string to `float`.
                   # If string is not a valid `float`, 
                   # it'll raise `ValueError` exception
    except ValueError:
        return False
    return True


def a():
    wb = Workbook()

    for filename in glob.glob('*.xlsx'):
        wb = openpyxl.load_workbook(filename)

    #row_count = 0
    index = 0
    #on prépare un nouveau workbook
    wb2 = Workbook()

    dest_filename = 'sortie_MARSOI.xlsx'

    ws = wb2.active
    ws.title = "MARSOI"
    
    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    iVin = sheet['E']
    iMillesime = sheet['G']
    iFormatb = sheet['D']
    iPrix = sheet['I']
    iQte = sheet['H']
          
    for index_cellule  in range (sheet.max_row):
        # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
        if iVin[index_cellule].value is None :
            index = index + 1
            continue
        elif 'REFERENCE' in str(iVin[index_cellule].value) :
            index = index + 1
            continue
        else:
            vin = iVin[index_cellule].value
            
            if iMillesime[index_cellule].value == 'NM':
                millesime = 'NV'
            else:
                millesime = iMillesime[index_cellule].value

            if iFormatb[index_cellule].value is None :
                formatb = 'BO'
            else :
                if iFormatb[index_cellule].value == '1,5L':
                    formatb = 'MG'
                elif iFormatb[index_cellule].value == '3L' :
                    formatb = 'DM'
                elif iFormatb[index_cellule].value == '5L' :
                    formatb = 'JE'
                elif iFormatb[index_cellule].value == '6L' :
                    formatb = 'IM'
                elif iFormatb[index_cellule].value == '9L' :
                    formatb = 'SA'
                elif iFormatb[index_cellule].value == '12L' :
                    formatb = 'BA'
                elif iFormatb[index_cellule].value == '15L' :
                    formatb = 'NA'
                elif iFormatb[index_cellule].value == '18L' :
                    formatb = 'ME'
                elif iFormatb[index_cellule].value == '27L' :
                    formatb = 'BABY'
                elif iFormatb[index_cellule].value == '0,375L':
                    formatb = 'DE'

            prix = iPrix[index_cellule].value / 1.2
            prix = round(prix, 2)
            
            quantite = iQte[index_cellule].value

            conditionnement = formatterConditionnementParDefaut(formatb, quantite)

            commentaire = 'Verif cdt'

            # On affecte les valeurs dans l'ordre des colonnes voulues.
            c1 = ws.cell(row = index, column = 1)
            c1.value = vin

            c2 = ws.cell(row = index, column = 2)
            c2.value = millesime

            c3 = ws.cell(row = index, column = 3)
            c3.value = formatb

            c4 = ws.cell(row = index, column = 4)
            c4.value = prix

            c5 = ws.cell(row = index, column = 5)
            c5.value = quantite

            c6 = ws.cell(row = index, column = 6)
            c6.value = conditionnement

            c7 = ws.cell(row = index, column = 7)
            c7.value = commentaire

            index = index + 1

    index_row = []

    for n in range(ws.max_row) :
        if ws.cell(n+1, 1).value is None:
            index_row.append(n+1)

    for row_del in range(len(index_row)):
        ws.delete_rows(idx=index_row[row_del], amount=1)
        index_row = list(map(lambda k: k -1, index_row))
    
    wb2.save(dest_filename)

a()
