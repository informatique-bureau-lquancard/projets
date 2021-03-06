# coding: utf-8
import csv
import glob
import re
import unidecode
import pandas as pd
import openpyxl_dictreader
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.chart import BarChart3D, Reference
from openpyxl.chart.layout import Layout, ManualLayout
#from app import monChateau


import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

for filename in glob.glob('allocations*'):

    reader = pd.read_csv(filename, sep = ';', header = 1)
    reader.to_csv('precha.csv', sep = ';', index = False, encoding='utf-8')
    
with open('precha.csv', newline='', encoding='utf-8') as monFichierEntre:
    
    new_reader = csv.DictReader(monFichierEntre, delimiter=';', doublequote=False)
    
    # nom_sortie='sortie_PRECHA.csv'
    # with open(nom_sortie,'w',newline='', encoding='utf-8') as monFichierSortie:
    #     writer = csv.writer(monFichierSortie, delimiter=';', quoting=csv.QUOTE_NONE, escapechar='', quotechar='')

    df = pd.read_csv('precha.csv', sep = ';')
    total_cols = len(df.axes[1])
    print(total_cols)

    wb = Workbook()
    dest_filename_bis = 'sortie_PRECHA.xlsx'
   
    ws = wb.active
    ws.title = 'precampagne'

    index_sortie : int = 2
    r = 1

    t_offre_N5 = []
    t_offre_N4 = []
    t_offre_N3 = []
    t_offre_N2 = []
    t_offre_N1 = []
    t_achete_N5 = []
    t_achete_N4 = []
    t_achete_N3 = []
    t_achete_N2 = []
    t_achete_N1 = []
    t = 2
    
    for row in new_reader:
        r = r+1

        if total_cols == 15 :
            # sur 5 ans
            quantite_offerte_5 : int = int(row['Qté offerte N-5'])
            quantite_achetee_5 : int = int(row['Qté offerte N-5'])
            quantite_offerte_4 : int = int(row['Qté offerte N-4'])
            quantite_achetee_4 : int = int(row['Qté offerte N-4'])
            quantite_offerte_3 : int = int(row['Qté offerte N-3'])
            quantite_achetee_3 : int = int(row['Qté achetée N-3'])
            quantite_offerte_2 : int = int(row['Qté offerte N-2'])
            quantite_achetee_2 : int = int(row['Qté achetée N-2'])
            quantite_offerte_1 : int = int(row['Qté offerte N-1'])
            quantite_achetee_1 : int = int(row['Qté achetée N-1'])
            negociant = row['Négociant']
            commentaire = row['unofficialCommentPrecampaign']
            demande_allocation = int(row['Demande d\'allocation'])

            header = ws.cell(row = 1, column = 1)
            header.value = 'Offre N-5'

            header = ws.cell(row = 1, column = 2)
            header.value = 'Achat N-5'

            header = ws.cell(row = 1, column = 3)
            header.value = 'Offre N-4'

            header = ws.cell(row = 1, column = 4)
            header.value = 'Achat N-4'

            header = ws.cell(row = 1, column = 5)
            header.value = 'Offre N-3'

            header = ws.cell(row = 1, column = 6)
            header.value = 'Achat N-3'

            header = ws.cell(row = 1, column = 7)
            header.value = 'Offre N-2'

            header = ws.cell(row = 1, column = 8)
            header.value = 'Achat N-2'

            header = ws.cell(row = 1, column = 9)
            header.value = 'Offre N-1'

            header = ws.cell(row = 1, column = 10)
            header.value = 'Achat N-1'

            header = ws.cell(row = 1, column = 11)
            header.value = 'Négociants'

            header = ws.cell(row = 1, column = 12)
            header.value = 'Commentaires'

            header = ws.cell(row = 1, column = 13)
            header.value = 'Demande BLQ'

            c1 = ws.cell(row = index_sortie, column = 1)
            c1.value = quantite_offerte_5

            c2 = ws.cell(row = index_sortie, column = 2)
            c2.value = quantite_achetee_5

            c3 = ws.cell(row = index_sortie, column = 3)
            c3.value = quantite_offerte_4

            c4 = ws.cell(row = index_sortie, column = 4)
            c4.value = quantite_achetee_4

            c5 = ws.cell(row = index_sortie, column = 5)
            c5.value = quantite_offerte_3

            c6 = ws.cell(row = index_sortie, column = 6)
            c6.value = quantite_achetee_3

            c7 = ws.cell(row = index_sortie, column = 7)
            c7.value = quantite_offerte_2

            c8 = ws.cell(row = index_sortie, column = 8)
            c8.value = quantite_achetee_2

            c9 = ws.cell(row = index_sortie, column = 9)
            c9.value = quantite_offerte_1

            c10 = ws.cell(row = index_sortie, column = 10)
            c10.value = quantite_achetee_1

            c11 = ws.cell(row = index_sortie, column = 11)
            c11.value = negociant

            c12 = ws.cell(row = index_sortie, column = 12)
            c12.value = commentaire

            c13 = ws.cell(row = index_sortie, column = 13)
            c13.value = demande_allocation

            index_sortie += 1
            continue        
        if total_cols == 13 :
            # sur 4 ans
            quantite_offerte_4 : int = int(row['Qté offerte N-4'])
            quantite_achetee_4 : int = int(row['Qté offerte N-4'])
            quantite_offerte_3 : int = int(row['Qté offerte N-3'])
            quantite_achetee_3 : int = int(row['Qté achetée N-3'])
            quantite_offerte_2 : int = int(row['Qté offerte N-2'])
            quantite_achetee_2 : int = int(row['Qté achetée N-2'])
            quantite_offerte_1 : int = int(row['Qté offerte N-1'])
            quantite_achetee_1 : int = int(row['Qté achetée N-1'])
            negociant = row['Négociant']
            commentaire = row['unofficialCommentPrecampaign']
            demande_allocation = int(row['Demande d\'allocation'])

            header = ws.cell(row = 1, column = 1)
            header.value = 'Offre N-4'

            header = ws.cell(row = 1, column = 2)
            header.value = 'Achat N-4'

            header = ws.cell(row = 1, column = 3)
            header.value = 'Offre N-3'

            header = ws.cell(row = 1, column = 4)
            header.value = 'Achat N-3'

            header = ws.cell(row = 1, column = 5)
            header.value = 'Offre N-2'

            header = ws.cell(row = 1, column = 6)
            header.value = 'Achat N-2'

            header = ws.cell(row = 1, column = 7)
            header.value = 'Offre N-1'

            header = ws.cell(row = 1, column = 8)
            header.value = 'Achat N-1'

            header = ws.cell(row = 1, column = 9)
            header.value = 'Négociants'

            header = ws.cell(row = 1, column = 10)
            header.value = 'Commentaires'

            header = ws.cell(row = 1, column = 11)
            header.value = 'Demande BLQ'
                       
            c1 = ws.cell(row = index_sortie, column = 1)
            c1.value = quantite_offerte_4

            c2 = ws.cell(row = index_sortie, column = 2)
            c2.value = quantite_achetee_4

            c3 = ws.cell(row = index_sortie, column = 3)
            c3.value = quantite_offerte_3

            c4 = ws.cell(row = index_sortie, column = 4)
            c4.value = quantite_achetee_3

            c5 = ws.cell(row = index_sortie, column = 5)
            c5.value = quantite_offerte_2

            c6 = ws.cell(row = index_sortie, column = 6)
            c6.value = quantite_achetee_2

            c7 = ws.cell(row = index_sortie, column = 7)
            c7.value = quantite_offerte_1

            c8 = ws.cell(row = index_sortie, column = 8)
            c8.value = quantite_achetee_1

            c9 = ws.cell(row = index_sortie, column = 9)
            c9.value = negociant

            c10 = ws.cell(row = index_sortie, column = 10)
            c10.value = commentaire

            c11 = ws.cell(row = index_sortie, column = 11)
            c11.value = demande_allocation

            index_sortie += 1
            continue
        if total_cols == 11 :
            # sur 3 ans
            quantite_offerte_3 : int = int(row['Qté offerte N-3'])
            quantite_achetee_3 : int = int(row['Qté achetée N-3'])
            quantite_offerte_2 : int = int(row['Qté offerte N-2'])
            quantite_achetee_2 : int = int(row['Qté achetée N-2'])
            quantite_offerte_1 : int = int(row['Qté offerte N-1'])
            quantite_achetee_1 : int = int(row['Qté achetée N-1'])
            negociant = row['Négociant']
            commentaire = row['unofficialCommentPrecampaign']
            demande_allocation : int = int(row['Demande d\'allocation'])

            if quantite_achetee_2 < quantite_achetee_1 :
                variation = '↑'
            elif quantite_achetee_2 > quantite_achetee_1 :
                variation = '↓'
            elif quantite_achetee_2 == quantite_achetee_1 :
                variation = '='
            
            header = ws.cell(row = 1, column = 1)
            header.value = 'Offre N-3'

            header = ws.cell(row = 1, column = 2)
            header.value = 'Achat N-3'

            header = ws.cell(row = 1, column = 3)
            header.value = 'Offre N-2'

            header = ws.cell(row = 1, column = 4)
            header.value = 'Achat N-2'

            header = ws.cell(row = 1, column = 5)
            header.value = 'Offre N-1'

            header = ws.cell(row = 1, column = 6)
            header.value = 'Achat N-1'

            header = ws.cell(row = 1, column = 7)
            header.value = 'Variation'

            header = ws.cell(row = 1, column = 8)
            header.value = 'Négociants'

            header = ws.cell(row = 1, column = 9)
            header.value = 'Commentaires'

            header = ws.cell(row = 1, column = 10)
            header.value = 'Demande BLQ'
            
            c1 = ws.cell(row = index_sortie, column = 1)
            c1.value = quantite_offerte_3

            c2 = ws.cell(row = index_sortie, column = 2)
            c2.value = quantite_achetee_3

            c3 = ws.cell(row = index_sortie, column = 3)
            c3.value = quantite_offerte_2

            c4 = ws.cell(row = index_sortie, column = 4)
            c4.value = quantite_achetee_2

            c5 = ws.cell(row = index_sortie, column = 5)
            c5.value = quantite_offerte_1

            c6 = ws.cell(row = index_sortie, column = 6)
            c6.value = quantite_achetee_1

            c7 = ws.cell(row = index_sortie, column = 7)
            c7.value = variation

            c8 = ws.cell(row = index_sortie, column = 8)
            c8.value = negociant

            c9 = ws.cell(row = index_sortie, column = 9)
            c9.value = commentaire

            c10 = ws.cell(row = index_sortie, column = 10)
            c10.value = demande_allocation

            index_sortie += 1
            #print(quantite_offerte_3, quantite_achetee_3, quantite_offerte_2, quantite_achetee_2, quantite_offerte_1, quantite_achetee_1, negociant, commentaire, demande_allocation)
            continue 
        if total_cols == 9 :
            # sur 2 ans
            quantite_offerte_2 : int = int(row['Qté offerte N-2'])
            quantite_achetee_2 : int = int(row['Qté achetée N-2'])
            quantite_offerte_1 : int = int(row['Qté offerte N-1'])
            quantite_achetee_1 : int = int(row['Qté achetée N-1'])
            negociant = row['Négociant']
            commentaire = row['unofficialCommentPrecampaign']
            demande_allocation : int= int(row['Demande d\'allocation'])

            header = ws.cell(row = 1, column = 1)
            header.value = 'Offre N-2'

            header = ws.cell(row = 1, column = 2)
            header.value = 'Achat N-2'

            header = ws.cell(row = 1, column = 3)
            header.value = 'Offre N-1'

            header = ws.cell(row = 1, column = 4)
            header.value = 'Achat N-1'

            header = ws.cell(row = 1, column = 5)
            header.value = 'Négociants'

            header = ws.cell(row = 1, column = 6)
            header.value = 'Commentaires'

            header = ws.cell(row = 1, column = 7)
            header.value = 'Demande BLQ'

            c1 = ws.cell(row = index_sortie, column = 1)
            c1.value = quantite_offerte_2

            c2 = ws.cell(row = index_sortie, column = 2)
            c2.value = quantite_achetee_2

            c3 = ws.cell(row = index_sortie, column = 3)
            c3.value = quantite_offerte_1

            c4 = ws.cell(row = index_sortie, column = 4)
            c4.value = quantite_achetee_1

            c5 = ws.cell(row = index_sortie, column = 5)
            c5.value = negociant

            c6 = ws.cell(row = index_sortie, column = 6)
            c6.value = commentaire

            c7 = ws.cell(row = index_sortie, column = 7)
            c7.value = demande_allocation

            index_sortie += 1
            continue
        if total_cols == 7 :
            # sur 1 an
            quantite_offerte_1 : int = int(row['Qté offerte N-1'])
            quantite_achetee_1 : int = int(row['Qté achetée N-1'])
            negociant = row['Négociant']
            commentaire = row['unofficialCommentPrecampaign']
            demande_allocation = int(row['Demande d\'allocation'])

            header = ws.cell(row = 1, column = 1)
            header.value = 'Offre N-1'

            header = ws.cell(row = 1, column = 2)
            header.value = 'Achat N-1'

            header = ws.cell(row = 1, column = 3)
            header.value = 'Négociants'

            header = ws.cell(row = 1, column = 4)
            header.value = 'Commentaires'

            header = ws.cell(row = 1, column = 5)
            header.value = 'Demande BLQ'


            c1 = ws.cell(row = index_sortie, column = 1)
            c1.value = quantite_offerte_1

            c2 = ws.cell(row = index_sortie, column = 2)
            c2.value = quantite_achetee_1

            c3 = ws.cell(row = index_sortie, column = 3)
            c3.value = negociant

            c4 = ws.cell(row = index_sortie, column = 4)
            c4.value = commentaire

            c5 = ws.cell(row = index_sortie, column = 5)
            c5.value = demande_allocation

            index_sortie += 1
            continue
    
    print(r)
    if total_cols == 15 :
        ws[f'A{r+1}'] = f'=SUM(A2:A{r})'
        ws[f'B{r+1}'] = f'=SUM(B2:B{r})'
        ws[f'C{r+1}'] = f'=SUM(C2:C{r})'
        ws[f'D{r+1}'] = f'=SUM(D2:D{r})'
        ws[f'E{r+1}'] = f'=SUM(E2:E{r})'
        ws[f'F{r+1}'] = f'=SUM(F2:F{r})'
        ws[f'G{r+1}'] = f'=SUM(G2:G{r})'
        ws[f'H{r+1}'] = f'=SUM(H2:H{r})'
        ws[f'I{r+1}'] = f'=SUM(I2:I{r})'
        ws[f'J{r+1}'] = f'=SUM(J2:J{r})'
        ws[f'M{r+1}'] = f'=SUM(M2:M{r})'

    if total_cols == 11 :
        ws[f'A{r+1}'] = f'=SUM(A2:A{r})'
        ws[f'B{r+1}'] = f'=SUM(B2:B{r})'
        ws[f'C{r+1}'] = f'=SUM(C2:C{r})'
        ws[f'D{r+1}'] = f'=SUM(D2:D{r})'
        ws[f'E{r+1}'] = f'=SUM(E2:E{r})'
        ws[f'F{r+1}'] = f'=SUM(F2:F{r})'
        ws[f'J{r+1}'] = f'=SUM(J2:J{r})'

        
        
        for t in range(2,r+1):
            t_offre_N3.append(ws[f'A{t}'].value)
            t_offre_N2.append(ws[f'C{t}'].value)
            t_offre_N1.append(ws[f'E{t}'].value)
            t_achete_N3.append(ws[f'B{t}'].value)
            t_achete_N2.append(ws[f'D{t}'].value)
            t_achete_N1.append(ws[f'F{t}'].value)
            t = t+1
        
        total_offer_N3 = sum(t_offre_N3)
        total_offer_N2 = sum(t_offre_N2)
        total_offer_N1 = sum(t_offre_N1)
        total_achete_N3 = sum(t_achete_N3)
        total_achete_N2 = sum(t_achete_N2)
        total_achete_N1 = sum(t_achete_N1)

        ws['T29'] = 'Total offre'
        ws['U29'] = 'Total acheté'

        ws['S30'] = 'N-1'
        ws['S31'] = 'N-2'
        ws['S32'] = 'N-3'
        
        ws['T30'] = total_offer_N3
        ws['T31'] = total_offer_N2
        ws['T32'] = total_offer_N1
        ws['U30'] = total_achete_N3
        ws['U31'] = total_achete_N2
        ws['U32'] = total_achete_N1

        values : int = Reference(ws, min_col = 20, min_row = 29 , max_row = 32, max_col = 21)
        titles = Reference(ws, min_col = 19, min_row = 30, max_row = 32, max_col = 19)

        chart_offre = BarChart3D()
    
        chart_offre.add_data(data = values, titles_from_data = True) 
        chart_offre.set_categories(titles)
    
        chart_offre.title = " EVOLUTION ALLOCATIONS SUR 3 ANS "
    
        chart_offre.x_axis.title = "Année "
    
        chart_offre.y_axis.title = "Volume"
   
        ws.add_chart(chart_offre, f"D{r+4}")

        
        # Si besoin de faire un graphique séparé
        # values : int = Reference(ws, min_col = 21, min_row = 30 , max_row = 32, max_col = 21)

        # chart_achats = BarChart3D()
    
        # chart_achats.add_data(values) 
    
        # chart_achats.title = " EVOLUTION ACHATS SUR 3 ANS "
    
        # chart_achats.x_axis.title = " Quantité "
    
        # chart_achats.y_axis.title = " Année "
    
        # ws.add_chart(chart_achats, "H30")


    if total_cols == 7 :
        ws[f'A{r+1}'] = f'=SUM(A2:A{r})'
        ws[f'B{r+1}'] = f'=SUM(B2:B{r})'
        ws[f'E{r+1}'] = f'=SUM(E2:E{r})'


######################################################################
# Définition du style du tableau pour le cas ou historique sur 3 ans #
######################################################################
   


    ws.column_dimensions['I'].width = 50
    ws['I4'].alignment = Alignment(wrap_text = True)

    def set_border(ws, r_min, c_min, c_max, r_max):
    
        border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

        rows = ws.iter_rows(min_row = r_min, min_col = c_min, max_col = c_max, max_row = r_max)
        for row in rows:    
            for cell in row:
                cell.border = border
    
    set_border(ws, 2, 1, 10, r)

    def set_alignment_commentaires(ws, r_min, c_min, c_max, r_max) :
        rows = ws.iter_rows(min_row = r_min, min_col = c_min, max_col = c_max, max_row = r_max)
        for row in rows :
            for cell in row :
                cell.alignment = Alignment(wrap_text = True)
    
    set_alignment_commentaires(ws, 1, 9, 9, r)

    def set_dimension_negociants(ws, r_min, c_min, c_max, r_max, col) :
        max_length = 0
        rows = ws.iter_rows(min_row = r_min, min_col = c_min, max_col = c_max, max_row = r_max)
        for row in rows :
            for cell in row :
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                
                adjusted_width = (max_length + 2 ) * 1.2
                ws.column_dimensions[col].width = adjusted_width
    
    set_dimension_negociants(ws, 1, 8, 8, r, 'H')

    def set_alignement_colonne(ws, r_min, c_min, c_max, r_max, col) :
        rows = ws.iter_rows(min_row = r_min, min_col = c_min, max_col = c_max, max_row = r_max)
        for row in rows :
            for cell in row :
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
                if cell.value == '↑':
                    cell.font = Font(color = '0099CC00', bold=True)
                elif cell.value == '↓':
                    cell.font = Font(color = '00FF0000', bold=True)
                else:
                    cell.font = Font(color = '00000000', bold=True)
                    
    set_alignement_colonne(ws, 1, 7, 7, r, 'G')

    for cell in ws[1:1]:
        cell.font = Font(color = '00FF0000', italic=True)

    for cell in ws[r+1:r+1]:
        cell.font = Font(bold=True)

    # server_run = app_rules()
    # chateau = monChateau()
    # print(chateau)
    wb.save(dest_filename_bis)

