# coding: utf-8
import glob
from operator import index
from unidecode import unidecode
from time import sleep
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook
import xlrd
import xlwt
import random


import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft


def is_int(n):
    try:
        int(n)   # Type-casting the string to `float`.
                   # If string is not a valid `float`, 
                   # it'll raise `ValueError` exception
    except ValueError:
        return False
    return True

nomProfil : str = "COMVIG"
extensionDebut : str = ".xls"
extensionFin : str = ".csv"

# Ouverture du fichier avec xlrd
# l'obectif est d'identifier les lignes dont le background est rouge rgb(255, 0, 0)
# puis de les supprimer, car nous ne voulons pas importer ces lignes.
for xls_file in glob.glob("*"+extensionDebut):
    
    # formatting_info = True, pour récupérer le format des cellules
    xls_wb = xlrd.open_workbook(xls_file, formatting_info=True)
     
    active_sheet = xls_wb.sheet_by_index(0)
    rows = active_sheet.nrows
    cols = active_sheet.ncols
    
    # Ouverture d'un workbook qui recevra les lignes voulues
    new_wb = xlwt.Workbook() 
    new_sheet = new_wb.add_sheet('cleaned sheet')
    

    for r in range(0, rows) :
        for c in range(0, cols) :
            
            # Parcours des lignes du tableau, sur la première colonne
            cell = active_sheet.cell(r,0) 
            
            # Récupération du code format des cellules
            cell_format = cell.xf_index 
            myFormatting = xls_wb.xf_list[cell_format]        
            
            # Récupération du code rgb de la cellule
            rgb = xls_wb.colour_map[myFormatting.background.pattern_colour_index]

            # Ecriture dans le workbook réceptacle, sans les lignes rouges
            if rgb == (255, 0, 0):
                continue
            cell_value = active_sheet.cell_value(r,c)
            new_sheet.write(r, c, cell_value)

Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC', 'CBN1', 'CBN2', 'CBN3','CBN4', 'CBN6','CBN12', 'CBN24DE']

# Enregistrement du nouveau workbook dans un fichier xls (xlwt ne permet d'écrire directement en xlsx)
new_wb.save(xls_file)
        
# Conversion du fichier xls en xlsx pour traitement classique
reader = pd.read_excel(xls_file)
reader.to_excel('COMVIG.xlsx', index = False)

wb = openpyxl.load_workbook('COMVIG.xlsx')

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename_bis = 'sortie_'+ 'bis_' + nomProfil.upper()+extensionFin
dest_filename = 'sortie_'+ nomProfil.upper()+extensionFin
ws = wb2.active
ws.title = "COMVIG sans lignes rouges"

# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['A']
iAppellation = sheet['B']
iMillesime = sheet['D']
iFormatb = sheet['G']
iPrix = sheet['E']
iQte = sheet['F']
iCond_Com = sheet['H']

row_count = sheet.max_row
column_count = sheet.max_column

for index_cellule  in range (1,sheet.max_row):
    if iVin[index_cellule].value is None :
        continue
    vin = unidecode(str(iVin[index_cellule].value)).upper()
    if iMillesime[index_cellule].value is None :
        millesime = 'NV'
    elif '/' in str(iMillesime[index_cellule].value) :
        millesime = 'NV'
    else :
        millesime = iMillesime[index_cellule].value

    prix = iPrix[index_cellule].value

    i_formatb = unidecode(str(iFormatb[index_cellule].value)).upper()
    if 'BOUTEILLE' in i_formatb or 'BLLE' in i_formatb or 'ASSORT' in i_formatb:
        formatb = 'BO'
    elif 'DEMI' in i_formatb :
        formatb = 'DE'
    elif 'CINQ' in i_formatb :
        formatb = 'CL'
    elif 'MAG' in i_formatb and 'DOUBLE' not in i_formatb :
        formatb = 'MG'
    elif 'DOUBLE' in i_formatb and 'MAG' in i_formatb or 'DMG' in i_formatb:
        formatb = 'DM'
    elif 'JERO' in i_formatb :
        formatb = 'JE'
    elif 'IMP' in i_formatb :
        formatb = 'IM'
    elif 'MATHUSA' in i_formatb :
        formatb = 'MAT'
    else :
        formatb = 'Non reconnu'
    
    quantite = iQte[index_cellule].value
    if quantite is None :
        quantite = 0
    elif is_int(quantite) is False :
        quantite = 0

    icdt = 0
    rec_cond = re.findall('[C][ABOC][\dON]?[\d]+', str(iCond_Com[index_cellule].value))
    
    if rec_cond :
        icdt = 0        
        for i_rec in range(len(rec_cond)) :
            if '12' in rec_cond[i_rec] :
                conditionnement = rec_cond[i_rec]
            else:
                conditionnement = rec_cond[0]
            conditionnement = conditionnement.replace('CAO','CC').replace('CO','CC')
            if conditionnement not in Dict_cond :
                if 'ASSORT' in vin :
                    conditionnement = 'COLLEC'
                    icdt = 2
                else:
                    conditionnement = 'UNITE'
                    icdt = 1    
    else:               
        i_cond = unidecode(str(iCond_Com[index_cellule].value)).upper()
        if 'COFFRET' in i_cond :
            conditionnement = 'COF'
            icdt = 0
        elif 'COLLEC' in i_cond :
            conditionnement = 'COLLEC'
            icdt = 0
        else :
            if i_cond  == 'NONE' :
                conditionnement = ft.conditionnementParDefaut(formatb, quantite)
                icdt = 1
            else:
                if 'COLLECTION' in vin or 'VERTICALE' in vin :
                    conditionnement = 'COLLEC'
                    icdt = 0
                elif 'ASSORT' in vin :
                    conditionnement = 'COLLEC'
                    icdt = 2
                else:
                    conditionnement = 'UNITE'
                    icdt = 1
        
    
    
            
    commentaire = unidecode(str(iCond_Com[index_cellule].value)).upper()
    if icdt == 1 :
        commentaire = 'verif cdt - '+unidecode(str(iCond_Com[index_cellule].value)).upper()
    elif icdt == 2 :
        commentaire = str(iAppellation[index_cellule].value)+' '+unidecode(str(iCond_Com[index_cellule].value)).upper()
    
    commentaire = commentaire.replace(conditionnement,'').replace(' - NONE','').replace('NONE','').replace('TION ','')
    
    # On affecte les valeurs dans l'ordre des colonnes voulues.

    c1 = ws.cell(row = index_cellule, column = 1)
    c1.value = vin

    c2 = ws.cell(row = index_cellule, column = 2)
    c2.value = millesime

    c3 = ws.cell(row = index_cellule, column = 3)
    c3.value = formatb

    c4 = ws.cell(row = index_cellule, column = 4)
    c4.value = prix

    c5 = ws.cell(row = index_cellule, column = 5)
    c5.value = quantite

    c6 = ws.cell(row = index_cellule, column = 6)
    c6.value = conditionnement

    c7 = ws.cell(row = index_cellule, column = 7)
    c7.value = commentaire

    if index_cellule == 1 :
        index_cellule = index_cellule
    else:
        if  ws.cell(row = index_cellule, column = 1).value == ws.cell(row = index_cellule-1, column = 1).value and ws.cell(row = index_cellule, column = 2).value == ws.cell(row = index_cellule-1, column = 2).value and ws.cell(row = index_cellule, column = 3).value == ws.cell(row = index_cellule-1, column = 3).value :
                while ws.cell(row = index_cellule, column = 6).value == ws.cell(row = index_cellule-1, column = 6).value :
                    c6.value = random.choice(Dict_cond)
                    c7.value = 'ATTN : VERIF CDT - DOUBLON LIGNE PRECEDENTE - '+c7.value


index_row = []
for n in range(ws.max_row) :
    if ws.cell(n+1, 1).value is None:
        index_row.append(n+1)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

    
wb2.save(dest_filename_bis)

read_file = pd.read_excel(dest_filename_bis)
read_file.to_csv(dest_filename, index = False, encoding="utf-8", sep = ';')            


    