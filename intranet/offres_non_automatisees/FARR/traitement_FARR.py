# coding: utf-8
from cmath import e
import glob
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook
import pandas as pd

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

############################################################
#   Script basé sur /TWMOFF by Mr Clement Escande          #
############################################################


Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO4','CBO6','CBO12','CBO24DE','COLLEC','CBN1','CBN2','CBN3','CBN4','CBN6','CBN12','CC','CBO','CBN']

def is_int(n):
    try:
        int(n)   # Type-casting the string to `float`.
                   # If string is not a valid `float`, 
                   # it'll raise `ValueError` exception
    except ValueError:
        return False
    return True

def is_float(n):
    try:
        float(n)   # Type-casting the string to `float`.
                   # If string is not a valid `float`, 
                   # it'll raise `ValueError` exception
    except ValueError:
        return False
    return True


def a():
    wb = Workbook()

    for filename in glob.glob('*.xlsx'):
        wb = openpyxl.load_workbook(filename)

    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename = 'sortie_FARR.xlsx'
    ws = wb2.active
    ws.title = "FARR"


    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]
    
    iVin = sheet['C']
    iMillesime = sheet['D']
    iFormatB = sheet['F']
    iPrix = sheet['O']
    iQte_cs = sheet['H']
    iQte_units = sheet['I']
    iUqte = sheet['M']
    iCond = sheet['L']
    iUcond = sheet['E']
    
    for index_cellule  in range (1,sheet.max_row):
        vin = unidecode(str(iVin[index_cellule].value)).upper()
        
        millesime = iMillesime[index_cellule].value
        if is_int(millesime) is False:
           millesime = 'NV'
        else:

            millesime_int : int = int(millesime)
            millesime : str = str(millesime_int)
        
        rec_formatB = re.findall('[0-9]?[0-9]?[0-9][0-9][c][l]', str(iFormatB[index_cellule].value))
        if rec_formatB :
            rec_formatB = re.findall('\d+',rec_formatB[0])[0]
            formatB = ft.formaterFormatBouteille(rec_formatB)
        else:
            formatB = 'Non reconnu'  
        
        # prix = iPrix[index_cellule].value
        # if prix is None :
        #     prix = 'N/C'
        # elif is_float(prix) is False :
        #     prix = 'N/C'
        # else:
        if iPrix[index_cellule].value is None or iPrix[index_cellule].value =='Unit Sell EUR':
            prix = 'N/C'
        else:
            prix = iPrix[index_cellule].value
            prix = round(prix,2)

        if iQte_cs[index_cellule].value == 0 :
            quantite = iQte_units[index_cellule].value
        else :
            if iQte_cs[index_cellule].value is None or iUqte[index_cellule].value is None :
                quantite = "Error"
            elif is_int(iQte_cs[index_cellule].value) is False or is_int(iUqte[index_cellule].value) is False :
                quantite = 'Error'              
            else:
                quantite = int(iQte_cs[index_cellule].value) * int(iUqte[index_cellule].value)

        dict_collec = ['ASSORTMENT', 'MIXED', 'COLLECTION']
        iCom = 0
        if iCond[index_cellule].value == 'cs' :
            uQte = re.findall('\d+',str(iUcond[index_cellule].value))
            for c in dict_collec :
                if c in vin :
                    conditionnement = 'COLLEC'
            
            if iMillesime[index_cellule].value == 'M.V.' :
                conditionnement = 'COLLEC'
            elif uQte and iMillesime[index_cellule].value != 'M.V.':
                conditionnement = 'CBO'+uQte[0] 
                if conditionnement not in Dict_cond :
                    iCom = 1   
                    conditionnement = 'UNITE'      
            else:
                if quantite < 12 :
                    conditionnement = 'UNITE'
                else:
                    conditionnement = 'CBO12'
         
        else:
            conditionnement = 'UNITE'

        
        if conditionnement == 'COLLEC' :
            quantite = iQte_cs[index_cellule].value


        if iCom == 1 :
            commentaire = 'verif cdt'
        else:
            commentaire = ''

        # On affecte les valeurs dans l'ordre des colonnes voulues.

        c1 = ws.cell(row = index_cellule, column = 1)
        c1.value = vin

        c2 = ws.cell(row = index_cellule, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = index_cellule, column = 3)
        c3.value = formatB

        c4 = ws.cell(row = index_cellule, column = 4)
        c4.value = prix

        c5 = ws.cell(row = index_cellule, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = index_cellule, column = 6)
        c6.value = conditionnement

        c7 = ws.cell(row = index_cellule, column = 7)
        c7.value = commentaire

    index_row = []

    for n in range(ws.max_row) :
        if ws.cell(n+1, 1).value is None:
            index_row.append(n+1)

    for row_del in range(len(index_row)):
        ws.delete_rows(idx=index_row[row_del], amount=1)
        index_row = list(map(lambda k: k -1, index_row))

    wb2.save(dest_filename)
    csv_converter = pd.read_excel(dest_filename)
    csv_converter.to_csv('sortie_FARR.csv', index = False, encoding="utf-8", sep = ';')
a()
