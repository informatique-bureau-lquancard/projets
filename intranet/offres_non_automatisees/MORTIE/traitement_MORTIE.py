# coding: utf-8
import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
from Fonction_tarifs import formatterConditionnementParDefaut
import Fonction_tarifs

# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_MORTIE.xlsx'
ws = wb2.active
ws.title = "MORTIE"


# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[1]] # lecture de l'onglet 2

iVin = sheet['A']
iMillesime = sheet['C']
iFormatb = sheet['F']
iPrix = sheet['E']
iQte = sheet['D']
#iCond = sheet['M']
#iColor = sheet['C']
#iRegion = sheet['A']
iAppellation = sheet['B']

row_count = sheet.max_row
column_count = sheet.max_column

for index_cellule  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if iPrix[index_cellule].value is None :
        continue
    elif str(iAppellation[index_cellule].value) == 'APPELLATION' :
        continue
    else:
        vin = unidecode(str(iVin[index_cellule].value).upper())
        vin = vin.replace('"','').replace("'","")         
       
        millesime = iMillesime[index_cellule].value

        prix = iPrix[index_cellule].value

        iSize = re.findall('[0-9]?[.]?\d+',str(iFormatb[index_cellule].value).replace(',','.'))
        #print(iSize)

        if iSize :
            iSize = iSize[0].replace('.',',')
        
        formatb = Fonction_tarifs.formatterFormatBouteille2(iSize)
        #print(formatb)
        
        estim_quantite = iQte[index_cellule].value
        if estim_quantite is not None :
            quantite = 60
            commentaire = ''
        else:
            quantite = 24
            commentaire = 'Quantite < 60'
        
        conditionnement = 'CBO12' 

        #print(vin, millesime,formatb, prix, quantite, conditionnement, commentaire)   

        # On affecte les valeurs dans l'ordre des colonnes voulues.
        c1 = ws.cell(row = index_cellule, column = 1)
        c1.value = vin

        c2 = ws.cell(row = index_cellule, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = index_cellule, column = 3)
        c3.value = formatb

        c4 = ws.cell(row = index_cellule, column = 4)
        c4.value = prix

        c5 = ws.cell(row = index_cellule, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = index_cellule, column = 6)
        c6.value = conditionnement

        c7 = ws.cell(row = index_cellule, column = 7)
        c7.value = commentaire

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)



   

