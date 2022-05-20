# coding: utf-8
import csv
import glob
import codecs
from unidecode import unidecode
from time import sleep
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook
import random

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC','CBN1','CBN2','CBN3','CBN4','CBN6','CBN12','CC','CBO','CBN']
# Conversion du fichier xls en xlsx
for xls_file in glob.glob('*.xls'):
    reader = pd.read_excel(xls_file)
    reader.to_excel('RDRG.xlsx', index = False)

sleep(1)

# On load le le tarif au format xlsx dans un workbook
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_RDRG.xlsx'
ws = wb2.active
ws.title = "R.D.R.G"

# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['C']
iAppellation = sheet['D']
iMillesime = sheet['E']
iFormatB = sheet['F']
iPrix = sheet['H']
iQte = sheet['G']
iCond = sheet['I']
#iDate_offre = sheet['I']  #validité
iRegion = sheet['B']
#iType_vendeur = sheet['H']
iCom = sheet['J']

row_count = sheet.max_row
column_count = sheet.max_column

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if iMillesime[i].value is None :
        continue
    elif 'MILL' in str(iMillesime[i].value) :
        continue
    else:
        rec_vin = unidecode(str(iVin[i].value).upper())
        region = unidecode(str(iRegion[i].value).upper())
        appellation = unidecode(str(iAppellation[i].value).upper())

        # if region == 'BORDEAUX':
        #     vin = rec_vin
        # else:
        #     vin = rec_vin+' '+appellation

        vin : str = rec_vin + " " + appellation
        
        millesime = iMillesime[i].value

        formatB : str = ft.formaterFormatBouteille(str(iFormatB[i].value))

        prix = iPrix[i].value

        quantite = iQte[i].value

        conditionnement = str(iCond[i].value).upper().replace('CTN','CC').replace('UNIT','UNITE').replace('CTO','CC')
        Cond_Com = 0
        if 'COLLECTION' in vin :
            conditionnement = 'COLLEC'
            quantite = re.findall('[0-9][0-9]?[0-9]?[0-9]?',quantite)[0]
        elif conditionnement not in Dict_cond :
            conditionnement = 'UNITE'
            Cond_Com = 1
               
        
        if Cond_Com == 1 :
            if iCom[i].value is None :
                commentaire = 'VERIF CDT - '+unidecode(str(iCond[i].value))
            else:
                commentaire = 'VERIF CDT - '+unidecode(str(iCond[i].value))+' - '+unidecode(str(iCom[i].value))
        else:
            if iCom[i].value is None :
                commentaire = ''
            else:
                commentaire = unidecode(str(iCom[i].value))


        # On affecte les valeurs dans l'ordre des colonnes voulues.
        c1 = ws.cell(row = i, column = 1)
        c1.value = vin

        c2 = ws.cell(row = i, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = i, column = 3)
        c3.value = formatB

        c4 = ws.cell(row = i, column = 4)
        c4.value = prix

        c5 = ws.cell(row = i, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = i, column = 6)
        c6.value = conditionnement

        c7 = ws.cell(row = i, column = 7)
        c7.value = commentaire

        if  ws.cell(row = i, column = 1).value == ws.cell(row = i-1, column = 1).value and ws.cell(row = i, column = 2).value == ws.cell(row = i-1, column = 2).value and ws.cell(row = i, column = 3).value == ws.cell(row = i-1, column = 3).value :
            while ws.cell(row = i, column = 6).value == ws.cell(row = i-1, column = 6).value :
                c6.value = random.choice(Dict_cond)
                c7.value = 'ATTN : VERIF CDT - DOUBLON LIGNE PRECEDENTE - '+c7.value


# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)