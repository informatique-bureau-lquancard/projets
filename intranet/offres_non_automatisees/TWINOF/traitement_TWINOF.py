# coding: utf-8
import csv
import glob
import codecs
from unidecode import unidecode
from time import sleep
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook
import fichierFonction_TWINOF

extensionDebut = ".xls"
nomFichier = "TWINOF" 
extensionFin = ".xlsx"

# Conversion du fichier csv en xlsx
# for xls_file in glob.glob('*'+extensionDebut):
#     reader = pd.read_excel(xls_file)
#     reader.to_excel(nomFichier+'*'+extensionFin, index = False)

# sleep(1)

# On load le le tarif au format xlsx
for filename in glob.glob('*'+extensionFin):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_TWINOF.xlsx'
ws = wb2.active
ws.title = nomFichier.upper()

# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

colonneVin = sheet['C']
colonneAppelation = sheet['A']
colonneMillesime = sheet['D']
colonneFormatB = sheet['E']
colonnePrix = sheet['F']

colonneQuantite = 0
colonneConditionnement = 'CBO12'
#iDate_offre = sheet['I']  #validité
#iRegie = sheet['J']
#iType_vendeur = sheet['H']
#iCom = sheet['N']

colonneVinBis = sheet['J']
colonneAppelationBis = sheet['I']
colonneMillesimeBis = sheet['K']
colonneFormatBBis = sheet['L']
colonnePrixBis = sheet['M']


row_count = sheet.max_row
column_count = sheet.max_column

j = 1

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if ((colonneVin[i].value is None) or (colonnePrix[i].value is None) or (colonnePrix[i].value == 'Price') ):
        continue

    vin = unidecode(str(colonneVin[i].value).upper())
        
    millesime = colonneMillesime[i].value

    if colonneMillesime[i].value is None :
        millesime = 'NV'

    prix = colonnePrix[i].value

    if('MAGNUM' in vin):
        formaB = 'MG'
        iVin = vin.find('MAGNUM')
        chateau = vin[0:iVin]
    else:
        formatB = str(colonneFormatB[i].value)
        formatB = fichierFonction_TWINOF.formatterFormatBouteille(formatB)

    quantite:int = 0
    #quantite = int(quantite)

    appellation = colonneAppelation[i].value

    commentaireSupplementaire = ''
    
    if ('MIXED CASE' in appellation):
        for i, c in enumerate(vin):
            if c.isdigit():
                commentaireSupplementaire = vin[i:len(vin)]
                vin = vin[0:i]
                break

    conditionnement = ""

    if(quantite == 0):
        if(formatB == 'DE'):
            conditionnement = 'CBO24DE'
        else:
            conditionnement = 'CBO12'

    # Pour d'éventuel changement
    commentaire = ''
    commentaire = commentaire + commentaireSupplementaire

    # On affecte les valeurs dans l'ordre des colonnes voulues.
    fichierFonction_TWINOF.affectationLignes(j, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaire)

    j = j + 1

    #####
    # La deuxième partie 
    #####

    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if ((colonneVinBis[i].value is None) or (colonnePrixBis[i].value is None) or (colonnePrixBis[i].value == 'Price') ):
        continue

    #print(colonneVinBis[i].value)
    #print(colonnePrixBis[i].value)

    vin = unidecode(str(colonneVinBis[i].value).upper())
        
    millesime = colonneMillesimeBis[i].value

    if colonneMillesimeBis[i].value is None :
        millesime = 'NV'

    prix = colonnePrixBis[i].value

    formatB = str(colonneFormatBBis[i].value)

    if('MAGNUM' in vin):
        formaB = 'MG'
        iVin = vin.find('MAGNUM')
        chateau = vin[0:iVin]
    else:
        formatB = str(colonneFormatB[i].value)
        formatB = fichierFonction_TWINOF.formatterFormatBouteille(formatB)

    quantite = 0
    #quantite = int(quantite)

    appellation = colonneAppelationBis[i].value

    if ('MIXED CASE' in appellation):
        for i, c in enumerate(vin):
            if c.isdigit():
                commentaireSupplementaire = vin[i:len(vin)]
                vin = vin[0:i]
                break
    
    if(quantite == 0):
        if(formatB == 'DE'):
            conditionnement = 'CBO24DE'
        else:
            conditionnement = 'CBO12'

    # Pour d'éventuel changement
    commentaire = ''
    commentaire = commentaire + commentaireSupplementaire

    # On affecte les valeurs dans l'ordre des colonnes voulues.
    fichierFonction_TWINOF.affectationLignes(j, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaire)

    j = j + 1


# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)



   

