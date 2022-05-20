# coding: utf-8
import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook
import random
import pandas as pd

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

nomProfil : str = "CAVHUG"
extensionDebut : str = ".xlsx"
extensionFin_bis : str = '.xlsx'
extensionFin : str = ".csv"

Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC']

# On load le le tarif au format xlsx
for filename in glob.glob('*' + extensionDebut):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename_bis = 'sortie_bis_' + nomProfil + extensionFin_bis
dest_filename = 'sortie_' + nomProfil + extensionFin
ws = wb2.active
ws.title = nomProfil

# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['A']
iMillesime = sheet['B']
iFormatb = sheet['D']
iPrix = sheet['E']
iQte = sheet['C']
#iCond = sheet['H']
#iDate_offre = sheet['I']  #validité
#iRegie = sheet['J']
#iType_vendeur = sheet['H']
#iCom = sheet['N']

row_count = sheet.max_row
column_count = sheet.max_column

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    
    vin = unidecode(str(iVin[i].value).upper())
    prix : str = ft.formaterPrix(str(iPrix[i].value))

    # print("prix : " + prix)

    if( ft.bLigneIncorrecte(prix, ["PRIX(HT)"], vin) ):
        continue
        
    millesime = ft.formaterAnnee(str(iMillesime[i].value))  

    rec_quantite = str(iQte[i].value)
    #quantite = int(quantite)

    #Reconnaissance du vin dans la chaine du nom du vin
    #si on trouve un conditionnement, on cherche le nombre d'unité par conditionnement et on divise le prix par le nombre d'unité / conditionnement
    #On multiplie la quantite par le nombre d'unité par conditionnement (uQte)
    #on nettoie la chaine du vin
    rec_cond = re.findall('[C][BA][OR][T ]?[O]?[N]?[ ]?[D]?[E]?[ ]?\d+?',vin)

    commentaire = ''

    if rec_cond :

        vin = vin.replace(rec_cond[0],'')
        conditionnement = rec_cond[0].replace(' ','').replace('CARTONDE','CC')
        commentaire = ''
        uQte = re.findall('\d+',conditionnement)

        if uQte :
            quantite = int(rec_quantite)*int(uQte[0])
            prix = float(prix) / int(uQte[0])

    # Gestion des ASSORTIMENTS et mise en commentaire du détail des assortiments.
    elif 'ASSORTIMENT' in vin :

        conditionnement = 'COLLEC'
        quantite = rec_quantite
        # rec_assort = re.findall('[ ][A][S][S][O][R][T][I][M][E][N][T][ ]\d+[ ][-]', vin)
        rec_assort = re.findall('ASSORTIMENT', vin)

        if rec_assort :
            commentaire = vin.split(rec_assort[0])[1]

            vin = vin.split(rec_assort[0])[0]+' ASSORTIMENT'

    else:
        quantite = rec_quantite
        conditionnement = 'UNITE'

    # Reconnaissance du format de bouteille, on cherche uniquement les caractères numérique
    rec_formatb = str(iFormatb[i].value)
    iSize = re.findall('\d+',rec_formatb)

    if iSize :
        iSize = str(iSize[0])
        if iSize == '37,5' or iSize == '37.5':
            formatb = 'DE'
        elif iSize == '75' or iSize == '70':
            formatb = 'BO'
        elif iSize == '150':
            formatb = 'MG'
        elif iSize == '300' or iSize == '3':
            formatb = 'DM'
        elif iSize == '600' or iSize == '6':
            formatb = 'IM'
        else:
            formatb = iSize
    else:
        formatb = rec_formatb
    
    #nettoyage du nom du vin
    mill_finder = re.findall('[ ]?[-]?[ ]?[0-9][0-9][0-9][0-9]',vin)
    
    if mill_finder :
        f = vin.find(mill_finder[0])
        vin = vin[:f].strip()

    #On affecte les valeurs dans l'ordre des colonnes voulues.
    c1 = ws.cell(row = i, column = 1)
    c1.value = vin

    c2 = ws.cell(row = i, column = 2)
    c2.value = millesime

    c3 = ws.cell(row = i, column = 3)
    c3.value = formatb

    c4 = ws.cell(row = i, column = 4)
    c4.value = prix

    c5 = ws.cell(row = i, column = 5)
    c5.value = quantite

    c6 = ws.cell(row = i, column = 6)
    c6.value = conditionnement

    c7 = ws.cell(row = i, column = 7)
    c7.value = commentaire

    if  ws.cell(row = i, column = 1).value == ws.cell(row = i-1, column = 1).value and ws.cell(row = i, column = 2).value == ws.cell(row = i-1, column = 2).value and ws.cell(row = i, column = 3).value == ws.cell(row = i-1, column = 3).value :
        
        while ws.cell(row = i, column = 6).value == ws.cell(row = i-1, column = 6).value :

            c7.value = 'ATTN : CONDITIONNEMENT : ' + conditionnement + " : " + commentaire
            c6.value = random.choice(Dict_cond)

#Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename_bis)

read_file = pd.read_excel(dest_filename_bis)
read_file.to_csv(dest_filename, index = False, encoding="utf-8", sep = ';')




   

