# coding: utf-8
import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook
import pandas as pd
import random

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

def conditionnement_par_defaut(formatb,quantite):
    if formatb == 'DE':
        if quantite < 24 :
            conditionnement = 'UNITE'
        else:
            conditionnement = 'CBO24DE'
    elif formatb == 'BO' or 'CL':
        if quantite < 6 :
            conditionnement = 'UNITE'
        elif quantite > 6 and quantite < 12 :
            conditionnement = 'CBO6'
        else:
            conditionnement = 'CBO12'
    elif formatb == 'MG' :
        if quantite < 6 :
            conditionnement = 'UNITE'
        else:
            conditionnement = 'CBO6'
    elif formatb == 'DM':
        if quantite < 3 :
            conditionnement = 'UNITE'
        else:
            conditionnement = 'CBO3'
    else:
        conditionnement = 'CBO1'
    
    return conditionnement

Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC']

# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename = 'sortie_DEWITT.xlsx'
    ws = wb2.active
    ws.title = "DEWITT"

    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    iVin = sheet['A']
    iMillesime = sheet['B']
    iFormatb = sheet['C']
    iPrix = sheet['E']
    iQte = sheet['D']
    iCond = sheet['J']
    iCouleur = sheet['K']

    row_count = sheet.max_row
    column_count = sheet.max_column

    for index_cellule  in range (1, row_count):
        # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
        if iVin[index_cellule].value is None or iFormatb[index_cellule].value is None :
            continue
        else:  
            vin = unidecode(str(iVin[index_cellule].value).upper())
            #print(vin)
            
            millesime = str(iMillesime[index_cellule].value).replace('-','NV')

            prix = unidecode(str(iPrix[index_cellule].value)).replace(' ','').replace('EUR','').replace('.',',')

            quantite = unidecode(str(iQte[index_cellule].value)).replace(' ','')

            formatb = ft.formaterFormatBouteille(iFormatb[index_cellule].value)
            
            conditionnement = str(iCond[index_cellule].value).replace('OWC','CBO').replace('OC','CC').replace('UNIT','UNITE').replace('GB','GIBO')


        c1 = ws.cell(row = index_cellule, column = 1)
        c1.value = vin

        c2 = ws.cell(row = index_cellule, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = index_cellule, column = 3)
        c3.value = formatb

        c4 = ws.cell(row = index_cellule, column = 4)
        c4.value = prix

        c5 = ws.cell(row = index_cellule, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = index_cellule, column = 6)
        c6.value = conditionnement

        # c7 = ws.cell(row = index_cellule, column = 7)
        # c7.value = commentaire

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)
csv_converter = pd.read_excel(dest_filename)
csv_converter.to_csv('sortie_DEWITT.csv', index = False, encoding = "utf-8", sep = ";")



   

