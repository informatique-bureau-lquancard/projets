import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook
import gspread
import pandas as pd

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

nomProfil : str = 'NGVINS'
extensionDebut : str = '.xlsx'
extensionFin_bis : str = '.xlsx'
extensionFin : str = '.csv'

# coding: utf-8

# authentification à google API,  compte j.martin.blq@gmail.com
# fichier credential.json dans le répertoire du script
gc = gspread.service_account(filename='credential.json')

# ouverture du spreadsheet google de Maion B
sh = gc.open_by_key("1XhdHSoBUcvXwTfq_vxluCHw6O_Hrnh89pVO_4KIu7ec")

# Récupération de la liste des onglets
worksheet_list = sh.worksheets()
w =0

writer = pd.ExcelWriter('sortie_NGVINS.xlsx', engine = 'openpyxl')
sheets = []

# Parcours des onglet et création de dataframe Pandas par onglets 
while w in range(len(worksheet_list)) :
    active_worksheet = sh.get_worksheet(w)
    sheet_title = unidecode(str(active_worksheet).split(' ')[1].replace("'",""))
    sheets.append(sheet_title)
    sheets[w] = pd.DataFrame(active_worksheet.get_all_records())
    sheets[w].to_excel(writer,sheet_title, index = False)
    w = w+1

writer.save()


# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename_bis = 'sortie_bis_' + nomProfil + extensionFin_bis
dest_filename = 'sortie_' + nomProfil + extensionFin
ws = wb2.active
ws.title = nomProfil

# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['G']
iRegion = sheet['A']
iCouleur = sheet['C']
iAppellation = sheet['B']
iMillesime = sheet['D']
iFormatB = sheet['F']
iPrix = sheet['H']
iQte = sheet['E']
iCond_formatb = sheet['I']
iCond = sheet['K']
#iRegie = sheet['J']
#iCom = sheet['H']

row_count = sheet.max_row
column_count = sheet.max_column

j : int = 1

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if iVin[i].value is None :
        continue
    elif unidecode(str(iPrix[i].value)).upper() == 'NOUS CONSULTER' or iPrix[i].value is None:
        continue
    elif unidecode(str(iVin[i].value)).upper() == 'APPELLATION':
        continue
    elif unidecode(str(iFormatB[i].value)).upper() is None or 'CL' not in unidecode(str(iFormatB[i].value)).upper() :
        continue
    else:
        vin = unidecode(str(iVin[i].value).upper())
        if '(' in vin :
            iCom = vin.find('(')
            com_vin = vin[int(iCom):]
            vin = vin[:int(iCom)]
        else:
            com_vin = ''

        if iMillesime[i].value is None:
            millesime : str = "NV"
        else:
            millesime : str = ft.formaterAnnee(str(iMillesime[i].value))
     
        prix = unidecode(str(iPrix[i].value)).replace('EUR','').replace(',','').replace('.',',')
        
        formatB : str  = str(iFormatB[i].value).upper()

        if 'X' in formatB :
            formatB = str(iFormatB[i].value).upper().split('X')

        formatB = ft.formaterFormatBouteille(formatB[1])

        quantite : str = unidecode( str(iQte[i].value).upper() )
        quantite = ft.formaterQuantite( quantite )
        
        # Fonction conditionnement par defaut si iCond is None
        if iCond[i].value is None:
            conditionnement = ft.conditionnementParDefaut(formatB, int(quantite))
        else:
            condBis = unidecode(str(iCond[i].value).upper())
            conditionnement = ft.formaterConditionnement(formatB, int(quantite), condBis, vin)
            
        commentaire = ""

        # Fonction Attribution des valeurs dans les bonnes colonnes pour le nouveau workbook
        ft.affectationLignes(j, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaire)

        j = j + 1

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

# Ecriture du nouveau workbook
wb2.save(dest_filename_bis)

read_file = pd.read_excel(dest_filename_bis)
read_file.to_csv(dest_filename, index = False, encoding="utf-8", sep = ';')