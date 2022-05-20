# coding: utf-8
from audioop import add
import csv
import glob
import codecs
from unidecode import unidecode
from time import sleep
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

nomProfil = "CMVINS"
extensionDebut = ".xls"
extensionFin = ".csv"

#Conversion du fichier xls en xlsx
for xls_file in glob.glob("*"+extensionDebut):
    reader = pd.read_excel(xls_file)
    reader.to_excel(nomProfil + ".xlsx", index = False, engine = 'openpyxl')

sleep(1)

# On load le le tarif au format xlsx dans un workbook
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename_bis = 'sortie_'+ 'bis_' + nomProfil.upper()+extensionFin
    dest_filename = 'sortie_'+ nomProfil.upper()+extensionFin
    ws = wb2.active
    ws.title = "CMVINS"

    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    iVin = sheet['B']
    iAppellation = sheet['D']
    iMillesime = sheet['C']
    iFormatb = sheet['F']
    iPrix = sheet['G']
    iQte = sheet['E']
    iCond = sheet['H']
    istatut_offre = sheet['A']  #validité

    bordeaux_dict = ['PAUILLAC','HAUT MEDOC','ST EMILION GCC','AUSONE', 'MOUTON', 'CHEVAL', 'MISSION', 'YQUEM', 'PETRUS', 'HTBRION', 'LAFITE', 'MARGAUX','PESSAC','ST JULIEN','POMEROL','ST ESTEPHE','MOULIS']

    #iRegie = sheet['J']
    #iType_vendeur = sheet['H']
    iCom = sheet['I']

    row_count = sheet.max_row
    column_count = sheet.max_column

    j : int = 1

    for i  in range (1, row_count):
        # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
        millesime = iMillesime[i].value
        prix = iPrix[i].value
        appellation = unidecode(str(iAppellation[i].value)).upper()
        quantite = iQte[i].value

        if prix is None or prix == "€ / unité":
            continue

        for a in bordeaux_dict :
            if a in appellation :
                vin = unidecode(str(iVin[i].value)).upper()
                break
            else:
                vin = appellation+' '+unidecode(str(iVin[i].value)).upper()

        if millesime is None :
            millesime = 'NV'
        else:
            millesime = iMillesime[i].value
        
        formatB : str = ft.formaterFormatBouteille(iFormatb[i].value)

        conditionnement : str = iCond[i].value        

        if(conditionnement is None):
            conditionnement = "UNITE"
        else:
            rec_cond = str(unidecode(iCond[i].value)).upper()
            if 'CB' in rec_cond :
                conditionnement = rec_cond.replace('CB','CBO')
                if formatB == 'DE':
                    conditionnement = conditionnement+'DE'
            elif 'CCO' in rec_cond :
                conditionnement = rec_cond.replace('CCO','CC')
                if formatB == 'DE':
                    conditionnement = conditionnement+'DE'
            elif 'COFFRET' in rec_cond :
                conditionnement = 'COF'
            elif 'COLLECTION' in vin :
                conditionnement = 'COLLEC'
            else:
                conditionnement = 'UNITE'
        
        save_cond = conditionnement
        conditionnement = conditionnement.split(' ')
        
        if len(conditionnement) > 1 :
            com_Finder = int(save_cond.find(' '))       
            add_com = save_cond[com_Finder:]
        else:      
            add_com = ''
        
        conditionnement = conditionnement[0]

        if(iCom[i].value is None):
            rec_com = ''
        else:
            rec_com = str(unidecode(iCom[i].value))
        
        commentaire = add_com+' '+rec_com

        #On affecte les valeurs dans l'ordre des colonnes voulues.
        ft.affectationLignes(j, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaire)
        j += 1

#Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename_bis)

read_file = pd.read_excel(dest_filename_bis)
read_file.to_csv(dest_filename, index = False, encoding="utf-8", sep = ';')