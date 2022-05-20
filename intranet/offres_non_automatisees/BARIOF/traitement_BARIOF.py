# coding: utf-8
import csv
import glob
import codecs
from unidecode import unidecode
import re
import openpyxl
from openpyxl import Workbook


# On load le le tarif au format xlsx
for filename in glob.glob('*.xlsx'):
    wb = openpyxl.load_workbook(filename)

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_BARIOF.xlsx'
ws = wb2.active
ws.title = "BARIOF"


# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

iVin = sheet['B']
iMillesime = sheet['C']
iFormatb = sheet['E']
iPrix = sheet['D']
#Qte = 0
iCond = sheet['F']
#iDate_offre = sheet['I']  #validité
#iRegie = sheet['J']
#iType_vendeur = sheet['H']
#iCom = sheet['N']

row_count = sheet.max_row
column_count = sheet.max_column

for i  in range (1, row_count):
    # si les cellules de la premiere colonne sont vide on n'écrit pas de façon à avoir "None" dans le workbook de sortie.
    if iVin[i].value is None :
        continue
    else:  
        vin = unidecode(str(iVin[i].value).upper())
        
        millesime = iMillesime[i].value

        prix = iPrix[i].value

        iSize = str(iFormatb[i].value).upper()
        if iSize == 'DEMIBT' :
            formatb = 'DE'
        elif iSize == 'BTL':
            formatb = 'BO'
        elif iSize == 'MGN':
            formatb = 'MG'
        elif iSize == 'DMGN':
            formatb = 'DM'
        elif iSize == 'DMGN':
            formatb = 'DM'
        elif iSize == 'JERO':
            formatb = 'JE'
        elif iSize == 'MARJEA':
            formatb = 'MJ'
        elif iSize == 'SALM':
            formatb = 'SA'
        elif iSize == 'IMP':
            formatb = 'IM'
        else:
            formatb = iSize

        quantite = 0
        #quantite = int(quantite)

        rec_cond = str(iCond[i].value).upper().replace('OWC','CBO').replace('0','').replace('OW','CBO').replace('PK','CC').replace('CT','CC').replace('12CC','CC12').replace('6CC','CC6').replace(' ','').replace('CARDBOARD','')
        if 'CBO' in rec_cond :
            uqte = re.findall('\d+', rec_cond)[0]
            conditionnement = 'CBO'+uqte
        elif 'CC' in rec_cond :
            uqte = re.findall('\d+', rec_cond)[0]
            conditionnement = 'CC'+uqte
        else :
            conditionnement = 'UNITE'

        # On affecte les valeurs dans l'ordre des colonnes voulues.
        c1 = ws.cell(row = i, column = 1)
        c1.value = vin

        c2 = ws.cell(row = i, column = 2)
        c2.value = millesime

        c3 = ws.cell(row = i, column = 3)
        c3.value = formatb

        c4 = ws.cell(row = i, column = 4)
        c4.value = prix

        c5 = ws.cell(row = i, column = 5)
        c5.value = quantite

        c6 = ws.cell(row = i, column = 6)
        c6.value = conditionnement

        # c7 = ws.cell(row = i, column = 7)
        # c7.value = commentaire

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)



   

