## coding: utf-8
import csv
import glob
from tkinter.ttk import Separator
from unidecode import unidecode

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

import pandas as pd
import openpyxl
from openpyxl import Workbook

from time import sleep

# Revoir suppression des fichiers et traitement sur site web

nomProfil = "REDCIR"
extensionDebut = ".xlsx"
extensionFin = ".csv"

designation_prix1 = "PVUnitaire"
designation_prix2 = "PrixdeVente"

designation_prix = [designation_prix1, designation_prix2]

suppression_designation = [",", "(ELA)", "(US LABEL)", "(ETA)"]

# def csv_from_excel(wb : Workbook, chemin_de_sortie : str):

#     # workbook = xlsxwriter.Workbook(chemin_de_sortie)

#     # sh = workbook.add_worksheet()

#     # book = xlrd.open_workbook(chemin_de_sortie)
#     # sh = book.sheet_by_index(1)

#     # wb = xlrd.open_workbook('excel.xlsx')
#     sh = wb.sheet_by_name(nomProfil)
#     sh = wb.create_sheet(nomProfil)

#     your_csv_file = open(chemin_de_sortie, 'w')
#     wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

#     for rownum in range(sh.nrows):
#         wr.writerow(sh.row_values(rownum))

#     your_csv_file.close()

# On load le le tarif au format xlsx
for filename in glob.glob('*'+extensionDebut):

    wb = openpyxl.load_workbook(filename)

    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename_bis = 'sortie_'+ 'bis_' + nomProfil + extensionFin
    dest_filename = 'sortie_'+ nomProfil + extensionFin
    ws = wb2.active
    ws.title = nomProfil

    #A revoir si le tableau tab_onglets n'est pas renseigné
    index_onglet_sortie = 0

    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.

    nombre_onglets_entree = len(sheet_list)

    sheets = []

    # Utile pour la maintenance et pour automatiser le processus
    index_onglet = 1

    # Lecture du workbook d'origine
    # tab_sheet_name_ok = ['Bordeaux CBO', 'Bordeaux Vrac', 'Bordeaux Primeurs 2020', 'Hors-Bordeaux', 'Stock Dugat-Py', 'Stock E. Reynaud']
    tab_sheet_name_ok = ['in']

    j : int = 1

    for sheet in wb:
  
        # Essai non concluant

        #   titre_onglet : str = sheet.title.replace(" ", "")

        # if(index_onglet == 1):
        #     tab_colonnes_onglet_actif = ["A", "D", "F", "G", "E", "I", "B", "C"]

        # if(index_onglet == 2):
        #     tab_colonnes_onglet_actif = ["A", "D", "F", "G", "E", "I", "B", "C"]

        # if(index_onglet == 3):
        #     tab_colonnes_onglet_actif = ["A", "D", "F", "G", "E", "", "B", "C"]

        # if(index_onglet == 4):
        #     tab_colonnes_onglet_actif = ["A", "E", "G", "H", "F", "J", "C", "D"]

        # if(index_onglet == 5):
        #     tab_colonnes_onglet_actif = ["A", "D", "F", "H", "G", "E", "B", "C"]

        # if(index_onglet == 6):
        #     tab_colonnes_onglet_actif = ["A", "C", "E", "G", "F", "D", "", "C"]
        
        # dictionnaire = dict(zip (tab_designations_colonnes, tab_colonnes_onglet_actif))
            
        # dataframe = ff.formatterColonnesOnglets(dictionnaire, sheet)

        # print(dataframe)
        # print(dataframe['VIN'][0])

        # vin, millesime, formatB, prix, quantite, conditionnement, commentaire, appellation, couleur
  
        erreurs = []

        if(sheet.title not in tab_sheet_name_ok):
            erreurs.append(f"sheet not OK : {sheet.title}")
            continue

        if(sheet.title == tab_sheet_name_ok[0]):

            iVin : tuple = sheet['A']

            iMillesime : tuple = sheet['C']
            iFormatB : tuple = sheet['E']
            iPrix : tuple = sheet['G']
            iQuantite : tuple = sheet['F']
            iConditionnement : tuple = sheet['D']
            iAppellation : str = ""
            iCouleur : tuple = sheet['B']
        else:
            erreurs.append(f"Erreurs : {sheet.title}")

        # if(sheet.title == tab_sheet_name_ok[0]):

        #     iVin : tuple = sheet['A']

        #     iMillesime : tuple = sheet['D']
        #     iFormatB : tuple = sheet['F']
        #     iPrix : tuple = sheet['G']
        #     iQuantite : tuple = sheet['E']
        #     iConditionnement : tuple = sheet['I']
        #     iAppellation : str = ""
        #     iCouleur : tuple = sheet['C']

        # elif(sheet.title == tab_sheet_name_ok[1]):

        ## Problème de maintabilité
        # if(sheet.title == tab_sheet_name_ok[0]):

        #     iVin : tuple = sheet['A']

        #     iMillesime : tuple = sheet['D']
        #     iFormatB : tuple = sheet['F']
        #     iPrix : tuple = sheet['G']
        #     iQuantite : tuple = sheet['E']
        #     iConditionnement : tuple = sheet['I']
        #     iAppellation : str = ""
        #     iCouleur : tuple = sheet['C']

        # elif(sheet.title == tab_sheet_name_ok[1]):

        #     iVin : tuple = sheet['A']

        #     iMillesime : tuple = sheet['D']
        #     iFormatB : tuple = sheet['F']
        #     iPrix : tuple = sheet['G']
        #     iQuantite : tuple = sheet['E']
        #     iConditionnement : tuple = sheet['I']
        #     iAppellation : str = ""
        #     iCouleur : tuple = sheet['C']
        
        # elif(sheet.title == tab_sheet_name_ok[2]):
            
        #     iVin : tuple = sheet['A']

        #     iMillesime : tuple = sheet['D']
        #     iFormatB : tuple = sheet['F']
        #     iPrix : tuple = sheet['G']
        #     iQuantite : tuple = sheet['E']
        #     iConditionnement : str = ""
        #     iAppellation : str = ""
        #     iCouleur : tuple = sheet['C']
        
        # elif(sheet.title == tab_sheet_name_ok[3]):
            
        #     iVin : tuple = sheet['A']

        #     iMillesime : tuple = sheet['E']
        #     iFormatB : tuple = sheet['G']
        #     iPrix : tuple = sheet['H']
        #     iQuantite : tuple = sheet['F']
        #     iConditionnement : tuple = sheet['J']
        #     iAppellation : str = ""
        #     iCouleur : tuple = sheet['D']
        
        # elif(sheet.title == tab_sheet_name_ok[4]):
                
        #     iVin : tuple = sheet['A']

        #     iMillesime : tuple = sheet['D']
        #     iFormatB : tuple = sheet['F']
        #     iPrix : tuple = sheet['H']
        #     iQuantite : tuple = sheet['G']
        #     iConditionnement : tuple = sheet['E']
        #     # Hors Bordeaux pour EV
        #     iAppellation : tuple = sheet['B']
        #     iCouleur : tuple = sheet['C']

        # elif(sheet.title == tab_sheet_name_ok[5]):
                
        #     iVin : tuple = sheet['A']

        #     iMillesime : tuple = sheet['C']
        #     iFormatB : tuple = sheet['E']
        #     iPrix : tuple = sheet['G']
        #     iQuantite : tuple = sheet['F']
        #     iConditionnement : tuple = sheet['D']
        #     iAppellation : str = ""
        #     iCouleur : tuple = sheet['B']

        # else:
        #     erreurs.append(f"Erreurs : {sheet.title}")
        
        row_count = sheet.max_row
        column_count = sheet.max_column

        for i in range (1, row_count):
            # prix
            prix = unidecode(str(iPrix[i].value))

            # formatte le prix pour qu'il ne reste que des chiffres
            prix = ft.formaterPrix(prix)

            # vin
            vin : str = unidecode(str(iVin[i].value)).upper()

            if(ft.bLigneIncorrecte(prix, designation_prix, vin)):
                continue

            # appellation
            appellation = ""
            
            if(type(iAppellation) is not str):
                appellation = str(iAppellation[i].value).upper()
                
                #charteau + appellation
                vin = vin + " " + appellation

            # chateau + couleur
            couleur = str(iCouleur[i].value).replace(' ', '')

            if(couleur == 'Blanc'):
                vin = vin + ' ' + couleur
                
            # millesime
            millesime = iMillesime[i].value
            millesime = ft.formaterAnnee(millesime)

            #format
            formatB : str = iFormatB[i].value

            # A revoir pour le moment pour les format on teste si le formatB se retrouve 
            # dans la chaine de caractere de tableau de format fixe.
            # L'inverse peut-être intéressant à mettre en place pour éviter le souci d'en dessous
            
            index_formatB : int

            if("MAG" in formatB):

                if("DMAG" in formatB):
                    index_formatB = formatB.find("(DMAG)")
                else:
                    index_formatB = formatB.find("(MAG)")

                formatB = formatB[0 : index_formatB]

            if("x" in formatB):
                index_formatB = formatB.find("x")

                quantite = formatB[0:index_formatB]

                formatB = formatB[index_formatB +1 : len(formatB)] 

            formatB = ft.formaterFormatBouteille(formatB)

            # quantite
            quantite : str = str(iQuantite[i].value)
            quantite = quantite.replace(' ', '')
            quantite = ft.formaterQuantite(quantite)

            index_vin : int = ft.testVinCollection(vin)

            # conditionnement
            conditionnement : str = ""

            if(index_vin != -1):

                conditionnement = "COLLEC"
                vin = vin[0 : index_vin - 1]

            # if(ff.existe(vin, suppression_designation) != -1):
            vin  = ft.supprimeChaineCaracteres(vin,suppression_designation)

            if(conditionnement == "" ):

                if(type(iConditionnement) is not str):
                    conditionnement : str = str(iConditionnement[i].value).upper()

                if("COFFRET" in conditionnement):
                    conditionnement = "CF"+ quantite
                else:
                    conditionnement = ft.formaterConditionnement(formatB, int(quantite), conditionnement, vin)

                index_conditionnement : int = conditionnement.find("(BANDED)")
            
                commentaire : str = ""

                if(index_conditionnement != -1):
                    commentaire = conditionnement[index_conditionnement : len(conditionnement)]
                    conditionnement = conditionnement[0 : index_conditionnement]

            # commentaire
            commentaire += ""

            # On affecte les valeurs dans l'ordre des colonnes voulues.
            ft.affectationLignes(j, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaire)

            j = j + 1

            # Revoir ce qui est intéressant t ce qui ne l'est pas

            # # Ancien avec fichier csv
            # # écriture de la ligne
            # # On fabrique la nouvelle ligne dans l'ordre voulu
            # newRow=[chateau,annee,formatB,prix,quantite,conditionnement,commentaire]
            # writer.writerow(newRow)

            # monFichierEntre.close()

        # Fonction de suppression des lignes vides d'un workbook
        # index_row = []
        # for n in range(1, ws.max_row):
        #     if ws.cell(n, 1).value is None:
        #         index_row.append(n)

        # for row_del in range(len(index_row)):
        #     ws.delete_rows(idx=index_row[row_del], amount=1)
        #     index_row = list(map(lambda k: k -1, index_row))

        wb2.save(dest_filename_bis)

        index_onglet += 1

        read_file = pd.read_excel(dest_filename_bis)
        read_file.to_csv(dest_filename, index = False, encoding="utf-8", sep = ';')
