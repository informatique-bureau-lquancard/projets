import csv
import glob
from importlib.machinery import ExtensionFileLoader
import re
from unidecode import unidecode
import openpyxl_dictreader
import openpyxl
from openpyxl import Workbook
import requests as req
import pandas as pd

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft
# coding: utf-8

URL = 'https://facebook.us15.list-manage.com/track/click?u=677339c726d03928bcb98cb80&id=9bccd7cd21&e=6435d1cb63'
file = req.get(URL, allow_redirects=True, verify = False)

content = file.headers['Content-Type']

if content == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' :
    extension = 'xlsx'
elif content == 'application/excel' :
    extension = 'xls'
else:
    print('error : extension not reconized')

open('Tarifs.'+extension, 'wb').write(file.content)

## Reconnaissance de l'extension du fichier
### @ Clement : Faire une fonction globale #####
if extension == 'xls' :
    reader = pd.read_excel('Tarifs.xls', header = 4)
    reader.dropna(how = "all", inplace = True)
    reader.to_excel('Tarifs.xlsx', index=False, header=False)


file = openpyxl_dictreader.DictReader('Tarifs.xlsx')

wb = openpyxl.load_workbook('Tarifs.xlsx')

sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

field_list = file.fieldnames
dic_alpha = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

# On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_BORBLE.xlsx'
ws = wb2.active
ws.title = "BORBLE"

row_count = sheet.max_row
column_count = sheet.max_column

## Recherche index du fieldname dans list_fieldnames ##
col_Vin = field_list.index("Vin / Wine")
col_Millesime = field_list.index("Mill./Vintage")
col_Appellation = field_list.index("Appellation")
col_Prix = field_list.index("EXW Price")
col_Formatb = field_list.index("Format")
col_Conditionnement = field_list.index("Condit. Packing")
col_Quantite = field_list.index("Quantité / Quantity")

## Faire correspondre l'index des fieldname avec la lettre de la colonne du Workbook (A, B, C etc...)
iVin = sheet[dic_alpha[col_Vin]]
iMillesime = sheet[dic_alpha[col_Millesime]]
iAppellation = sheet[dic_alpha[col_Appellation]]
iPrix = sheet[dic_alpha[col_Prix]]
iFormatb = sheet[dic_alpha[col_Formatb]]
iConditionnement = sheet[dic_alpha[col_Conditionnement]]
iQuantite = sheet[dic_alpha[col_Quantite]]

## Les commentaires n'on pas de fieldname, correspondance est faite avec le nombre max de colonnes
iCom1 = sheet[dic_alpha[column_count-2]]
iCom2 = sheet[dic_alpha[column_count-1]]

j = 1

for i in range(1, row_count):
    
    vin = str(iVin[i].value).upper().strip().replace('"','')
    millesime = iMillesime[i].value
    formatB = ft.formaterFormatBouteille(str(iFormatb[i].value))
    prix = iPrix[i].value
    quantite = iQuantite[i].value
    conditionnement = ft.formaterConditionnement(formatB,quantite, str(iConditionnement[i].value), vin)
    if iCom1[i].value is None and iCom2[i].value is not None :
        commentaires = str(iCom2[i].value)
    elif iCom1[i].value is not None and iCom2[i].value is None :
        commentaires = str(iCom1[i].value)
    elif iCom1[i].value is not None and iCom2[i].value is not None :
        commentaires = str(iCom1[i].value+' '+str(iCom2[i].value))
    else:
        commentaires = ''

    #On affecte les valeurs dans l'ordre des colonnes voulues.
    ft.affectationLignes(j, ws, vin, millesime, formatB, prix, quantite, conditionnement, commentaires)
    j += 1

#Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)

