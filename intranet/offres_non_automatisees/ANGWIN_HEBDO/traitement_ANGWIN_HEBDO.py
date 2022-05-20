# coding: utf-8
import csv
import os
import glob
import re
import io
import time
from typing import ItemsView
from unidecode import unidecode
from lxml import etree
import sys
import xml.etree.ElementTree as ET
import urllib
import urllib.request
from urllib.request import Request, urlopen
import feedparser
import pandas as pd
import openpyxl
from openpyxl import Workbook
import random

url = 'https://www.vinsetmillesimes.com/gmerchantcenterpro87f0116d88fbb4ac32bbbe5f4fd44afc.fr.EUR.shop1.product.xml'
feed = feedparser.parse(url)

nom_sortie = 'sortie_ANGWIN_HEBDO.csv'

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

def conditionnement_par_defaut(formatb,quantite):
    if formatb == 'DE':
        if quantite < 24 :
            conditionnement = 'UNITE'
        else:
            conditionnement = 'CBO24DE'
    elif formatb == 'BO' or 'CL':
        if quantite < 12 :
            conditionnement = 'UNITE'
        else:
            conditionnement = 'CBO12'
    elif formatb == 'MG' :
        if quantite < 6 :
            conditionnement = 'UNITE'
        else:
            conditionnement = 'CBO6'
    elif formatb == 'DM':
        if quantite < 3 :
            conditionnement = 'UNITE'
        else:
            conditionnement = 'CBO3'
    else:
        conditionnement = 'CBO1'
    
    return conditionnement

with open(nom_sortie,'w',newline='', encoding='utf-8') as monFichierSortie:
    writer = csv.writer(monFichierSortie, delimiter=';', quoting=csv.QUOTE_NONE, escapechar=';', quotechar='')
    

    i : int = 0
    for entry in feed.entries:
        item = entry.description

        vin : str = unidecode(str(item)).upper()

        # print("vin : " + vin)

        prix : str = ft.formaterPrix(str(entry.g_price))
        

        designation_prix = ["€/unit"]

        if( ft.bLigneIncorrecte(prix, designation_prix, vin)):
            continue;

        # prix TTC
        # prix = entry.g_price.replace(' EUR','')
        # prix HT
        prixTTC : float = round(float(prix.replace(',', '.')), 2)
        
        # prix_int : float =  prix_float
        
        
        # if( int(prix_int) == prix_int ) :
        #     prix_int = int(prix_int)

        # prix = str(prix_int).replace(".", ',')

        prix_au_rabais : str = ""

        

        if ( "g_sale_price" in entry ):
            prix_au_rabais : str = entry.g_sale_price.replace(' EUR','').replace('.',',')
            prixTTC : float = round(float(prix_au_rabais.replace(',', '.')), 2)
            prixINIT : str = ft.formaterPrix(str(entry.g_price))
        
        prix = prixTTC / 1.2

        

        quantite = entry.g_sell_on_google_quantity
        couleur = entry.g_color
    
        rec_millesime = re.findall('[0-9][0-9][0-9][0-9]',item)
        
        if rec_millesime :
            millesime = rec_millesime[0]
        else:
            millesime = 'NV'

        commentaire : str = ""       

        iCom_Vin = ''
        rec_assort = 0
        if("ASSORTIMENT" in vin):
            rec_assort = 1
            liste_vin = vin.split('ASSORTIMENT')
            if len(liste_vin) == 2 :
                if 'LOT ' in liste_vin[0]:
                    vin = liste_vin[1]
                    rec_details = re.findall('\d+[ ][B][O][U][T][E][I][L][L][E][S][ ]', vin)
                    if rec_details :
                        vin = vin.replace(rec_details[0],'')
                        iCom_Vin = liste_vin[0]+rec_details[0]
                else:
                    vin = liste_vin[0]
                    iCom_Vin = liste_vin[1]
            vin += ' ASSORTIMENT'
        
        vin = vin.replace(' '+millesime, '').replace(';','').replace('"','').strip()


        
        rec_formatb = re.findall('[0-9]?[,]?[0-9][eE0-9][cC]?[lL]',item)
        formatb_dict = ['Magnum','Double Magnum', 'Jeroboam', 'Imper']
        formatb = ''
        for f in formatb_dict :
            if f in item :
                iFormat = item.find(f)
                formatb = item[int(iFormat):]
            else:
                break

        if rec_formatb and formatb == '':
            formatb = rec_formatb[0]
        elif 'Imperiale' in item :
            formatb = 'IM'
        elif formatb == '' :
            formatb = 'BO'

        formatb = formatb.replace('Double Magnum','DM').replace('Magnum','MG').replace('Jeroboam','JE').replace('62cl','BO').replace('70cl','BO').replace('300cl','DM').replace('300CL','DM').replace(' 3L','')
        
        if rec_assort == 1 :
            conditionnement = 'COLLEC'
            commentaire += ' '+iCom_Vin
        else:
            conditionnement = conditionnement_par_defaut(formatb,int(quantite))
            commentaire = 'Verif CDT ' + commentaire
        
        if( prix_au_rabais != ""):
            commentaire += "Remise speciale, prix initial : " + prixINIT
        
        if 'TACHE' in vin :
            print(vin, millesime, formatb, prix, quantite, conditionnement, commentaire)
        
        newRow=[vin, millesime, formatb, prix, quantite, conditionnement, commentaire]
        writer.writerow(newRow)

reader = pd.read_csv('sortie_ANGWIN_HEBDO.csv', sep = ';', header = None)
reader.to_excel('sortie_ANGWIN_HEBDO_temp.xlsx', header = None, index = None)
new_reader = pd.read_excel('sortie_ANGWIN_HEBDO_temp.xlsx', header = None)
new_reader.sort_values(by = 0, inplace=True)
new_reader.to_excel('sortie_ANGWIN_HEBDO_temp.xlsx', header = None, index = None)


wb = openpyxl.load_workbook('sortie_ANGWIN_HEBDO_temp.xlsx')

#On prépare un nouveau workbook
wb2 = Workbook()
dest_filename = 'sortie_ANGWIN_HEBDO.xlsx'
ws = wb2.active
ws.title = 'ANGWIN'

# Lecture du workbook d'origine
sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
sheet = wb[sheet_list[0]]

row_count = sheet.max_row
column_count = sheet.max_column

Dict_cond = ['UNITE','CC1','CC2','CC3','CC4','CC6','CC12','CC24DE','CBO1','CBO2','CBO3','CBO6','CBO12','CBO24DE','COLLEC']

vin = sheet['A']
millesime = sheet['B']
formatb = sheet['C']
prix = sheet['D']
quantite = sheet['E']
conditionnement = sheet['F']
commentaire = sheet['G']

for i  in range (1, row_count):

    c1 = ws.cell(row = i, column = 1)
    c1.value = vin[i].value

    c2 = ws.cell(row = i, column = 2)
    c2.value = millesime[i].value

    c3 = ws.cell(row = i, column = 3)
    c3.value = formatb[i].value

    c4 = ws.cell(row = i, column = 4)
    c4.value = prix[i].value

    c5 = ws.cell(row = i, column = 5)
    c5.value = quantite[i].value

    c6 = ws.cell(row = i, column = 6)
    c6.value = conditionnement[i].value

    c7 = ws.cell(row = i, column = 7)
    c7.value = commentaire[i].value

    if i == 1:
        i = i
    else:
        if  ws.cell(row = i, column = 1).value == ws.cell(row = i-1, column = 1).value and ws.cell(row = i, column = 2).value == ws.cell(row = i-1, column = 2).value and ws.cell(row = i, column = 3).value == ws.cell(row = i-1, column = 3).value :
            while ws.cell(row = i, column = 6).value == ws.cell(row = i-1, column = 6).value :
                ws.cell(row = i, column = 6).value = random.choice(Dict_cond)
                ws.cell(row = i, column = 7).value = 'ATTN : VERIF CDT - DOUBLON LIGNE PRECEDENTE - '+str(ws.cell(row = i, column = 7).value)

wb2.save(dest_filename)
# new_reader.to_excel('sortie_ANGWIN_HEBDO_bis.xlsx', header = None, index = None)

