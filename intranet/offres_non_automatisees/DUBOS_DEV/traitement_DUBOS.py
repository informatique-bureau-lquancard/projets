# -*- coding: utf-8 -*-
import csv
import glob
import openpyxl
import openpyxl_dictreader
from openpyxl import Workbook
import re
# coding: utf-8

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

j : int = 1
for filename in glob.glob('*.xlsx'):
    file = openpyxl_dictreader.DictReader(filename)

    wb = openpyxl.load_workbook(filename)

    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    field_list = file.fieldnames
    print(field_list)
    
    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename = 'sortie_DUBOS.xlsx'
    ws = wb2.active
    ws.title = "DUBOS"

    row_count = sheet.max_row
    column_count = sheet.max_column

    veriftab = ['Sauternes', 'Barsac']
    for row in file :
        
        vin = row['NomCourt']
        if 'Couleur' == 'Blanc' and appellation not in veriftab :
            vin = vin+' Blanc'  
        
        millesime = row['Mill.']
        formatb = row['Format']
        couleur = row['Couleur']
        appellation = row['AppellationArt']
        caissage = row['Caissage']
        prix = row['Prix Place Dubos']
        quantite = row['Qaffectées']

        if vin is None :
            continue

        formatb = formatb.split(" ")
        f = len(formatb)
        formatb = formatb[f-2]
        formatb = ft.formaterFormatBouteille(formatb)

        prix = str(prix).replace('.',',')

        conditionnement = 'CBO'+str(caissage)
        if formatb == 'DE':
            conditionnement = conditionnement+'DE'

        commentaires = ''
        
        ft.affectationLignes(j, ws, vin, millesime, formatb, prix, quantite, conditionnement, commentaires)
        j += 1

#Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)
        
        
    




