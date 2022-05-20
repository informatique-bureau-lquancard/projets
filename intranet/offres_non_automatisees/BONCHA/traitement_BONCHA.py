# coding: utf-8
import csv
import glob
import codecs
from time import sleep
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook
from unidecode import unidecode

import sys
sys.path.append("/var/www/html/php/fonctions_tarifs")
import Fonction_tarifs as ft

nomProfil : str = "BONCHA"
extensionDebut : str = ".xls"
extensionFin : str = ".xlsx"

# Conversion du fichier xls en xlsx
for xls_file in glob.glob('*' + extensionDebut):
    reader = pd.read_excel(xls_file)
    reader.to_excel(nomProfil + extensionFin, index = False)

# sleep(1)

# On load le le tarif au format xlsx dans un workbook
for filename in glob.glob('*' + extensionFin):
    wb = openpyxl.load_workbook(filename)

    # On prépare un nouveau workbook
    wb2 = Workbook()
    dest_filename = 'sortie_' + nomProfil + extensionFin
    ws = wb2.active
    ws.title = nomProfil

    # Lecture du workbook d'origine
    sheet_list = wb.sheetnames  # on récupère le nom des onglets, pas necessaire ici.
    sheet = wb[sheet_list[0]]

    Mill_Domaine_Appellation = sheet['B']
    Vin = sheet['C']
    Formatb = sheet['D']
    Prix = sheet['F']
    Qte = sheet['E']
    Cond = sheet['G']

    row_count = sheet.max_row
    column_count = sheet.max_column

    for index_ligne  in range (1, row_count) :
        if Mill_Domaine_Appellation[index_ligne].value is None :
            continue
        if 'Vintage' in str(Mill_Domaine_Appellation[index_ligne].value):
            continue

        domaine : str = ""

        if Vin[index_ligne].value is None :    
            if "Rouge" in Mill_Domaine_Appellation[index_ligne].value or "Blanc" in Mill_Domaine_Appellation[index_ligne].value or "Rosé" in Mill_Domaine_Appellation[index_ligne].value :
                appellation = Mill_Domaine_Appellation[index_ligne].value
                continue

            if "Rouge" not in Mill_Domaine_Appellation[index_ligne].value or "Blanc" not in Mill_Domaine_Appellation[index_ligne].value or "Rosé" not in Mill_Domaine_Appellation[index_ligne].value :
                if 'D.R.C.' in Mill_Domaine_Appellation[index_ligne].value :
                    domaine = 'DRC'
                    commentaire = str(Mill_Domaine_Appellation[index_ligne].value)[6:]
                else:
                    domaine = Mill_Domaine_Appellation[index_ligne].value
                    commentaire = ''

                continue
        
        #Dict_bourgognes_blanc = ['Chassagne-Montrachet', 'Corton-Charlemagne','Montrachet', 'Puligny-Montrachet Folatières', 'Puligny-Montrachet La Garenne', 'Bouzeron', 'Aligoté', 'Chablis']

        ### Construction du vin ####
        i_Vin = Vin[index_ligne].value

        if( len(domaine) < 1 ):
            vin = i_Vin
        else:
            vin = domaine+' '+i_Vin

        if appellation == "Bourgogne Blanc" :
            vin = vin + ' blanc'

        #### Millesime ####
        millesime : str = ft.formaterAnnee( str( Mill_Domaine_Appellation[index_ligne].value ) )

        #### Format bouteille ####
        i_Formatb = str(Formatb[index_ligne].value).replace(' ','').upper()
        formatb = ft.formaterFormatBouteille(i_Formatb)

        #### Prix ####       
        prix = Prix[index_ligne].value

        #### Quantite ####
        quantite = Qte[index_ligne].value

        #### Conditionnement ####
        i_Cond = str(Cond[index_ligne].value).upper()
        rec_Cond = re.findall('[O][CW][C]?',i_Cond)
        rec_uQte = re.findall('\d+', i_Cond)
        
        
        if rec_Cond :            
            cond_found = rec_Cond[0].replace('OWC','CBO').replace('OC','CC')
            if rec_uQte :
                conditionnement = cond_found+rec_uQte[0]
                commentaire = unidecode(i_Cond.replace(rec_Cond[0],'').replace(rec_uQte[0],'')).strip()
            else:
                conditionnement = rec_Cond[0]
                commentaire = unidecode(i_Cond.replace(rec_Cond[0],'')).strip()
                       
        else:
            if 'UNIT' in i_Cond :
                conditionnement = 'UNITE'
                if 'DRC' not in vin :
                    commentaire = i_Cond.replace('UNIT','').strip()
            elif 'GIFT BOX' in i_Cond :
                conditionnement = 'GIBO1'
                commentaire = i_Cond.replace('GIFT BOX','').strip()
            else:
                conditionnement = 'UNITE'
                commentaire = i_Cond

        vin = unidecode(vin).replace("'","").upper()  
        commentaire = commentaire.lower()
        #print(vin, millesime, prix, formatb, quantite, conditionnement, commentaire)

        #On affecte les valeurs dans l'ordre des colonnes voulues.
        c1 = ws.cell(row = index_ligne, column = 1)
        c1.value = vin
        c2 = ws.cell(row = index_ligne, column = 2)
        c2.value = millesime
        c3 = ws.cell(row = index_ligne, column = 3)
        c3.value = formatb
        c4 = ws.cell(row = index_ligne, column = 4)
        c4.value = prix
        c5 = ws.cell(row = index_ligne, column = 5)
        c5.value = quantite
        c6 = ws.cell(row = index_ligne, column = 6)
        c6.value = conditionnement
        c7 = ws.cell(row = index_ligne, column = 7)
        c7.value = commentaire

# Fonction de suppression des lignes vides d'un workbook
index_row = []
for n in range(1, ws.max_row):
    if ws.cell(n, 1).value is None:
        index_row.append(n)

for row_del in range(len(index_row)):
    ws.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k -1, index_row))

wb2.save(dest_filename)
csv_converter = pd.read_excel(dest_filename)
csv_converter.to_csv('sortie_BONCHA.csv', index = False, encoding="utf-8", sep = ';')
